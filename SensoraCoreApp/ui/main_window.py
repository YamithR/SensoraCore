# =====================================================================================
# ARCHIVO PRINCIPAL DE INTERFAZ GRÁFICA PARA SENSORACORE
# =====================================================================================
# Ruta del archivo: main_window.py para SensoraCore/ui
# Función: Define la ventana principal y todas las interfaces de los sensores
# Autor: Sistema SensoraCore
# Propósito: Crear una aplicación desktop para monitoreo de sensores ESP32
from IMPORTACIONES import *  # Importar todo lo necesario desde el módulo de importaciones
from Modulos.SENSORA_SIMPLE_ANGLE import (LinearCalibration, anguloSimple_UI, AnguloSimpleMonitor)


# =====================================================================================
# CLASE: HILO PARA SENSOR DE BRAZO CON MÚLTIPLES ÁNGULOS
# =====================================================================================
# Propósito: Maneja la comunicación para brazo robótico con 3 potenciómetros + sensor capacitivo
# Funcionalidad: Recibe datos de 3 ángulos simultáneamente más estado de sensor capacitivo
# Hereda de: QThread (permite ejecución en segundo plano)

class BrazoAnguloThread(QThread):
    # --- SEÑAL PERSONALIZADA COMPLEJA ---
    # Signal con 7 parámetros: 3 lecturas ADC, 3 ángulos, 1 estado de sensor capacitivo
    # (lectura1, angulo1, lectura2, angulo2, lectura3, angulo3, sensor_capacitivo_activo)
    data_received = Signal(int, int, int, int, int, int, bool)
    
    def __init__(self, esp32_ip, port=8080):
        """
        Constructor del hilo para brazo con múltiples sensores
        
        Parámetros:
        - esp32_ip: IP del ESP32
        - port: Puerto TCP (8080 por defecto)
        """
        super().__init__()                        # Inicializar QThread
        self.esp32_ip = esp32_ip                 # IP del microcontrolador
        self.port = port                         # Puerto de comunicación
        self.running = False                     # Control de bucle principal
        self.sock = None                         # Socket TCP
    
    def run(self):
        """
        Método principal - maneja comunicación con brazo multi-sensor
        Formato esperado: "POT1:val,ANG1:ang,POT2:val,ANG2:ang,POT3:val,ANG3:ang,SENSOR:state"
        """
        self.running = True                      # Activar bandera de ejecución
        try:
            # --- ESTABLECER CONEXIÓN ---
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  # Socket TCP
            self.sock.settimeout(3)              # 3 segundos para conectar
            self.sock.connect((self.esp32_ip, self.port))  # Conectar al ESP32
            
            # --- ACTIVAR MODO BRAZO ---
            self.sock.sendall(b'MODO:BRAZO_ANGULO')  # Comando para modo brazo multi-sensor
            self.sock.settimeout(1)              # 1 segundo para datos
            
            # --- BUCLE DE RECEPCIÓN DE DATOS ---
            while self.running:
                try:
                    # Recibir hasta 128 bytes (más datos que sensor simple)
                    data = self.sock.recv(128)
                    if not data:                 # Conexión cerrada
                        break
                    
                    # --- PROCESAR MENSAJE COMPLETO ---
                    msg = data.decode(errors='ignore').strip()  # Bytes a string
                    for line in msg.split('\n'):               # Cada línea por separado
                        if line.startswith('POT1:'):            # Identificar datos del brazo
                            try:
                                # --- PARSEAR DATOS COMPLEJOS ---
                                # Formato: POT1:val,ANG1:ang,POT2:val,ANG2:ang,POT3:val,ANG3:ang,SENSOR:state
                                parts = line.split(',')         # Separar por comas
                                
                                # Extraer valores de cada potenciómetro y ángulo
                                lectura1 = int(parts[0].split(':')[1])  # Lectura ADC potenciómetro 1
                                angulo1 = int(parts[1].split(':')[1])   # Ángulo calculado 1
                                lectura2 = int(parts[2].split(':')[1])  # Lectura ADC potenciómetro 2
                                angulo2 = int(parts[3].split(':')[1])   # Ángulo calculado 2
                                lectura3 = int(parts[4].split(':')[1])  # Lectura ADC potenciómetro 3
                                angulo3 = int(parts[5].split(':')[1])   # Ángulo calculado 3
                                
                                # Estado del sensor capacitivo (True/False)
                                sensor_estado = parts[6].split(':')[1] == 'True'
                                
                                # --- EMITIR TODOS LOS DATOS ---
                                self.data_received.emit(lectura1, angulo1, lectura2, angulo2, 
                                                      lectura3, angulo3, sensor_estado)
                            except:
                                pass                             # Ignorar errores de formato
                                
                except socket.timeout:                          # Timeout en recepción
                    continue                                    # Continuar esperando
                    
        except Exception as e:                                  # Error de conexión
            pass                                                # Ignorar y terminar
            
        finally:
            # --- LIMPIEZA DE RECURSOS ---
            if self.sock:
                try:
                    self.sock.sendall(b'STOP')                 # Detener modo brazo
                except:
                    pass
                self.sock.close()                              # Cerrar socket
    
    def stop(self):
        """Detener hilo del brazo de forma segura"""
        self.running = False                                   # Desactivar bucle
        self.wait()                                           # Esperar finalización

# =====================================================================================
# CLASE: HILO PARA SENSOR DE DISTANCIA INFRARROJO (IR)
# =====================================================================================
# Propósito: Maneja sensor de distancia por infrarrojos con salida digital
# Funcionalidad: Detecta presencia/ausencia de objetos (True/False)
# Tipo de sensor: Sensor IR digital de proximidad

class DistanciaIRThread(QThread):
    # --- SEÑAL SIMPLE DIGITAL ---
    # Signal(bool) = True cuando detecta objeto, False cuando no hay objeto
    data_received = Signal(bool)  # Solo estado digital ON/OFF
    
    def __init__(self, esp32_ip, port=8080):
        """
        Constructor para hilo de sensor IR
        
        Parámetros:
        - esp32_ip: Dirección IP del ESP32
        - port: Puerto TCP (8080 predeterminado)
        """
        super().__init__()                        # Inicializar clase padre QThread
        self.esp32_ip = esp32_ip                 # Almacenar IP del microcontrolador
        self.port = port                         # Puerto de comunicación TCP
        self.running = False                     # Flag de control de bucle
        self.sock = None                         # Socket de conexión TCP
    
    def run(self):
        """
        Método principal del hilo - maneja comunicación con sensor IR
        Formato esperado del ESP32: "IR_DIGITAL:True" o "IR_DIGITAL:False"
        """
        self.running = True                      # Activar bandera de ejecución
        try:
            # --- ESTABLECER COMUNICACIÓN TCP ---
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  # Crear socket
            self.sock.settimeout(3)              # Timeout de conexión: 3 segundos
            self.sock.connect((self.esp32_ip, self.port))  # Conectar al ESP32
            
            # --- CONFIGURAR MODO SENSOR IR ---
            self.sock.sendall(b'MODO:DISTANCIA_IR')  # Enviar comando de activación
            self.sock.settimeout(1)              # Timeout de recepción: 1 segundo
            
            # --- BUCLE PRINCIPAL DE MONITOREO ---
            while self.running:                  # Ejecutar mientras esté activo
                try:
                    # Recibir datos del sensor (hasta 128 bytes)
                    data = self.sock.recv(128)
                    if not data:                 # Si no hay datos, conexión perdida
                        break
                    
                    # --- PROCESAR DATOS DIGITALES ---
                    msg = data.decode(errors='ignore').strip()  # Convertir a string
                    for line in msg.split('\n'):               # Procesar línea por línea
                        if line.startswith('IR_DIGITAL:'):      # Buscar datos del sensor IR
                            try:
                                # --- PARSEAR ESTADO DIGITAL ---
                                # Formato: "IR_DIGITAL:True" o "IR_DIGITAL:False"
                                estado = line.split(':')[1] == 'True'  # Convertir string a booleano
                                
                                # --- EMITIR ESTADO ---
                                self.data_received.emit(estado)        # Enviar estado a interfaz
                            except:
                                pass                             # Ignorar errores de formato
                                
                except socket.timeout:                          # Manejo de timeout
                    continue                                    # Continuar esperando datos
                    
        except Exception as e:                                  # Capturar errores de conexión
            pass                                                # Ignorar y finalizar
            
        finally:
            # --- LIMPIEZA DE RECURSOS ---
            if self.sock:                                       # Si socket está activo
                try:
                    self.sock.sendall(b'STOP')                 # Enviar comando de parada
                except:
                    pass                                        # Ignorar errores de envío
                self.sock.close()                              # Cerrar conexión TCP
    
    def stop(self):
        """Método para detener el hilo de sensor IR de forma segura"""
        self.running = False                                   # Desactivar bucle principal
        self.wait()                                           # Esperar finalización del hilo

# =====================================================================================
# CLASE: HILO PARA SENSOR DE DISTANCIA CAPACITIVO
# =====================================================================================
# Propósito: Maneja sensor capacitivo para detección de proximidad
# Funcionalidad: Detecta objetos cercanos usando cambios en capacitancia
# Salida: Digital (True = objeto detectado, False = sin objeto)

class DistanciaCapThread(QThread):
    # --- SEÑAL DIGITAL CAPACITIVA ---
    # Signal(bool) para estado del sensor capacitivo
    data_received = Signal(bool)  # Solo estado digital ON/OFF
    
    def __init__(self, esp32_ip, port=8080):
        """
        Constructor del hilo para sensor capacitivo
        
        Parámetros:
        - esp32_ip: IP del ESP32 conectado
        - port: Puerto TCP para comunicación
        """
        super().__init__()                        # Inicializar QThread padre
        self.esp32_ip = esp32_ip                 # Dirección IP del microcontrolador
        self.port = port                         # Puerto de conexión TCP
        self.running = False                     # Control del bucle principal
        self.sock = None                         # Socket de comunicación
    
    def run(self):
        """
        Bucle principal del hilo capacitivo
        Formato esperado: "CAP_DIGITAL:True" o "CAP_DIGITAL:False"
        """
        self.running = True                      # Activar ejecución del hilo
        try:
            # --- ESTABLECER CONEXIÓN TCP ---
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  # Socket TCP
            self.sock.settimeout(3)              # 3 segundos para establecer conexión
            self.sock.connect((self.esp32_ip, self.port))  # Conectar al ESP32
            
            # --- ACTIVAR MODO CAPACITIVO ---
            self.sock.sendall(b'MODO:DISTANCIA_CAP')  # Comando para activar sensor capacitivo
            self.sock.settimeout(1)              # 1 segundo para recibir datos
            
            # --- BUCLE DE MONITOREO ---
            while self.running:                  # Ejecutar mientras esté activo
                try:
                    # Recibir datos del sensor capacitivo
                    data = self.sock.recv(128)
                    if not data:                 # Verificar si hay datos
                        break
                    
                    # --- DECODIFICAR MENSAJE ---
                    msg = data.decode(errors='ignore').strip()  # Bytes a string limpio
                    for line in msg.split('\n'):               # Cada línea independiente
                        if line.startswith('CAP_DIGITAL:'):     # Identificar datos capacitivos
                            try:
                                # --- EXTRAER ESTADO DIGITAL ---
                                # Parsear formato: "CAP_DIGITAL:True" o "CAP_DIGITAL:False"
                                estado = line.split(':')[1] == 'True'  # String a booleano
                                
                                # --- TRANSMITIR ESTADO ---
                                self.data_received.emit(estado)        # Enviar a interfaz principal
                            except:
                                pass                             # Ignorar errores de conversión
                                
                except socket.timeout:                          # Timeout en recepción
                    continue                                    # Continuar monitoreando
                except Exception as e:                          # Otros errores de comunicación
                    pass                                        # Ignorar y continuar
                    
        finally:
            # --- FINALIZACIÓN Y LIMPIEZA ---
            if self.sock:                                       # Si socket existe
                try:
                    self.sock.sendall(b'STOP')                 # Comando de parada al ESP32
                except:
                    pass                                        # Ignorar errores de envío
                self.sock.close()                              # Cerrar conexión TCP
    
    def stop(self):
        """Detener hilo capacitivo de manera controlada"""
        self.running = False                                   # Desactivar bucle
        self.wait()                                           # Esperar finalización completa

# =====================================================================================
# CLASE: HILO PARA SENSOR ULTRASÓNICO DE DISTANCIA
# =====================================================================================
# Propósito: Maneja sensor ultrasónico HC-SR04 para medición de distancia
# Funcionalidad: Mide distancia real en centímetros usando ondas ultrasónicas
# Datos: Solo distancia calculada (no ADC/voltaje simulados)

class DistanciaUltrasonicThread(QThread):
    # --- SEÑAL SIMPLIFICADA ---
    # Signal(float) = distancia_en_cm
    data_received = Signal(float)  # solo distancia_cm
    
    def __init__(self, esp32_ip, port=8080):
        """
        Constructor para hilo de sensor ultrasónico
        
        Parámetros:
        - esp32_ip: Dirección IP del ESP32
        - port: Puerto TCP (8080 por defecto)
        """
        super().__init__()                        # Inicializar clase padre QThread
        self.esp32_ip = esp32_ip                 # IP del microcontrolador ESP32
        self.port = port                         # Puerto de comunicación TCP        self.running = False                     # Flag de control del bucle
        self.sock = None                         # Socket de conexión TCP
    
    def run(self):
        """
        Método principal - maneja sensor ultrasónico HC-SR04
        Formato esperado: "ULTRA_CM:25.4"
        """
        self.running = True                      # Activar bandera de ejecución
        try:
            # --- ESTABLECER COMUNICACIÓN ---
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  # Socket TCP
            self.sock.settimeout(3)              # Timeout de conexión: 3 segundos
            self.sock.connect((self.esp32_ip, self.port))  # Conectar al ESP32
            
            # --- CONFIGURAR MODO ULTRASÓNICO ---
            self.sock.sendall(b'MODO:DISTANCIA_ULTRA')  # Activar modo sensor ultrasónico
            self.sock.settimeout(1)              # Timeout para recepción: 1 segundo
            
            # --- BUCLE DE MEDICIÓN CONTINUA ---
            while self.running:                  # Ejecutar mientras esté activo
                try:
                    # Recibir datos del sensor ultrasónico
                    data = self.sock.recv(128)
                    if not data:                 # Verificar conexión activa
                        break
                    
                    # --- PROCESAR DATOS DE DISTANCIA ---
                    msg = data.decode(errors='ignore').strip()  # Convertir a string
                    for line in msg.split('\n'):               # Procesar cada línea
                        if line.startswith('ULTRA_CM:'):        # Identificar datos ultrasónicos
                            try:
                                # --- PARSEAR SOLO DISTANCIA ---
                                # Formato: "ULTRA_CM:25.4"
                                distancia_cm = float(line.split(':')[1])  # Distancia en centímetros
                                
                                # --- EMITIR SOLO DISTANCIA REAL ---
                                self.data_received.emit(distancia_cm)
                            except:
                                pass                             # Ignorar errores de conversión
                                
                except socket.timeout:                          # Manejo de timeout
                    continue                                    # Continuar esperando datos
                    
        except Exception as e:                                  # Capturar errores de comunicación
            pass                                                # Ignorar y finalizar hilo
            
        finally:
            # --- LIMPIEZA DE RECURSOS ---
            if self.sock:                                       # Si socket está activo
                try:
                    self.sock.sendall(b'STOP')                 # Detener modo ultrasónico
                except:
                    pass                                        # Ignorar errores de envío
                self.sock.close()                              # Cerrar conexión TCP
    
    def stop(self):
        """Detener hilo ultrasónico de forma segura"""
        self.running = False                                   # Desactivar bucle principal
        self.wait()                                           # Esperar finalización del hilo

# =====================================================================================
# CLASE PRINCIPAL: VENTANA PRINCIPAL DE LA APLICACIÓN SENSORACORE
# =====================================================================================
# Propósito: Ventana principal que contiene toda la interfaz de usuario
# Funcionalidad: Gestiona conexión ESP32, selección de sensores, y visualización de datos
# Hereda de: QMainWindow (ventana principal de Qt con menús, barras de herramientas, etc.)

class MainWindow(QMainWindow, AnguloSimpleMonitor):
    def __init__(self):
        """
        Constructor de la ventana principal
        Inicializa todos los componentes de la interfaz y variables de estado
        """
        super().__init__()                       # Inicializar la clase padre QMainWindow
        
        # --- CONFIGURACIÓN BÁSICA DE VENTANA ---
        self.setWindowTitle("SensoraCore")       # Título que aparece en la barra de título
        self.setMinimumSize(1000, 700)          # Tamaño mínimo permitido (ancho x alto)
        self.resize(1200, 800)                  # Tamaño inicial de la ventana
        
        # =====================================================================================
        # VARIABLES DE ESTADO PARA CONEXIÓN Y HILOS
        # =====================================================================================
        
        # --- Variables para cliente ESP32 y hilos de sensores ---
        self.esp_client = None                   # Cliente para comunicación básica con ESP32
        self.angulo_thread = None               # Hilo para sensor de ángulo simple
        self.brazo_thread = None                # Hilo para brazo con múltiples sensores
        self.distancia_ir_thread = None         # Hilo para sensor infrarrojo
        self.distancia_cap_thread = None        # Hilo para sensor capacitivo
          # --- Banderas de estado general ---
        self.is_connected = False               # True cuando ESP32 está conectado
        self.is_monitoring = False              # True cuando algún sensor está monitoreando

       
        # =====================================================================================
        # SISTEMA DE CALIBRACIÓN
        # =====================================================================================
        
        # --- Instancia de calibración para sensor de ángulo simple ---
        self.angulo_calibration = LinearCalibration()  # Sistema de calibración lineal
        # =====================================================================================
        # VARIABLES PARA DATOS DE SENSOR DE ÁNGULO SIMPLE (con optimización de memoria)
        # =====================================================================================
        
        self.angulos = []                       # Lista para almacenar ángulos medidos
        self.lecturas = []                      # Lista para almacenar lecturas ADC
        self.max_points = 100                   # Límite máximo de puntos en gráfica (evita lag)
        
        # --- Timer para optimizar actualizaciones de gráfica ---
        self.graph_update_timer = QTimer()      # Timer para actualizar gráficas
        self.graph_update_timer.timeout.connect(self.update_graph_display)  # Conectar a método de actualización
        self.graph_update_timer.setInterval(100)  # Actualizar cada 100ms (10 FPS)

        # =====================================================================================
        # VARIABLES PARA DATOS DE BRAZO MULTI-SENSOR (con optimización)
        # =====================================================================================
        
        self.brazo_angulos = [[], [], []]       # Listas para 3 ángulos del brazo [ang1, ang2, ang3]
        self.brazo_lecturas = [[], [], []]      # Listas para 3 lecturas ADC [lec1, lec2, lec3]
        self.brazo_capacitive_states = []       # Lista para estados del sensor capacitivo
        self.brazo_max_points = 100              # Límite reducido para mejor rendimiento
        self.brazo_is_monitoring = False        # Flag específico para monitoreo del brazo
        
        # =====================================================================================
        # VARIABLES PARA SENSORES DE DISTANCIA
        # =====================================================================================
        
        # --- Variables para sensor IR ---
        self.distancia_ir_lecturas = []         # Lecturas del sensor IR
        self.distancia_ir_voltajes = []         # Voltajes del sensor IR
        self.distancia_ir_cm = []               # Distancias calculadas del IR
          # --- Variables para sensor capacitivo ---
        self.distancia_cap_lecturas = []        # Lecturas del sensor capacitivo
        self.distancia_cap_voltajes = []        # Voltajes del sensor capacitivo
        self.distancia_cap_cm = []              # Distancias del capacitivo
        
        # --- Variables para sensor ultrasónico ---
        self.distancia_ultra_cm = []            # Distancias reales del ultrasónico
        
        # --- Configuración común para sensores de distancia ---
        self.distancia_max_points = 100         # Límite de puntos para sensores de distancia
        self.distancia_ir_is_monitoring = False # Flag para monitoreo IR
        self.distancia_cap_is_monitoring = False # Flag para monitoreo capacitivo
        self.distancia_ultra_is_monitoring = False # Flag para monitoreo ultrasónico
        
        # =====================================================================================
        # SISTEMA DE ACTUALIZACIONES OPTIMIZADO
        # =====================================================================================
        
        # --- Variables para optimizar actualizaciones de interfaz ---
        self.pending_updates = False            # Flag para evitar actualizaciones excesivas
        self.pending_simple_data = None         # Datos pendientes del sensor simple
        self.pending_brazo_data = None          # Datos pendientes del brazo
        self.pending_distancia_ir_data = None   # Datos pendientes del IR
        self.pending_distancia_cap_data = None  # Datos pendientes del capacitivo
        self.pending_distancia_ultra_data = None # Datos pendientes del ultrasónico
        
        # =====================================================================================
        # INICIALIZACIÓN DE LA INTERFAZ
        # =====================================================================================
        
        # Configurar estilo visual de la aplicación
        self.setup_styles()
          # Crear todos los elementos de la interfaz de usuario
        self.setup_ui()

    # =====================================================================================
    # MÉTODO AUXILIAR: GESTIÓN DEL TIMER DE GRÁFICAS
    # =====================================================================================
    
    def manage_graph_timer(self):
        """
        Gestiona el timer de actualización de gráficas basado en el estado de todos los sensores
        
        Propósito: Evitar conflictos entre múltiples sensores que comparten el mismo timer
        Lógica: Solo detiene el timer si TODOS los sensores están inactivos
        Sensores: Ángulo simple, brazo robótico, IR, capacitivo, ultrasónico
        Optimización: Evita detener el timer prematuramente causando gráficas estáticas
        """
        
        # --- VERIFICAR ESTADO DE TODOS LOS SENSORES ---
        sensors_active = (
            self.is_monitoring or                       # Sensor de ángulo simple
            self.brazo_is_monitoring or                 # Brazo robótico multi-sensor
            self.distancia_ir_is_monitoring or          # Sensor infrarrojo
            self.distancia_cap_is_monitoring or         # Sensor capacitivo
            self.distancia_ultra_is_monitoring          # Sensor ultrasónico
        )
        
        # --- GESTIONAR TIMER SEGÚN ESTADO GLOBAL ---
        if sensors_active:
            # Si hay al menos un sensor activo, asegurar que el timer esté corriendo
            if not self.graph_update_timer.isActive():
                self.graph_update_timer.start()
        else:
            # Solo detener timer si TODOS los sensores están inactivos
            if self.graph_update_timer.isActive():
                self.graph_update_timer.stop()

      # =====================================================================================
    # MÉTODO: CONFIGURACIÓN DE ESTILOS VISUALES
    # =====================================================================================
    def setup_styles(self):
        """
        Configura los estilos globales de la aplicación usando CSS-like syntax
        
        Propósito: Define la apariencia visual de todos los elementos de la interfaz
        Tecnología: QSS (Qt Style Sheets) - similar a CSS para web
        Resultado: Interfaz moderna y profesional con colores consistentes
        """
        self.setStyleSheet("""
            /* ======================== VENTANA PRINCIPAL ======================== */
            /* Configuración del fondo general de toda la aplicación */
            QMainWindow {
                background-color: #f8f9fa;  /* Color de fondo: gris muy claro (#f8f9fa) */
            }
            
            /* ======================== CAJAS DE GRUPO ======================== */
            /* Estilo para todas las cajas de grupo (secciones) de la aplicación */
            /* QGroupBox se usa para agrupar controles relacionados con un borde y título */
            QGroupBox {
                font-weight: bold;              /* Texto en negrita para destacar títulos */
                border: 2px solid #dee2e6;     /* Borde sólido gris claro de 2px de grosor */
                border-radius: 8px;            /* Esquinas redondeadas (8px de radio) */
                margin-top: 1ex;               /* Margen superior para acomodar el título */
                padding-top: 10px;             /* Espacio interno superior (10px) */
                background-color: white;       /* Fondo blanco para contraste */
            }
            
            /* Estilo específico para los títulos de las cajas de grupo */
            QGroupBox::title {
                subcontrol-origin: margin;     /* El título se origina desde el margen */
                left: 10px;                    /* Posición izquierda del título (10px desde borde) */
                padding: 0 8px 0 8px;         /* Padding: arriba=0, derecha=8px, abajo=0, izquierda=8px */
                color: #495057;               /* Color del texto: gris oscuro profesional */
                background-color: white;       /* Fondo blanco para que se vea sobre el borde */
            }
            
            /* ======================== BOTONES ESTÁNDAR ======================== */
            /* Estilo base para todos los botones de la aplicación */
            QPushButton {
                border: 2px solid #007bff;     /* Borde azul Bootstrap de 2px */
                border-radius: 6px;            /* Esquinas redondeadas moderadas */
                padding: 8px 16px;             /* Padding interno: 8px vertical, 16px horizontal */
                background-color: #007bff;     /* Fondo azul primario Bootstrap */
                color: black;                  /* Texto negro para máximo contraste */
                font-weight: bold;             /* Texto en negrita para legibilidad */
                min-height: 20px;             /* Altura mínima para botones uniformes */
            }
            
            /* Estilo cuando el mouse pasa por encima del botón (hover effect) */
            QPushButton:hover {
                background-color: #0056b3;     /* Azul más oscuro al hacer hover */
                border-color: #0056b3;         /* Borde también más oscuro */
            }
            
            /* Estilo cuando el botón está siendo presionado (pressed state) */
            QPushButton:pressed {
                background-color: #004085;     /* Azul muy oscuro para feedback visual */
            }
            
            /* Estilo cuando el botón está deshabilitado (disabled state) */
            QPushButton:disabled {
                background-color: #6c757d;     /* Gris para indicar estado inactivo */
                border-color: #6c757d;         /* Borde gris consistente */
                color: #adb5bd;               /* Texto gris claro para menor contraste */
            }
            
            /* ======================== CAMPOS DE TEXTO ======================== */
            /* Estilo para todos los campos de entrada de texto (QLineEdit) */
            QLineEdit {
                border: 2px solid #ced4da;     /* Borde gris neutro */
                border-radius: 6px;            /* Esquinas redondeadas suaves */
                padding: 8px 12px;             /* Padding: 8px vertical, 12px horizontal */
                font-size: 14px;              /* Tamaño de fuente legible */
                background-color: white;       /* Fondo blanco limpio */
                color: black;                  /* Texto negro para contraste */
            }
            
            /* Estilo cuando el campo de texto tiene foco (usuario está escribiendo) */
            QLineEdit:focus {
                border-color: #007bff;         /* Borde azul para indicar campo activo */
                outline: none;                 /* Remover outline por defecto del sistema */
            }
            
            /* ======================== LISTAS DE SENSORES ======================== */
            /* Estilo para la lista de sensores disponibles (QListWidget) */
            QListWidget {
                border: 1px solid #dee2e6;     /* Borde sutil gris claro */
                border-radius: 6px;            /* Esquinas redondeadas */
                background-color: white;       /* Fondo blanco principal */
                alternate-background-color: #f8f9fa;  /* Color alternativo para filas zebra */
            }
            
            /* Estilo para cada elemento individual de la lista de sensores */
            QListWidget::item {
                padding: 12px;                 /* Espacio interno generoso para legibilidad */
                border-bottom: 1px solid #e9ecef;  /* Línea separadora sutil entre elementos */
                color: black;                  /* Texto negro */
            }
            
            /* Estilo cuando el mouse pasa por encima de un elemento de la lista */
            QListWidget::item:hover {
                background-color: #7db9f9;     /* Azul claro suave para hover */
            }
            
            /* Estilo cuando un elemento de la lista está seleccionado */
            QListWidget::item:selected {
                background-color: #007bff;     /* Azul primario para selección */
                color: black;                  /* Mantener texto negro para contraste */
            }
        """)
    # =====================================================================================
    # MÉTODO: CONFIGURACIÓN DE LA INTERFAZ DE USUARIO
    # =====================================================================================
    def setup_ui(self):
        """
        Configura la estructura principal de la interfaz de usuario
        
        Propósito: Crear el layout principal con panel izquierdo y derecho
        Diseño: Interfaz dividida (split) con conexión/sensores a la izquierda y detalles a la derecha
        Tecnología: QSplitter para división ajustable entre paneles
        """
        
        # --- WIDGET CENTRAL PRINCIPAL ---
        # Toda aplicación QMainWindow necesita un widget central
        central_widget = QWidget()               # Crear widget contenedor principal
        self.setCentralWidget(central_widget)    # Establecer como widget central de la ventana
        
        # --- LAYOUT PRINCIPAL HORIZONTAL ---
        # QHBoxLayout organiza elementos horizontalmente (lado a lado)
        main_layout = QHBoxLayout(central_widget)  # Aplicar layout al widget central
        main_layout.setSpacing(20)               # Espacio entre elementos: 20 píxeles
        main_layout.setContentsMargins(20, 20, 20, 20)  # Márgenes: 20px en todos los lados
        
        # --- CREAR SPLITTER PARA DIVISIÓN RESPONSIVE ---
        # QSplitter permite al usuario ajustar el tamaño de los paneles arrastrando
        splitter = QSplitter(Qt.Horizontal)      # Divisor horizontal (izquierda-derecha)
        main_layout.addWidget(splitter)          # Agregar splitter al layout principal
        
        # --- PANEL IZQUIERDO (conexión y lista de sensores) ---
        self.left_panel = self.create_left_panel()  # Crear panel izquierdo (método separado)
        splitter.addWidget(self.left_panel)     # Agregar al splitter
        
        # --- PANEL DERECHO (detalles del sensor seleccionado) ---
        self.right_panel = self.create_right_panel()  # Crear panel derecho (método separado)
        splitter.addWidget(self.right_panel)    # Agregar al splitter
        
        # --- CONFIGURAR PROPORCIONES DEL SPLITTER ---
        # setStretchFactor define qué tanto espacio ocupa cada panel relativo al otro
        splitter.setStretchFactor(0, 1)         # Panel izquierdo: factor 1 (1/3 del espacio)
        splitter.setStretchFactor(1, 2)         # Panel derecho: factor 2 (2/3 del espacio)
        splitter.setSizes([400, 800])           # Tamaños iniciales en píxeles [izquierdo, derecho]    
    # =====================================================================================
    # MÉTODO: CREACIÓN DEL PANEL IZQUIERDO
    # =====================================================================================
    def create_left_panel(self):
        """
        Crea el panel izquierdo con conexión ESP32 y lista de sensores
        
        Propósito: Interface para conectar al ESP32 y seleccionar sensores disponibles
        Contenido: Título, configuración de conexión, estado, y lista de sensores
        Retorna: QFrame configurado con todos los elementos del panel izquierdo
        """
        
        # --- CONTENEDOR PRINCIPAL DEL PANEL ---
        panel = QFrame()                         # Crear frame contenedor
        panel.setFrameStyle(QFrame.StyledPanel)  # Estilo de panel con borde
        panel.setStyleSheet("QFrame { background-color: white; border-radius: 8px; }")  # Fondo blanco redondeado
        
        # --- LAYOUT VERTICAL DEL PANEL ---
        layout = QVBoxLayout(panel)              # Layout vertical (elementos apilados)
        layout.setSpacing(20)                    # Espacio entre elementos: 20px
        layout.setContentsMargins(20, 20, 20, 20)  # Márgenes internos: 20px todos los lados
        
        # =====================================================================================
        # SECCIÓN: TÍTULO Y SUBTÍTULO DE BIENVENIDA
        # =====================================================================================
        
        # --- TÍTULO PRINCIPAL ---
        title_label = QLabel("SensoraCore")      # Crear etiqueta con nombre de la aplicación
        title_label.setAlignment(Qt.AlignCenter) # Centrar texto horizontalmente
        title_label.setStyleSheet("""
            font-size: 24px;                    /* Tamaño de fuente grande */
            font-weight: bold;                  /* Texto en negrita */
            color: #007bff;                     /* Color azul corporativo */
            margin: 10px 0;                     /* Margen vertical de 10px */
        """)
        layout.addWidget(title_label)           # Agregar título al layout
        
        # --- SUBTÍTULO DESCRIPTIVO ---
        subtitle_label = QLabel("Sistema de Monitoreo de Sensores WiFi")  # Descripción del sistema
        subtitle_label.setAlignment(Qt.AlignCenter)  # Centrar texto
        subtitle_label.setStyleSheet("""
            font-size: 14px;                    /* Tamaño mediano */
            color: #6c757d;                     /* Color gris suave */
            margin-bottom: 20px;                /* Margen inferior para separación */
        """)
        layout.addWidget(subtitle_label)       # Agregar subtítulo al layout
        
        # =====================================================================================
        # SECCIÓN: CONFIGURACIÓN DE CONEXIÓN ESP32
        # =====================================================================================
        
        # --- GRUPO DE CONEXIÓN ---
        connection_group = QGroupBox("Configuración de Conexión")  # Caja agrupada con título
        connection_layout = QVBoxLayout(connection_group)  # Layout vertical para la caja
        
        # --- ETIQUETA PARA CAMPO IP ---
        ip_label = QLabel("IP del ESP32:")       # Etiqueta descriptiva
        ip_label.setStyleSheet("font-weight: bold; color: #495057;")  # Negrita y color oscuro
        connection_layout.addWidget(ip_label)   # Agregar etiqueta al grupo
        
        # --- CAMPO DE ENTRADA DE IP ---
        self.ip_input = QLineEdit()              # Campo de texto para ingresar IP
        self.ip_input.setPlaceholderText("Ejemplo: 192.168.1.100")  # Texto de ayuda
        self.ip_input.setText("192.168.20.25")  # IP predeterminada (cambiar según red)
        connection_layout.addWidget(self.ip_input)  # Agregar campo al grupo
        
        # --- BOTÓN DE CONEXIÓN ---
        self.connect_btn = QPushButton("🔌 Conectar ESP32")  # Botón con emoji para visual
        self.connect_btn.clicked.connect(self.test_connection)  # Conectar señal click al método
        connection_layout.addWidget(self.connect_btn)  # Agregar botón al grupo
        
        # --- INDICADOR DE ESTADO DE CONEXIÓN ---
        self.status_label = QLabel("⚪ Desconectado")  # Etiqueta de estado inicial
        self.status_label.setAlignment(Qt.AlignCenter)  # Centrar texto
        self.status_label.setStyleSheet("""
            padding: 8px;                       /* Espacio interno */
            border-radius: 4px;                 /* Esquinas redondeadas */
            background-color: #f8d7da;          /* Fondo rojo claro (desconectado) */
            color: #721c24;                     /* Texto rojo oscuro */
            font-weight: bold;                  /* Texto en negrita */
        """)
        connection_layout.addWidget(self.status_label)  # Agregar indicador al grupo
        
        layout.addWidget(connection_group)     # Agregar grupo completo al panel
        
        # =====================================================================================
        # SECCIÓN: LISTA DE SENSORES DISPONIBLES
        # =====================================================================================
        
        # --- GRUPO DE SENSORES (inicialmente oculto) ---
        self.sensors_group = QGroupBox("Sensores Disponibles")  # Caja para lista de sensores
        self.sensors_group.setVisible(False)   # Ocultar hasta que se conecte ESP32
        
        sensors_layout = QVBoxLayout(self.sensors_group)  # Layout para el grupo de sensores
        
        # --- LISTA DE SENSORES ---
        self.sensors_list = QListWidget()       # Widget de lista para sensores
        self.sensors_list.itemClicked.connect(self.on_sensor_selected)  # Conectar evento de selección
        
        # --- POBLAR LISTA CON SENSORES DISPONIBLES ---
        self.add_sensor_items()                 # Método que agrega los sensores a la lista
        
        sensors_layout.addWidget(self.sensors_list)  # Agregar lista al grupo
        layout.addWidget(self.sensors_group)   # Agregar grupo al panel principal
          # =====================================================================================
        # SECCIÓN: BOTÓN DE REINICIO EN ESQUINA SUPERIOR IZQUIERDA
        # =====================================================================================
        
        # --- BOTÓN DE REINICIO ---
        self.restart_btn = QPushButton("🔄 Reiniciar Interfaz")  # Botón con emoji de reinicio
        self.restart_btn.setMinimumHeight(40)    # Altura mínima para mayor visibilidad
        self.restart_btn.setMaximumWidth(180)    # Ancho máximo controlado
        self.restart_btn.clicked.connect(self.restart_application)  # Conectar al método de reinicio
        self.restart_btn.setEnabled(False)      # Deshabilitado hasta conectar ESP32
        self.restart_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;     /* Verde para acción positiva */
                color: white;                   /* Texto blanco */
                border: none;                   /* Sin borde */
                border-radius: 8px;             /* Esquinas redondeadas */
                font-weight: bold;              /* Texto en negrita */
                font-size: 12px;                /* Tamaño de fuente legible */
                padding: 8px 12px;              /* Espaciado interno */
            }
            QPushButton:hover {
                background-color: #218838;     /* Verde más oscuro al pasar mouse */
            }
            QPushButton:pressed {
                background-color: #1e7e34;     /* Verde aún más oscuro al presionar */
            }
            QPushButton:disabled {
                background-color: #6c757d;     /* Gris cuando está deshabilitado */
                color: #dee2e6;                 /* Texto gris claro */
            }
        """)
        layout.addWidget(self.restart_btn)       # Agregar botón al panel principal
        
        # --- ESPACIADOR FLEXIBLE ---
        layout.addStretch()                     # Agregar espacio flexible al final
        
        return panel                            # Retornar panel completo configurado
    # =====================================================================================
    # MÉTODO: AGREGAR ELEMENTOS A LA LISTA DE SENSORES
    # =====================================================================================
    def add_sensor_items(self):
        """
        Agrega los sensores disponibles a la lista del panel izquierdo
        
        Propósito: Poblar la lista con todos los tipos de sensores soportados
        Funcionalidad: Crear elementos de lista con descripción, emoji y estado de disponibilidad
        Datos: Cada sensor tiene emoji, descripción y ID único para identificación
        """
        
        # --- DEFINIR SENSORES DISPONIBLES ---
        # Tupla con: (emoji_nombre, descripción_funcional, identificador_único)
        sensors = [
            ("🎛️ Ángulo Simple", "Potenciómetro como sensor de ángulo", "angulo_simple"),
            ("🦾 Brazo Ángulo", "Sensor de ángulo para brazo robótico", "brazo_angulo"),
            ("📏 Distancia IR", "Sensor de distancia infrarrojo", "distancia_ir"),
            ("🔍 Distancia Capacitivo", "Sensor de distancia capacitivo", "distancia_cap"),
            ("📡 Distancia Ultrasónico", "Sensor HC-SR04", "distancia_ultra"),
            ("💨 Velocidad Óptica", "Sensor óptico de velocidad", "velocidad_optica")
        ]
        
        # --- CREAR ELEMENTOS DE LISTA PARA CADA SENSOR ---
        for icon_name, description, sensor_id in sensors:  # Iterar por cada sensor definido
            # Crear elemento individual de la lista
            item = QListWidgetItem()             # Nuevo elemento de lista
            item.setText(f"{icon_name}\n{description}")  # Texto: emoji + descripción en 2 líneas
            item.setData(Qt.UserRole, sensor_id) # Guardar ID único en datos del elemento
            
            # --- VERIFICAR DISPONIBILIDAD DEL SENSOR ---
            # Lista de sensores actualmente implementados y funcionales
            if sensor_id not in ["angulo_simple", "brazo_angulo", "distancia_ir", "distancia_cap", "distancia_ultra"]:
                # Para sensores no implementados:
                item.setFlags(item.flags() & ~Qt.ItemIsEnabled)  # Deshabilitar elemento
                item.setText(f"{icon_name}\n{description}\n(Próximamente)")  # Agregar nota de estado
                item.setToolTip("Esta función será implementada en futuras versiones")  # Tooltip informativo
            
            # --- AGREGAR ELEMENTO A LA LISTA ---
            self.sensors_list.addItem(item)      # Añadir elemento configurado a la lista    
    # =====================================================================================
    # MÉTODO: CREACIÓN DEL PANEL DERECHO
    # =====================================================================================
    def create_right_panel(self):
        """
        Crea el panel derecho para mostrar detalles del sensor seleccionado
        
        Propósito: Área principal donde se muestran gráficas, datos y controles del sensor activo
        Estados: Pantalla de bienvenida inicial → Interfaz específica del sensor seleccionado
        Contenido: Mensaje de bienvenida, instrucciones, y espacio para interfaces de sensores
        """
        
        # --- CONTENEDOR PRINCIPAL DEL PANEL DERECHO ---
        panel = QFrame()                         # Frame contenedor principal
        panel.setFrameStyle(QFrame.StyledPanel)  # Estilo de panel con borde
        panel.setStyleSheet("QFrame { background-color: white; border-radius: 8px; }")  # Fondo blanco redondeado
        
        # --- LAYOUT VERTICAL DEL PANEL ---
        layout = QVBoxLayout(panel)              # Layout vertical para apilar elementos
        layout.setContentsMargins(20, 20, 20, 20)  # Márgenes internos de 20px
        
        # =====================================================================================
        # SECCIÓN: MENSAJE DE BIENVENIDA INICIAL
        # =====================================================================================
        
        # --- WIDGET DE BIENVENIDA ---
        self.welcome_widget = QWidget()          # Widget contenedor para la pantalla inicial
        welcome_layout = QVBoxLayout(self.welcome_widget)  # Layout vertical para el contenido
        
        # --- ICONO PRINCIPAL ---
        welcome_icon = QLabel("🔧")              # Emoji de herramienta como icono principal
        welcome_icon.setAlignment(Qt.AlignCenter)  # Centrar icono horizontalmente
        welcome_icon.setStyleSheet("font-size: 72px; margin: 50px;")  # Tamaño grande con margen
        welcome_layout.addWidget(welcome_icon)   # Agregar icono al layout de bienvenida
        
        # --- TEXTO INSTRUCTIVO PRINCIPAL ---
        welcome_text = QLabel("Conecta tu ESP32 y selecciona un sensor\npara comenzar a monitorear datos")
        welcome_text.setAlignment(Qt.AlignCenter)  # Centrar texto
        welcome_text.setStyleSheet("""
            font-size: 16px;                    /* Tamaño de fuente legible */
            color: #6c757d;                     /* Color gris suave */
            line-height: 1.5;                   /* Espaciado entre líneas */
        """)
        welcome_layout.addWidget(welcome_text)  # Agregar texto al layout
        
        # --- INDICACIÓN SOBRE DIAGRAMAS DE CONEXIÓN ---
        diagram_hint = QLabel("📋 Una vez conectado, encontrarás el diagrama\nde conexiones ESP32 en cada sensor")
        diagram_hint.setAlignment(Qt.AlignCenter)  # Centrar texto
        diagram_hint.setStyleSheet("""
            font-size: 14px;                    /* Tamaño menor para texto secundario */
            color: #495057;                     /* Color más oscuro para destacar */
            background-color: #e2f4ff;          /* Fondo azul muy claro */
            border: 1px solid #b3daff;          /* Borde azul claro */
            border-radius: 6px;                 /* Esquinas redondeadas */
            padding: 10px;                      /* Espacio interno */
            margin: 10px 40px;                  /* Margen vertical y horizontal */
        """)
        welcome_layout.addWidget(diagram_hint)  # Agregar hint al layout
        
        # --- ESPACIADOR FLEXIBLE ---
        welcome_layout.addStretch()            # Espacio flexible para centrar contenido verticalmente
        
        layout.addWidget(self.welcome_widget)  # Agregar widget de bienvenida al panel principal
        
        # =====================================================================================
        # SECCIÓN: CONTENEDOR PARA DETALLES DEL SENSOR
        # =====================================================================================
        
        # --- ÁREA DE SCROLL PARA INTERFACES DE SENSORES ---
        self.sensor_details = QScrollArea()     # Área con scroll para contenido largo
        self.sensor_details.setVisible(False)  # Inicialmente oculto (se muestra al seleccionar sensor)
        self.sensor_details.setWidgetResizable(True)  # Permitir redimensionamiento automático
        self.sensor_details.setFrameShape(QFrame.NoFrame)  # Sin marco visible
        layout.addWidget(self.sensor_details)  # Agregar área de scroll al panel
        return panel                           # Retornar panel configurado completo

    # =====================================================================================
    # MÉTODO: MANEJO DE SELECCIÓN DE SENSORES
    # =====================================================================================    
    def on_sensor_selected(self, item):
        """
        Maneja la selección de un sensor de la lista del panel izquierdo
        
        Propósito: Cambiar la interfaz del panel derecho según el sensor seleccionado
        Funcionalidad: Validar conexión ESP32 y mostrar interfaz específica del sensor
        Parámetros: item - QListWidgetItem con datos del sensor seleccionado
        """
        
        # --- VALIDAR CONEXIÓN ANTES DE MOSTRAR SENSOR ---
        if not self.is_connected:                # Verificar si ESP32 está conectado
            QMessageBox.warning(self, "Sin conexión", "Debes conectar al ESP32 primero")
            return                               # Salir sin hacer nada si no hay conexión
        
        # --- DETENER TODOS LOS THREADS ACTIVOS ANTES DE CAMBIAR INTERFAZ ---
        self.stop_all_monitoring_threads()
        
        # --- OBTENER IDENTIFICADOR DEL SENSOR ---
        sensor_id = item.data(Qt.UserRole)       # Obtener ID único almacenado en el item
        
        # --- MOSTRAR INTERFAZ ESPECÍFICA SEGÚN SENSOR ---
        if sensor_id == "angulo_simple":         # Sensor de ángulo con potenciómetro simple
            anguloSimple_UI(self)  # Mostrar la interfaz de sensor simple
        elif sensor_id == "brazo_angulo":        # Brazo robótico con múltiples sensores
            self.show_brazo_angulo_interface()
        elif sensor_id == "distancia_ir":        # Sensor de distancia infrarrojo
            self.show_distancia_ir_interface()
        elif sensor_id == "distancia_cap":       # Sensor de distancia capacitivo
            self.show_distancia_cap_interface()
        elif sensor_id == "distancia_ultra":     # Sensor ultrasónico HC-SR04
            self.show_distancia_ultra_interface()
        else:                                    # Sensores no implementados aún
            QMessageBox.information(self, "Próximamente", 
                                  "Esta función será implementada en futuras versiones")

    # =====================================================================================
    # MÉTODO: INTERFAZ DEL SENSOR DE BRAZO CON MÚLTIPLES ÁNGULOS
    # =====================================================================================
    def show_brazo_angulo_interface(self):
        """
        Crea y muestra la interfaz para el brazo robótico con múltiples sensores
        
        Propósito: Interfaz completa para monitorear brazo con 3 potenciómetros + sensor capacitivo
        Funcionalidad: Diagrama de conexiones, controles de monitoreo, visualización multi-canal
        Sensores: 3 potenciómetros (GPIO 32, 33, 25) + sensor capacitivo (GPIO 4)
        """
        
        # --- OCULTAR PANTALLA DE BIENVENIDA ---
        self.welcome_widget.setVisible(False)   # Esconder mensaje inicial
        
        # --- CREAR WIDGET PRINCIPAL DEL SENSOR ---
        sensor_widget = QWidget()               # Contenedor principal de la interfaz
        layout = QVBoxLayout(sensor_widget)     # Layout vertical para organizar elementos
        layout.setSpacing(20)                   # Espacio entre secciones: 20px
        
        # =====================================================================================
        # SECCIÓN: TÍTULO Y DESCRIPCIÓN DEL SENSOR
        # =====================================================================================
        
        # --- TÍTULO PRINCIPAL ---
        title = QLabel("🦾 Sensor de Brazo Ángulo")  # Título con emoji de brazo robótico
        title.setStyleSheet("""
            font-size: 20px;                    /* Tamaño grande para destacar */
            font-weight: bold;                  /* Negrita para jerarquía visual */
            color: #007bff;                     /* Azul corporativo */
            margin-bottom: 10px;                /* Separación inferior */
        """)
        layout.addWidget(title)                 # Agregar título al layout principal
        
        # --- DESCRIPCIÓN FUNCIONAL ---
        description = QLabel("Monitorea 3 ángulos simultáneamente usando potenciómetros en GPIO 32, 33, 34 y sensor capacitivo en GPIO 25 del ESP32")
        description.setStyleSheet("""
            font-size: 14px;                    /* Tamaño legible */
            color: #6c757d;                     /* Gris suave */
            margin-bottom: 20px;                /* Separación inferior generosa */
        """)
        description.setWordWrap(True)           # Permitir salto de línea automático
        layout.addWidget(description)           # Agregar descripción al layout
        
        # =====================================================================================
        # SECCIÓN: DIAGRAMA DE CONEXIONES MÚLTIPLES ESP32
        # =====================================================================================
        
        # --- GRUPO DEL DIAGRAMA PARA BRAZO ---
        diagram_group = QGroupBox("🔌 Diagrama de Conexiones ESP32 - Brazo Ángulo")  # Título específico
        diagram_layout = QVBoxLayout(diagram_group)  # Layout vertical para el contenido
        
        # --- DIAGRAMA ASCII COMPLEJO PARA MÚLTIPLES SENSORES ---
        diagram_text = QLabel("""
<pre style="font-family: 'Courier New', monospace; font-size: 12px; line-height: 1.4; color: #495057;">
┌─────────────────────────────────┐
│  ESP32 DevKit V1                │
│                                 │
│  3V3  ○ ←── Potenciómetros (+)  │
│  GND  ○ ←── Potenciómetros (-)  │
│  D32  ○ ←── Potenciómetro 1 (S) │
│  D33  ○ ←── Potenciómetro 2 (S) │
│  D34  ○ ←── Potenciómetro 3 (S) │
│  D25  ○ ←── Sensor Capacitivo   │
│                                 │
│  LED integrado: GPIO 2          │
└─────────────────────────────────┘

<b>3 Potenciómetros 10kΩ:</b>
• Pin (+): Alimentación 3.3V (todos)
• Pin (S): Señales analógicas:
• Pin (-): Tierra (GND) (todos)
  - Potenciómetro 1 → GPIO 32 (Base)
  - Potenciómetro 2 → GPIO 33 (Articulación 1)  
  - Potenciómetro 3 → GPIO 34 (Articulación 2)

<b>Sensor Capacitivo:</b>
• Señal digital → GPIO 25 (con pull-up interno)
</pre>
        """)
        diagram_text.setWordWrap(True)          # Permitir ajuste de texto
        diagram_text.setStyleSheet("""
            background-color: #f8f9fa;          /* Fondo gris muy claro para diagrama */
            border: 2px solid #dee2e6;          /* Borde gris para definir área */
            border-radius: 6px;                 /* Esquinas redondeadas */
            padding: 15px;                      /* Espacio interno generoso */
            margin: 5px;                        /* Margen exterior pequeño */
        """)
        diagram_layout.addWidget(diagram_text)  # Agregar diagrama al grupo
        
        # --- NOTA ESPECÍFICA PARA BRAZO ROBÓTICO ---
        note_label = QLabel("💡 <b>Nota:</b> Este sensor simula un brazo robótico con 3 articulaciones. El sensor capacitivo simula el agarre.")
        note_label.setStyleSheet("""
            font-size: 13px;                    /* Tamaño menor para nota */
            color: #856404;                     /* Color ámbar oscuro */
            background-color: #fff3cd;          /* Fondo ámbar claro (alerta) */
            border: 1px solid #ffeaa7;          /* Borde ámbar */
            border-radius: 4px;                 /* Esquinas redondeadas */
            padding: 8px;                       /* Espacio interno */
            margin-top: 5px;                    /* Separación superior */
        """)
        note_label.setWordWrap(True)            # Permitir ajuste de línea
        diagram_layout.addWidget(note_label)    # Agregar nota al grupo
        
        layout.addWidget(diagram_group)         # Agregar grupo completo al layout principal
        
        # =====================================================================================
        # SECCIÓN: CONTROLES DE MONITOREO MÚLTIPLE
        # =====================================================================================
        
        # --- GRUPO DE CONTROLES ---
        controls_group = QGroupBox("Controles")  # Caja agrupada para controles
        controls_layout = QVBoxLayout(controls_group)  # Layout vertical para controles
        
        # =====================================================================================
        # SUB-SECCIÓN: ESTADO DE MÚLTIPLES POTENCIÓMETROS
        # =====================================================================================
        # Muestra el estado de los 3 potenciómetros del brazo robótico en tiempo real
        self.brazo_labels = {}                  # Diccionario para almacenar referencias a etiquetas
        for i in range(1, 4):                   # Crear etiquetas para potenciómetros 1, 2 y 3
            label = QLabel(f"Potenciómetro {i}: Lectura: -- | Ángulo: --°")  # Texto inicial
            
            # --- ESTILO PARA ETIQUETAS DE ESTADO DE POTENCIÓMETROS ---
            label.setStyleSheet("""
                font-size: 14px;                /* Tamaño de fuente legible */
                font-weight: bold;              /* Texto en negrita para destacar */
                color: #495057;                 /* Color gris oscuro para el texto */
                padding: 8px;                   /* Espaciado interno de 8px */
                background-color: #f8f9fa;      /* Fondo gris muy claro */
                border-radius: 6px;             /* Esquinas redondeadas de 6px */
                border: 2px solid #dee2e6;      /* Borde gris claro de 2px */
                margin: 2px;                    /* Margen externo pequeño */            """)
            self.brazo_labels[f'pot{i}'] = label
            controls_layout.addWidget(label)    # Agregar etiqueta al layout de controles
        
        # =====================================================================================
        # SUB-SECCIÓN: ESTADO DEL SENSOR CAPACITIVO
        # =====================================================================================
        # Muestra el estado digital (True/False) del sensor capacitivo del brazo
        self.capacitive_label = QLabel("Sensor Capacitivo: --")  # Etiqueta para sensor capacitivo
        
        # --- ESTILO PARA ETIQUETA DE SENSOR CAPACITIVO ---
        self.capacitive_label.setStyleSheet("""
            font-size: 14px;                    /* Tamaño consistente con otros sensores */
            font-weight: bold;                  /* Texto en negrita */
            color: #495057;                     /* Color gris oscuro */
            padding: 8px;                       /* Espaciado interno */
            background-color: #f8f9fa;          /* Fondo gris claro igual que potenciómetros */
            border-radius: 6px;                 /* Esquinas redondeadas */
            border: 2px solid #dee2e6;          /* Borde gris claro */
            margin: 2px;                        /* Margen pequeño */
        """)
        controls_layout.addWidget(self.capacitive_label)
        
        # ==================== BOTONES DE CONTROL PARA BRAZO ROBÓTICO ====================
        buttons_layout = QHBoxLayout()
          # BOTÓN INICIAR MONITOREO - Verde para acción positiva
        self.brazo_start_btn = QPushButton("▶️ Iniciar Monitoreo")
        self.brazo_start_btn.clicked.connect(self.toggle_brazo_monitoring)
        self.brazo_start_btn.setStyleSheet("QPushButton { background-color: #28a745; border-color: #28a745; }")
        buttons_layout.addWidget(self.brazo_start_btn)
        
        controls_layout.addLayout(buttons_layout)
        
        # ==================== BOTONES DE ACCIONES SECUNDARIAS ==================== 
        actions_layout = QHBoxLayout()
        
        # BOTÓN LIMPIAR GRÁFICA - Para borrar datos del brazo robótico
        self.brazo_clear_btn = QPushButton("🗑️ Limpiar Gráfica")
        self.brazo_clear_btn.clicked.connect(self.clear_brazo_graph)
        actions_layout.addWidget(self.brazo_clear_btn)
        
        # BOTÓN EXPORTAR - Para guardar datos en Excel (3 potenciómetros + capacitivo)
        self.brazo_export_btn = QPushButton("📊 Exportar Excel")
        self.brazo_export_btn.clicked.connect(self.export_brazo_to_excel)
        self.brazo_export_btn.setEnabled(False)  # Se habilita solo cuando hay datos
        actions_layout.addWidget(self.brazo_export_btn)
        
        controls_layout.addLayout(actions_layout)
        layout.addWidget(controls_group)
        
        # Gráfica mejorada para múltiples canales
        graph_group = QGroupBox("Gráfica en Tiempo Real - Múltiples Ángulos")
        graph_layout = QVBoxLayout(graph_group)
        
        # Configurar matplotlib con colores mejorados para múltiples líneas
        self.brazo_figure = Figure(figsize=(10, 6), dpi=100, facecolor='white')
        self.brazo_canvas = FigureCanvasQTAgg(self.brazo_figure)
        self.brazo_ax = self.brazo_figure.add_subplot(111)
        
        # Mejorar colores y estilo del gráfico
        self.brazo_ax.set_facecolor('#f8f9fa')
        self.brazo_ax.grid(True, linestyle='--', alpha=0.7, color='#dee2e6')
        self.brazo_ax.set_xlabel('Muestras', fontsize=12, fontweight='bold', color='#495057')
        self.brazo_ax.set_ylabel('Ángulo (°)', fontsize=12, fontweight='bold', color='#495057')
        self.brazo_ax.set_title('Monitoreo de Brazo Robótico - 3 Ángulos', fontsize=14, fontweight='bold', color='#007bff')
        
        # Líneas de datos con diferentes colores para cada potenciómetro
        colors = ["#cce744", "#00dafb", "#f500cc"]  # Azul, Verde, Rojo
        labels = ['Base (Pot 1)', 'Articulación 1 (Pot 2)', 'Articulación 2 (Pot 3)']
        
        self.brazo_lines = []
        for i, (color, label) in enumerate(zip(colors, labels)):
            line, = self.brazo_ax.plot([], [], 'o-', linewidth=3, markersize=4,
                                     color=color, label=label,
                                     markerfacecolor=color, 
                                     markeredgecolor='white', markeredgewidth=1)
            self.brazo_lines.append(line)
        
        # Agregar leyenda
        self.brazo_ax.legend(loc='upper right', fontsize=10)
          # Configurar límites iniciales
        self.brazo_ax.set_xlim(0, 100)
        self.brazo_ax.set_ylim(-135, 135)  # Rango de -135° a +135°
        
        # Mejorar el layout del gráfico
        self.brazo_figure.tight_layout(pad=2.0)
        
        # Inicializar el canvas con un dibujo inicial
        self.brazo_canvas.draw()
        
        graph_layout.addWidget(self.brazo_canvas)
        layout.addWidget(graph_group)
        
        # Inicializar listas de datos para los 3 potenciómetros
        self.brazo_angulos = [[], [], []]  # Listas para cada potenciómetro
        self.brazo_lecturas = [[], [], []]
        self.brazo_capacitive_states = []
        self.brazo_max_points = 100
        self.brazo_is_monitoring = False
          # Mostrar en el panel derecho
        self.sensor_details.setWidget(sensor_widget)
        self.sensor_details.setVisible(True)    
    # =====================================================================================
    # MÉTODO: INTERFAZ DEL SENSOR DE DISTANCIA INFRARROJO (IR)
    # =====================================================================================
    def show_distancia_ir_interface(self):
        """
        Crea y muestra la interfaz específica para el sensor de distancia infrarrojo digital
        
        Propósito: Interfaz para monitorear sensor IR digital que detecta presencia/ausencia
        Funcionalidad: Diagrama de conexiones, estado digital ON/OFF, controles de monitoreo
        Sensor: Sensor IR digital conectado al GPIO 14 del ESP32
        Tipo: Digital (Solo detección de presencia, no medición de distancia exacta)
        """
        
        # --- OCULTAR PANTALLA DE BIENVENIDA ---
        self.welcome_widget.setVisible(False)   # Esconder mensaje inicial
        
        # --- CREAR WIDGET PRINCIPAL DEL SENSOR ---
        sensor_widget = QWidget()               # Contenedor principal de la interfaz
        layout = QVBoxLayout(sensor_widget)     # Layout vertical para organizar elementos
        layout.setSpacing(20)                   # Espacio entre secciones: 20px
          # ==================== TÍTULO DEL SENSOR IR DIGITAL ====================
        title = QLabel("📡 Sensor Infrarrojo Digital (IR)")
        # ESTILO PARA TÍTULO - Gradiente rojo temático del sensor IR
        title.setStyleSheet("""
            font-size: 24px;                   /* Tamaño grande para título principal */
            font-weight: bold;                  /* Texto en negrita */
            color: #2c3e50;                     /* Color base (se sobrescribe) */
            background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0, 
                                      stop: 0 #e74c3c, stop: 1 #c0392b);  /* Gradiente rojo IR */
            color: white;                       /* Texto blanco sobre fondo rojo */
            padding: 15px;                      /* Espaciado interno generoso */
            border-radius: 10px;                /* Esquinas muy redondeadas */
            margin-bottom: 10px;                /* Margen inferior */
        """)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Layout horizontal para contenido principal
        main_layout = QHBoxLayout()
        
        # Panel izquierdo - Diagrama de conexión
        left_panel = QGroupBox("🔌 Diagrama de Conexión")
        left_layout = QVBoxLayout(left_panel)
        
        connection_diagram = QLabel()
        connection_text = """
        <div style='font-family: monospace; font-size: 12px; line-height: 1.5;'>
        <b>ESP32 ↔ Sensor IR Digital</b><br><br>
        
        <table border='1' style='border-collapse: collapse; width: 100%;'>
        <tr style='background-color: #f39c12; color: white;'>
            <th style='padding: 8px;'>ESP32</th>
            <th style='padding: 8px;'>Sensor IR</th>
            <th style='padding: 8px;'>Cable</th>
        </tr>
        <tr>
            <td style='padding: 8px;'>3.3V</td>
            <td style='padding: 8px;'>VCC</td>
            <td style='padding: 8px; color: brown;'>🟤 Marrón</td>
        </tr>
        <tr style='background-color: #ecf0f1;'>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;color: blue;'>🔵 Azul</td>
        </tr>        <tr>
            <td style='padding: 8px;'>GPIO 14</td>
            <td style='padding: 8px;'>OUT</td>
            <td style='padding: 8px; color: black;'>⚫ Negro</td>
        </tr>
        </table><br>
        
        <b>Especificaciones:</b><br>
        • Tipo: Digital (ON/OFF)<br>
        • Voltaje: 3.3V<br>
        • Pull-up interno: Activo<br>
        • Detección: Presencia/Ausencia
        </div>
        """        # ==================== DIAGRAMA DE CONEXIÓN IR ====================
        connection_diagram.setText(connection_text)
        connection_diagram.setWordWrap(True)
        # ESTILO PARA DIAGRAMA - Marco rojo temático del sensor IR
        connection_diagram.setStyleSheet("""
            background-color: #ffffff;         /* Fondo blanco para legibilidad */
            border: 2px solid #e74c3c;         /* Borde rojo tema IR */
            border-radius: 8px;                /* Esquinas redondeadas */
            padding: 15px;                     /* Espaciado interno */
            color: black;                      /* Texto negro sobre fondo blanco */
        """)
        left_layout.addWidget(connection_diagram)
        left_panel.setMaximumWidth(400)
        main_layout.addWidget(left_panel)
        
        # Panel derecho - Estado digital
        right_panel = QGroupBox("🎛️ Estado Digital")
        right_layout = QVBoxLayout(right_panel)
        
        # Estado actual
        status_group = QGroupBox("🔍 Estado Actual")
        status_layout = QVBoxLayout(status_group)
          # ==================== ESTADO DIGITAL DEL SENSOR IR ====================
        # Etiqueta que muestra DETECCIÓN (verde) o SIN DETECCIÓN (rojo)
        self.distancia_ir_status = QLabel("🔴 SIN DETECCIÓN")
        # ESTILO PARA ESTADO DIGITAL - Indicador visual grande y claro
        self.distancia_ir_status.setStyleSheet("""
            font-size: 28px;                   /* Tamaño grande para máxima visibilidad */
            font-weight: bold;                  /* Texto en negrita */
            color: #dc3545;                     /* Color rojo para estado 'sin detección' */
            padding: 20px;                      /* Espaciado interno generoso */
            background-color: #f8f9fa;          /* Fondo gris muy claro */
            border-radius: 12px;                /* Esquinas muy redondeadas */
            border: 3px solid #dc3545;          /* Borde rojo grueso para énfasis */
            text-align: center;                 /* Texto centrado */
        """)
        self.distancia_ir_status.setAlignment(Qt.AlignCenter)
        status_layout.addWidget(self.distancia_ir_status)
        right_layout.addWidget(status_group)
        
        # Controles de monitoreo
        controls_group = QGroupBox("🕹️ Controles")
        controls_layout = QVBoxLayout(controls_group)
        
        buttons_layout = QHBoxLayout()        # ==================== BOTÓN INICIAR MONITOREO IR ====================
        self.start_distancia_ir_btn = QPushButton("▶️ Iniciar Monitoreo")
        self.start_distancia_ir_btn.clicked.connect(self.toggle_distancia_ir_monitoring)
        # ESTILO: Color rojo temático del sensor IR para consistencia visual
        self.start_distancia_ir_btn.setStyleSheet("QPushButton { background-color: #e74c3c; border-color: #e74c3c; color: white; padding: 10px; }")
        buttons_layout.addWidget(self.start_distancia_ir_btn)
        
        controls_layout.addLayout(buttons_layout)
        right_layout.addWidget(controls_group)
        
        # Información adicional
        info_group = QGroupBox("ℹ️ Información")
        info_layout = QVBoxLayout(info_group)        # ==================== INFORMACIÓN DEL SENSOR IR ====================
        info_text = QLabel("""
        <b>Sensor IR Digital:</b><br>
        • Detección simple de presencia/ausencia<br>
        • No mide distancia exacta<br>
        • Ideal para detección de obstáculos<br>
        • Bajo consumo de energía<br>
        • Respuesta rápida ON/OFF
        """)
        info_text.setWordWrap(True)
        # ESTILO PARA INFORMACIÓN - Fondo amarillo suave para destacar info importante
        info_text.setStyleSheet("padding: 10px; background-color: #fff3cd; border-radius: 5px; color: #856404;")
        info_layout.addWidget(info_text)
        right_layout.addWidget(info_group)
        
        main_layout.addWidget(right_panel)
        layout.addLayout(main_layout)
          # Configurar el widget en el área principal
        self.sensor_details.setWidget(sensor_widget)
        self.sensor_details.setVisible(True)    # Hacer visible el área de detalles del sensor
    
    # =====================================================================================
    # MÉTODO: INTERFAZ DEL SENSOR DE DISTANCIA CAPACITIVO
    # =====================================================================================
    def show_distancia_cap_interface(self):
        """
        Crea y muestra la interfaz específica para el sensor de distancia capacitivo digital
        
        Propósito: Interfaz para monitorear sensor capacitivo digital que detecta proximidad
        Funcionalidad: Diagrama de conexiones, estado digital ON/OFF, controles de monitoreo
        Sensor: Sensor capacitivo digital conectado al GPIO 35 del ESP32
        Tipo: Digital (Solo detección de proximidad, no medición de distancia exacta)
        Ventajas: Sensible a materiales no metálicos, ideal para detección de proximidad
        """
        
        # --- OCULTAR PANTALLA DE BIENVENIDA ---
        self.welcome_widget.setVisible(False)   # Esconder mensaje inicial
        
        # --- CREAR WIDGET PRINCIPAL DEL SENSOR ---
        sensor_widget = QWidget()               # Contenedor principal de la interfaz
        layout = QVBoxLayout(sensor_widget)     # Layout vertical para organizar elementos
        layout.setSpacing(20)                   # Espacio entre secciones: 20px
          # ==================== TÍTULO DEL SENSOR CAPACITIVO ====================
        title = QLabel("📡 Sensor Capacitivo Digital")
        # ESTILO PARA TÍTULO - Gradiente azul temático del sensor capacitivo
        title.setStyleSheet("""
            font-size: 24px;                   /* Tamaño grande para título principal */
            font-weight: bold;                  /* Texto en negrita */
            color: #2c3e50;                     /* Color base (se sobrescribe) */
            background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0, 
                                      stop: 0 #3498db, stop: 1 #2980b9);  /* Gradiente azul capacitivo */
            color: white;                       /* Texto blanco sobre fondo azul */
            padding: 15px;                      /* Espaciado interno generoso */
            border-radius: 10px;                /* Esquinas muy redondeadas */
            margin-bottom: 10px;                /* Margen inferior */
        """)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Layout horizontal para contenido principal
        main_layout = QHBoxLayout()
        
        # Panel izquierdo - Diagrama de conexión
        left_panel = QGroupBox("🔌 Diagrama de Conexión")
        left_layout = QVBoxLayout(left_panel)
        
        connection_diagram = QLabel()
        connection_text = """
        <div style='font-family: monospace; font-size: 12px; line-height: 1.5;'>
        <b>ESP32 ↔ Sensor Capacitivo Digital</b><br><br>
        
        <table border='1' style='border-collapse: collapse; width: 100%;'>
        <tr style='background-color: #3498db; color: white;'>
            <th style='padding: 8px;'>ESP32</th>
            <th style='padding: 8px;'>Sensor Cap.</th>
            <th style='padding: 8px;'>Cable</th>
        </tr>
        <tr>
            <td style='padding: 8px;'>3.3V</td>
            <td style='padding: 8px;'>VCC</td>
            <td style='padding: 8px; color: brown;'>🟤 Marrón</td>
        </tr>
        <tr style='background-color: #ecf0f1;'>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;color: blue;'>🔵 Azul</td>
        </tr>        <tr>
            <td style='padding: 8px;'>GPIO 14</td>
            <td style='padding: 8px;'>OUT</td>
            <td style='padding: 8px; color: black;'>⚫ Negro</td>
        </tr>
        </table><br>
        
        <b>Especificaciones:</b><br>
        • Tipo: Digital (ON/OFF)<br>
        • Voltaje: 3.3V<br>
        • Pull-up interno: Activo<br>
        • Detección: Presencia/Ausencia
        </div>
        """        
        # ==================== DIAGRAMA DE CONEXIÓN CAPACITIVO ====================
        connection_diagram.setText(connection_text)
        connection_diagram.setWordWrap(True)
        # ESTILO PARA DIAGRAMA - Marco azul temático del sensor capacitivo
        connection_diagram.setStyleSheet("""
            background-color: #ffffff;         /* Fondo blanco para legibilidad */
            border: 2px solid #3498db;         /* Borde azul tema capacitivo */
            border-radius: 8px;                /* Esquinas redondeadas */
            padding: 15px;                     /* Espaciado interno */
            color: black;                      /* Texto negro sobre fondo blanco */
        """)
        left_layout.addWidget(connection_diagram)
        left_panel.setMaximumWidth(400)
        main_layout.addWidget(left_panel)
        
        # Panel derecho - Estado digital
        right_panel = QGroupBox("🎛️ Estado Digital")
        right_layout = QVBoxLayout(right_panel)
        
        # Estado actual
        status_group = QGroupBox("🔍 Estado Actual")
        status_layout = QVBoxLayout(status_group)
          # ==================== ESTADO DIGITAL DEL SENSOR CAPACITIVO ====================
        # Etiqueta que muestra DETECCIÓN (verde) o SIN DETECCIÓN (rojo)
        self.distancia_cap_status = QLabel("🔴 SIN DETECCIÓN")
        # ESTILO PARA ESTADO DIGITAL - Indicador visual grande y claro (idéntico al IR)
        self.distancia_cap_status.setStyleSheet("""
            font-size: 28px;                   /* Tamaño grande para máxima visibilidad */
            font-weight: bold;                  /* Texto en negrita */
            color: #dc3545;                     /* Color rojo para estado 'sin detección' */
            padding: 20px;                      /* Espaciado interno generoso */
            background-color: #f8f9fa;          /* Fondo gris muy claro */
            border-radius: 12px;                /* Esquinas muy redondeadas */
            border: 3px solid #dc3545;          /* Borde rojo grueso para énfasis */
            text-align: center;                 /* Texto centrado */
        """)
        self.distancia_cap_status.setAlignment(Qt.AlignCenter)
        status_layout.addWidget(self.distancia_cap_status)
        right_layout.addWidget(status_group)
        
        # Controles de monitoreo
        controls_group = QGroupBox("🕹️ Controles")
        controls_layout = QVBoxLayout(controls_group)
        
        buttons_layout = QHBoxLayout()        # ==================== BOTÓN INICIAR MONITOREO CAPACITIVO ====================
        self.start_distancia_cap_btn = QPushButton("▶️ Iniciar Monitoreo")
        self.start_distancia_cap_btn.clicked.connect(self.toggle_distancia_cap_monitoring)
        # ESTILO: Color azul temático del sensor capacitivo para consistencia visual
        self.start_distancia_cap_btn.setStyleSheet("QPushButton { background-color: #3498db; border-color: #3498db; color: white; padding: 10px; }")
        buttons_layout.addWidget(self.start_distancia_cap_btn)
        
        controls_layout.addLayout(buttons_layout)
        right_layout.addWidget(controls_group)
        
        # Información adicional
        info_group = QGroupBox("ℹ️ Información")
        info_layout = QVBoxLayout(info_group)
          # ==================== INFORMACIÓN DEL SENSOR CAPACITIVO ====================
        info_text = QLabel("""
        <b>Sensor Capacitivo Digital:</b><br>
        • Detección simple de presencia/ausencia<br>
        • No mide distancia exacta<br>
        • Ideal para detección de proximidad<br>
        • Sensible a materiales no metálicos<br>
        • Respuesta rápida ON/OFF
        """)
        info_text.setWordWrap(True)        # ESTILO PARA INFORMACIÓN - Fondo azul suave para destacar info del sensor capacitivo
        info_text.setStyleSheet("padding: 10px; background-color: #d1ecf1; border-radius: 5px; color: black;")
        info_layout.addWidget(info_text)
        right_layout.addWidget(info_group)
        main_layout.addWidget(right_panel)
        layout.addLayout(main_layout)
        
        # Configurar el widget en el área principal
        self.sensor_details.setWidget(sensor_widget)
        self.sensor_details.setVisible(True)    # Hacer visible el área de detalles del sensor
    
    # =====================================================================================
    # MÉTODO: INTERFAZ DEL SENSOR ULTRASÓNICO HC-SR04
    # =====================================================================================
    def show_distancia_ultra_interface(self):
        """
        Crea y muestra la interfaz específica para el sensor ultrasónico HC-SR04
        
        Propósito: Interfaz completa para monitorear sensor ultrasónico de alta precisión
        Funcionalidad: Diagrama de conexiones, medición continua de distancia, gráfica en tiempo real
        Sensor: HC-SR04 conectado a GPIO 26 (TRIG) y GPIO 27 (ECHO) del ESP32
        Tipo: Analógico (Medición precisa de distancia de 2-400 cm)
        Principio: Ondas ultrasónicas de 40kHz con cálculo basado en tiempo de vuelo
        """
        
        # --- OCULTAR PANTALLA DE BIENVENIDA ---
        self.welcome_widget.setVisible(False)   # Esconder mensaje inicial
        
        # --- CREAR WIDGET PRINCIPAL DEL SENSOR ---
        sensor_widget = QWidget()               # Contenedor principal de la interfaz
        layout = QVBoxLayout(sensor_widget)     # Layout vertical para organizar elementos
        layout.setSpacing(20)                   # Espacio entre secciones: 20px
          # ==================== TÍTULO DEL SENSOR ULTRASÓNICO ====================
        title = QLabel("📏 Sensor Ultrasónico HC-SR04")
        # ESTILO PARA TÍTULO - Gradiente verde azulado temático del sensor ultrasónico
        title.setStyleSheet("""
            font-size: 24px;                   /* Tamaño grande para título principal */
            font-weight: bold;                  /* Texto en negrita */
            color: #2c3e50;                     /* Color base (se sobrescribe) */
            background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 0, 
                                      stop: 0 #17a2b8, stop: 1 #138496);  /* Gradiente cyan ultrasónico */
            color: white;                       /* Texto blanco sobre fondo cyan */
            padding: 15px;                      /* Espaciado interno generoso */
            border-radius: 10px;                /* Esquinas muy redondeadas */
            margin-bottom: 10px;                /* Margen inferior */
        """)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        
        # Layout horizontal para contenido principal
        main_layout = QHBoxLayout()
        
        # Panel izquierdo - Diagrama de conexión
        left_panel = QGroupBox("🔌 Diagrama de Conexión")
        left_layout = QVBoxLayout(left_panel)
        
        connection_diagram = QLabel()
        connection_text = """
        <div style='font-family: monospace; font-size: 12px; line-height: 1.5;'>
        <b>ESP32 ↔ Sensor HC-SR04</b><br><br>
        
        <table border='1' style='border-collapse: collapse; width: 100%;'>
        <tr style='background-color: #17a2b8; color: white;'>
            <th style='padding: 8px;'>ESP32</th>
            <th style='padding: 8px;'>HC-SR04</th>
            <th style='padding: 8px;'>Cable</th>
        </tr>
        <tr>
            <td style='padding: 8px;'>5V</td>
            <td style='padding: 8px;'>VCC</td>
            <td style='padding: 8px; color: red;'>🔴 Rojo</td>
        </tr>
        <tr style='background-color: #ecf0f1;'>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;'>GND</td>
            <td style='padding: 8px;'>⚫ Negro</td>
        </tr>        <tr>
            <td style='padding: 8px;'>GPIO 26</td>
            <td style='padding: 8px;'>TRIG</td>
            <td style='padding: 8px; color: green;'>🟢 Verde</td>
        </tr>
        <tr style='background-color: #ecf0f1;'>
            <td style='padding: 8px;'>GPIO 27</td>
            <td style='padding: 8px;'>ECHO</td>
            <td style='padding: 8px; color: blue;'>🔵 Azul</td>
        </tr>
        </table><br>
        
        <b>Especificaciones:</b><br>
        • Rango: 2-400 cm<br>
        • Voltaje: 5V<br>
        • Tipo: Analógico<br>
        • Precisión: ±3mm<br>
        • Frecuencia: 40kHz
        </div>
        """        # ==================== DIAGRAMA DE CONEXIÓN ULTRASÓNICO ====================
        connection_diagram.setText(connection_text)
        connection_diagram.setWordWrap(True)
        # ESTILO PARA DIAGRAMA - Marco cyan temático del sensor ultrasónico
        connection_diagram.setStyleSheet("""
            background-color: #ffffff;         /* Fondo blanco para legibilidad */
            border: 2px solid #17a2b8;         /* Borde cyan tema ultrasónico */
            border-radius: 8px;                /* Esquinas redondeadas */
            padding: 15px;                     /* Espaciado interno */
            color: black;
        """)
        left_layout.addWidget(connection_diagram)
        left_panel.setMaximumWidth(400)
        main_layout.addWidget(left_panel)
        
        # Panel derecho - Controles y lecturas
        right_panel = QGroupBox("🎛️ Control y Monitoreo")
        right_layout = QVBoxLayout(right_panel)
        
        # Lecturas actuales
        readings_group = QGroupBox("📊 Lecturas Actuales")
        readings_layout = QVBoxLayout(readings_group)        
        # ==================== LECTURAS DEL SENSOR ULTRASÓNICO ====================
        # Etiqueta que muestra distancia, velocidad del sonido y ecuación de cálculo
        self.distancia_ultra_label = QLabel("Distancia: -- cm")
        # ESTILO PARA LECTURAS - Color cyan temático con diseño destacado
        self.distancia_ultra_label.setStyleSheet("""
            font-size: 18px;                   /* Tamaño de fuente grande para lecturas */
            font-weight: bold;                  /* Texto en negrita */
            color: #17a2b8;                     /* Color cyan tema ultrasónico */
            padding: 15px;                      /* Espaciado interno generoso */
            background-color: #f8f9fa;          /* Fondo gris muy claro */
            border-radius: 8px;                 /* Esquinas redondeadas */
            border: 2px solid #17a2b8;          /* Borde cyan para consistencia */
        """)
        readings_layout.addWidget(self.distancia_ultra_label)
        
        # Información sobre velocidad del sonido y cálculo
        self.sound_speed_label = QLabel("Velocidad del sonido: 343 m/s (20°C)")
        self.sound_speed_label.setStyleSheet("""
            font-size: 14px;
            color: #495057;
            padding: 10px;
            background-color: #e9ecef;
            border-radius: 6px;
            margin: 5px 0px;
        """)
        readings_layout.addWidget(self.sound_speed_label)
        
        # Ecuación de cálculo
        self.equation_label = QLabel("Cálculo: Distancia = (Tiempo × 343 m/s) / 2")
        self.equation_label.setStyleSheet("""
            font-size: 13px;
            color: #6c757d;
            padding: 8px;
            background-color: #f8f9fa;
            border-radius: 4px;
            border-left: 4px solid #17a2b8;
            margin: 5px 0px;
        """)
        readings_layout.addWidget(self.equation_label)
        right_layout.addWidget(readings_group)
        
        # Controles de monitoreo
        controls_group = QGroupBox("🕹️ Controles")
        controls_layout = QVBoxLayout(controls_group)
        
        buttons_layout = QHBoxLayout()
          # ==================== BOTÓN INICIAR MONITOREO ULTRASÓNICO ====================
        self.start_distancia_ultra_btn = QPushButton("▶️ Iniciar Monitoreo")
        self.start_distancia_ultra_btn.clicked.connect(self.toggle_distancia_ultra_monitoring)        # ESTILO: Color cyan temático del sensor ultrasónico para consistencia visual
        self.start_distancia_ultra_btn.setStyleSheet("QPushButton { background-color: #17a2b8; border-color: #17a2b8; color: white; padding: 10px; }")
        buttons_layout.addWidget(self.start_distancia_ultra_btn)
        
        controls_layout.addLayout(buttons_layout)
        
        # Acciones adicionales
        actions_layout = QHBoxLayout()
          # ==================== BOTONES DE ACCIONES SECUNDARIAS ULTRASÓNICO ====================        # BOTÓN LIMPIAR GRÁFICA - Para borrar datos del sensor ultrasónico
        self.clear_distancia_ultra_btn = QPushButton("🗑️ Limpiar Gráfica")
        self.clear_distancia_ultra_btn.clicked.connect(self.clear_distancia_ultra_graph)
        actions_layout.addWidget(self.clear_distancia_ultra_btn)
        
        # BOTÓN EXPORTAR - Para guardar datos en Excel (distancia)
        self.export_distancia_ultra_btn = QPushButton("📊 Exportar Excel")
        self.export_distancia_ultra_btn.clicked.connect(self.export_distancia_ultra_to_excel)
        self.export_distancia_ultra_btn.setEnabled(False)  # Se habilita solo cuando hay datos
        actions_layout.addWidget(self.export_distancia_ultra_btn)
        
        controls_layout.addLayout(actions_layout)
        right_layout.addWidget(controls_group)
        
        # Información adicional
        info_group = QGroupBox("ℹ️ Información")
        info_layout = QVBoxLayout(info_group)
          # ==================== INFORMACIÓN DEL SENSOR ULTRASÓNICO ====================
        info_text = QLabel("""
        <b>Sensor HC-SR04:</b><br>
        • Alta precisión en medición de distancia<br>
        • Ideal para navegación de robots<br>
        • No afectado por color o transparencia<br>
        • Funciona con ondas ultrasónicas<br>
        • Excelente para distancias largas
        """)
        info_text.setWordWrap(True)
        # ESTILO PARA INFORMACIÓN - Fondo cyan suave para destacar info del sensor ultrasónico
        info_text.setStyleSheet("padding: 10px; background-color: #d1ecf1; border-radius: 5px; color: black;")
        info_layout.addWidget(info_text)
        right_layout.addWidget(info_group)
        
        main_layout.addWidget(right_panel)
        layout.addLayout(main_layout)
        
        # Gráfica
        graph_group = QGroupBox("📈 Gráfica en Tiempo Real")
        graph_layout = QVBoxLayout(graph_group)
          # ==================== CONFIGURACIÓN GRÁFICA ULTRASÓNICA ====================
        # Configurar matplotlib para sensor ultrasónico con tema cyan
        self.figure_ultra = Figure(figsize=(10, 6), dpi=100, facecolor='white')
        self.canvas_ultra = FigureCanvasQTAgg(self.figure_ultra)
        self.ax_ultra = self.figure_ultra.add_subplot(111)
        
        # CONFIGURACIÓN DE ESTILO PARA GRÁFICA ULTRASÓNICA
        self.ax_ultra.set_title('Sensor Ultrasónico HC-SR04 - Tiempo Real', fontsize=14, fontweight='bold', color='#17a2b8')  # Título cyan
        self.ax_ultra.set_xlabel('Tiempo (s)', fontsize=12)        # Etiqueta eje X
        self.ax_ultra.set_ylabel('Distancia (cm)', fontsize=12)    # Etiqueta eje Y
        self.ax_ultra.grid(True, alpha=0.3)                        # Rejilla sutil
        self.ax_ultra.set_facecolor('#f8f9fa')                     # Fondo gris claro
        
        # LÍNEA DE DATOS - Color cyan para consistencia temática
        self.line_ultra, = self.ax_ultra.plot([], [], 'c-', linewidth=2, label='Distancia Ultrasónica')
        self.ax_ultra.legend()
          # CONFIGURAR LÍMITES INICIALES
        self.ax_ultra.set_xlim(0, 60)     # 60 segundos de visualización
        self.ax_ultra.set_ylim(0, 400)    # Rango 0-400 cm (rango del HC-SR04)
        
        self.figure_ultra.tight_layout()
          # Inicializar el canvas con un dibujo inicial
        self.canvas_ultra.draw()
        
        graph_layout.addWidget(self.canvas_ultra)
        layout.addWidget(graph_group)
        
        # Configurar el widget en el área principal
        self.sensor_details.setWidget(sensor_widget)
        self.sensor_details.setVisible(True)    # Hacer visible el área de detalles del sensor
    
    # =====================================================================================
    # MÉTODO: ACTUALIZACIÓN OPTIMIZADA DE GRÁFICAS
    # =====================================================================================
    def update_graph_display(self):
        """
        Actualiza todas las gráficas de sensores de forma optimizada usando timer
        
        Propósito: Centralizar y optimizar la actualización de múltiples gráficas
        Funcionamiento: Usa flags de datos pendientes para evitar actualizaciones innecesarias
        Optimización: Solo redibuja canvas cuando hay datos nuevos pendientes
        Rendimiento: Evita bloqueos de UI con actualizaciones frecuentes
        Sensores: Ángulo simple, brazo robótico, IR, capacitivo, ultrasónico
        """
        
        # --- VERIFICAR SI HAY DATOS PENDIENTES ---
        # Solo actualizar si hay datos pendientes y la aplicación está activa
        if not self.pending_updates:           # Si no hay actualizaciones pendientes
            return                             # Salir sin procesar nada
            
        try:
            # ==================== ACTUALIZAR GRÁFICA ÁNGULO SIMPLE ====================
            # Actualizar gráfica de ángulo simple si hay datos pendientes
            if (self.pending_simple_data is not None and 
                hasattr(self, 'canvas') and hasattr(self, 'line')):
                self.canvas.draw()             # Redibujar canvas del sensor de ángulo                
            # ==================== ACTUALIZAR GRÁFICA BRAZO ROBÓTICO ====================
            # Actualizar gráfica de brazo si hay datos pendientes  
            if (self.pending_brazo_data is not None and 
                hasattr(self, 'brazo_canvas') and hasattr(self, 'brazo_lines')):
                self.brazo_canvas.draw()       # Redibujar canvas del brazo robótico (3 líneas)
                
            # ==================== ACTUALIZAR GRÁFICA SENSOR IR ====================
            # Actualizar gráfica IR si hay datos pendientes
            if (self.pending_distancia_ir_data is not None and 
                hasattr(self, 'canvas_ir') and hasattr(self, 'line_ir')):
                self.canvas_ir.draw()          # Redibujar canvas del sensor IR
                
            # ==================== ACTUALIZAR GRÁFICA SENSOR CAPACITIVO ====================
            # Actualizar gráfica capacitiva si hay datos pendientes
            if (self.pending_distancia_cap_data is not None and 
                hasattr(self, 'canvas_cap') and hasattr(self, 'line_cap')):
                self.canvas_cap.draw()         # Redibujar canvas del sensor capacitivo
                
            # ==================== ACTUALIZAR GRÁFICA SENSOR ULTRASÓNICO ====================
            # Actualizar gráfica ultrasónica si hay datos pendientes
            if (self.pending_distancia_ultra_data is not None and 
                hasattr(self, 'canvas_ultra') and hasattr(self, 'line_ultra')):
                self.canvas_ultra.draw()       # Redibujar canvas del sensor ultrasónico
                
            # --- LIMPIAR FLAGS DE ACTUALIZACIÓN ---
            # Limpiar flags de datos pendientes para próxima iteración
            self.pending_updates = False              # Resetear flag principal de actualizaciones
            self.pending_simple_data = None           # Limpiar datos del ángulo simple
            self.pending_brazo_data = None            # Limpiar datos del brazo robótico
            self.pending_distancia_ir_data = None     # Limpiar datos del sensor IR
            self.pending_distancia_cap_data = None    # Limpiar datos del sensor capacitivo
            self.pending_distancia_ultra_data = None  # Limpiar datos del sensor ultrasónico            
        except Exception as e:
            # --- MANEJO DE ERRORES SILENCIOSO ---
            # Continuar silenciosamente si hay errores de actualización gráfica
            # Esto evita crashes por problemas temporales de rendering
            pass
            
    # =====================================================================================
    # SECCIÓN: FUNCIONES DE CONEXIÓN Y GESTIÓN DEL ESP32
    # =====================================================================================    
    # ============================================================================
    # FUNCIONES DE CONEXIÓN Y MONITOREO
    # ============================================================================
    
    # =====================================================================================
    # MÉTODO: PROBAR CONEXIÓN CON ESP32
    # =====================================================================================
    def test_connection(self):
        """
        Prueba la conexión TCP/IP con el ESP32 usando la IP proporcionada
        
        Propósito: Verificar conectividad antes de iniciar cualquier monitoreo
        Funcionamiento: Envía comando LED_ON para probar comunicación
        Validación: Verifica respuesta "LED_ON_OK" del ESP32
        UI: Actualiza estado visual y habilita/deshabilita controles
        Red: Usa cliente TCP para comunicación con ESP32
        """
        
        # --- VALIDAR ENTRADA DE IP ---
        esp32_ip = self.ip_input.text().strip()  # Obtener IP limpia sin espacios
        
        if not esp32_ip:                         # Verificar que no esté vacía
            QMessageBox.warning(self, "IP requerida", "Ingresa la IP del ESP32")
            return                               # Salir si no hay IP
        
        # --- ACTUALIZAR INTERFAZ DURANTE CONEXIÓN ---
        # Deshabilitar botón y mostrar estado de conexión en progreso
        self.connect_btn.setEnabled(False)       # Prevenir múltiples intentos
        self.status_label.setText("🔄 Conectando...")  # Estado visual de progreso
        self.status_label.setStyleSheet("""
            padding: 8px;                        /* Espaciado interno */
            border-radius: 4px;                  /* Esquinas redondeadas */
            background-color: #cce5ff;           /* Fondo azul claro */
            color: #004085;                      /* Texto azul oscuro */
            font-weight: bold;                   /* Texto en negrita */
        """)
        self.repaint()                           # Forzar actualización inmediata de la interfaz
        
        # --- PROBAR CONEXIÓN TCP ---
        # Probar conexión directamente con el ESP32
        try:
            client = ESP32Client(esp32_ip)       # Crear cliente con IP proporcionada
            response = client.led_on()           # Enviar comando de prueba LED_ON
            
            # --- VERIFICAR RESPUESTA DEL ESP32 ---
            if "LED_ON_OK" in response:          # Verificar respuesta esperada
                self.on_connected(esp32_ip)      # Conexión exitosa
            else:
                self.on_disconnected()           # Respuesta inesperada
                
        except Exception as e:
            # --- MANEJAR ERROR DE CONEXIÓN ---
            self.on_disconnected()               # Error de conexión (timeout, red, etc.)    
    # =====================================================================================
    # MÉTODO: CALLBACK DE CONEXIÓN EXITOSA
    # =====================================================================================
    def on_connected(self, esp32_ip):
        """
        Callback ejecutado cuando la conexión con el ESP32 es exitosa
        
        Propósito: Configurar la aplicación para modo conectado
        Funcionalidad: Inicializar cliente ESP32, actualizar UI, mostrar sensores
        Parámetros: esp32_ip - Dirección IP del ESP32 conectado
        UI: Cambia estado visual, habilita lista de sensores
        """
        
        # --- ESTABLECER ESTADO DE CONEXIÓN ---
        self.is_connected = True                 # Flag global de conexión
        self.esp_client = ESP32Client(esp32_ip)  # Cliente para comunicación TCP
          # --- ACTUALIZAR INTERFAZ DE CONEXIÓN ---
        self.connect_btn.setText("🔌 Conectado")  # Cambiar texto del botón
        self.connect_btn.setEnabled(False)       # Deshabilitar botón (ya conectado)
        self.restart_btn.setEnabled(True)        # Habilitar botón de reinicio
        self.status_label.setText("✅ Conectado al ESP32")  # Estado exitoso
        self.status_label.setStyleSheet("""
            padding: 8px;                        /* Espaciado interno */
            border-radius: 4px;                  /* Esquinas redondeadas */
            background-color: #d4edda;           /* Fondo verde claro */
            color: #155724;                      /* Texto verde oscuro */
            font-weight: bold;                   /* Texto en negrita */
        """)
        
        # --- MOSTRAR LISTA DE SENSORES CON ANIMACIÓN ---
        self.show_sensors_with_animation()       # Efecto visual de aparición
    
    # =====================================================================================
    # MÉTODO: CALLBACK DE CONEXIÓN FALLIDA
    # =====================================================================================
    def on_disconnected(self):
        """
        Callback ejecutado cuando la conexión con el ESP32 falla o se pierde
        
        Propósito: Configurar la aplicación para modo desconectado
        Funcionalidad: Limpiar cliente ESP32, actualizar UI, ocultar sensores
        UI: Cambia estado visual, deshabilita funcionalidades que requieren conexión
        Seguridad: Previene operaciones sin conexión válida
        """
        
        # --- LIMPIAR ESTADO DE CONEXIÓN ---
        self.is_connected = False                # Flag global de desconexión
        self.esp_client = None                   # Limpiar cliente TCP
          # --- ACTUALIZAR INTERFAZ DE CONEXIÓN ---
        self.connect_btn.setText("🔌 Conectar al ESP32")  # Restaurar texto original
        self.connect_btn.setEnabled(True)        # Habilitar botón para reconectar
        self.restart_btn.setEnabled(False)       # Deshabilitar botón de reinicio
        self.status_label.setText("❌ Error de conexión")  # Estado de error
        self.status_label.setStyleSheet("""
            padding: 8px;                        /* Espaciado interno */
            border-radius: 4px;                  /* Esquinas redondeadas */
            background-color: #f8d7da;           /* Fondo rojo claro */
            color: #721c24;                      /* Texto rojo oscuro */
            font-weight: bold;                   /* Texto en negrita */
        """)
        
        # --- OCULTAR FUNCIONALIDADES ---
        self.sensors_group.setVisible(False)     # Ocultar lista de sensores
    
    # =====================================================================================
    # MÉTODO: ANIMACIÓN DE LISTA DE SENSORES
    # =====================================================================================
    def show_sensors_with_animation(self):
        """
        Muestra la lista de sensores disponibles con efecto de desvanecimiento suave
        
        Propósito: Proporcionar feedback visual atractivo al usuario
        Animación: Efecto fade-in de 0.8 segundos con curva suave
        UX: Mejora la experiencia visual de la aplicación
        Timing: Activado solo después de conexión exitosa
        """
        
        # --- HACER VISIBLE EL GRUPO DE SENSORES ---
        self.sensors_group.setVisible(True)      # Mostrar contenedor de sensores
        
        # --- CONFIGURAR EFECTO DE OPACIDAD ---
        self.fade_effect = QGraphicsOpacityEffect()  # Efecto de transparencia
        self.sensors_group.setGraphicsEffect(self.fade_effect)  # Aplicar al grupo
        
        # --- CREAR ANIMACIÓN DE APARICIÓN ---
        self.animation = QPropertyAnimation(self.fade_effect, b"opacity")  # Animar opacidad
        self.animation.setDuration(800)          # Duración: 800ms (0.8 segundos)
        self.animation.setStartValue(0.0)        # Inicio: Completamente transparente
        self.animation.setEndValue(1.0)          # Final: Completamente opaco
        self.animation.setEasingCurve(QEasingCurve.InOutQuad)  # Curva suave
        
        # --- INICIAR ANIMACIÓN ---
        self.animation.start()                   # Ejecutar efecto de desvanecimiento

    # =====================================================================================
    # SECCIÓN: FUNCIONES DE MONITOREO - BRAZO ROBÓTICO MULTI-SENSOR
    # =====================================================================================
    
    # =====================================================================================
    # MÉTODO: ALTERNAR MONITOREO DEL BRAZO ROBÓTICO
    # =====================================================================================
    def toggle_brazo_monitoring(self):
        """
        Alterna entre iniciar y detener el monitoreo del brazo robótico multi-sensor
        
        Propósito: Control unificado para sistema complejo de 3 potenciómetros + sensor capacitivo
        Lógica: Verifica estado actual y ejecuta acción opuesta
        Sensores: Base, Articulación 1, Articulación 2, Sensor capacitivo
        Estado: Basado en flag self.brazo_is_monitoring
        """
        
        if not self.brazo_is_monitoring:         # Si no está monitoreando brazo
            self.start_brazo_monitoring()        # Iniciar monitoreo completo
        else:                                    # Si ya está monitoreando
            self.stop_brazo_monitoring()         # Detener monitoreo completo
    
    # =====================================================================================
    # MÉTODO: INICIAR MONITOREO DEL BRAZO ROBÓTICO
    # =====================================================================================
    def start_brazo_monitoring(self):
        """
        Inicia el monitoreo en tiempo real del brazo robótico con múltiples sensores
        
        Propósito: Comenzar adquisición simultánea de 3 potenciómetros + sensor capacitivo
        Thread: Crea BrazoAnguloThread para comunicación compleja con ESP32
        Datos: Recibe 3 lecturas ADC + estado capacitivo en un solo paquete
        Protocolo: "BRAZO_ANGULO" - comando especializado para múltiples sensores
        UI: Actualiza botones y habilita exportación
        Gráfica: Inicia visualización de 3 líneas simultáneas
        """
        
        # --- VERIFICAR CONEXIÓN REQUERIDA ---
        if not self.is_connected:                # Verificar conexión TCP activa
            QMessageBox.warning(self, "Sin conexión", "Debes conectar al ESP32 primero")
            return                               # Salir si no hay conexión
        
        try:
            # --- CREAR Y CONFIGURAR THREAD MULTI-SENSOR ---
            self.brazo_thread = BrazoAnguloThread(self.esp_client.esp32_ip)  # Thread especializado
            self.brazo_thread.data_received.connect(self.update_brazo_data)  # Conectar señal compleja
            
            # --- INICIAR MONITOREO MULTI-SENSOR ---
            self.brazo_thread.start()            # Iniciar thread de comunicación
            self.brazo_is_monitoring = True      # Marcar estado como monitoreando brazo            
            # --- INICIAR MONITOREO MULTI-SENSOR ---
            self.brazo_thread.start()            # Iniciar thread de comunicación
            self.brazo_is_monitoring = True      # Marcar estado como monitoreando brazo
              # --- ACTUALIZAR INTERFAZ DE CONTROL ---
            self.brazo_start_btn.setText("⏸️ Pausar")  # Cambiar botón a pausar
            self.brazo_start_btn.setStyleSheet("QPushButton { background-color: #ffc107; border-color: #ffc107; }")  # Amarillo pausa
            self.brazo_export_btn.setEnabled(True)  # Habilitar exportación multi-datos
              # --- INICIAR ACTUALIZACIÓN GRÁFICA MULTI-LÍNEA ---
            self.manage_graph_timer()            # Gestionar timer compartido inteligentemente
            
        except Exception as e:
            # --- MANEJAR ERRORES DE INICIALIZACIÓN MULTI-SENSOR ---
            QMessageBox.critical(self, "Error", f"Error al iniciar monitoreo: {str(e)}")
    
    # =====================================================================================
    # MÉTODO: DETENER MONITOREO DEL BRAZO ROBÓTICO
    # =====================================================================================
    def stop_brazo_monitoring(self):
        """
        Detiene el monitoreo del brazo robótico multi-sensor y limpia recursos
        
        Propósito: Parar adquisición de múltiples sensores y liberar thread
        Thread: Detiene BrazoAnguloThread de forma segura
        UI: Restaura botones a estado inicial
        Recursos: Limpia objetos para evitar memory leaks
        """
        
        # --- DETENER THREAD MULTI-SENSOR ---
        if self.brazo_thread and self.brazo_thread.isRunning():  # Si existe y está corriendo
            self.brazo_thread.stop()             # Detener thread de forma segura
            self.brazo_thread = None             # Limpiar referencia
          # --- ACTUALIZAR ESTADO Y TIMERS ---
        self.brazo_is_monitoring = False         # Marcar como no monitoreando brazo
        self.manage_graph_timer()                # Gestionar timer compartido inteligentemente
          # --- RESTAURAR INTERFAZ DE CONTROL ---
        self.brazo_start_btn.setText("▶️ Iniciar Monitoreo")  # Restaurar texto inicial
        self.brazo_start_btn.setStyleSheet("QPushButton { background-color: #28a745; border-color: #28a745; }")  # Verde inicial
    
    # =====================================================================================
    # MÉTODO: ACTUALIZAR DATOS DEL BRAZO ROBÓTICO
    # =====================================================================================
    def update_brazo_data(self, lectura1, angulo1, lectura2, angulo2, lectura3, angulo3, sensor_estado):
        """
        Procesa y actualiza los datos recibidos del brazo robótico multi-sensor
        
        Propósito: Manejar datos complejos de múltiples sensores en tiempo real
        Parámetros: lectura1-3 (int) - Valores ADC crudos de 3 potenciómetros
                   angulo1-3 (float) - Ángulos calculados en grados para cada potenciómetro
                   sensor_estado (bool) - Estado del sensor capacitivo (True/False)
        Almacenamiento: Mantiene 3 listas separadas + estado capacitivo
        UI: Actualiza 4 etiquetas independientes (3 potenciómetros + capacitivo)
        Gráfica: Prepara datos para 3 líneas simultáneas con colores diferentes
        """
        
        # --- ALMACENAR DATOS DE MÚLTIPLES SENSORES ---
        # Agregar datos a las listas de cada potenciómetro
        for i, (lectura, angulo) in enumerate([(lectura1, angulo1), (lectura2, angulo2), (lectura3, angulo3)]):
            self.brazo_lecturas[i].append(lectura)  # Lista por potenciómetro
            self.brazo_angulos[i].append(angulo)    # Lista por potenciómetro
            
            # --- MANTENER LÍMITE DE PUNTOS POR SENSOR ---
            if len(self.brazo_lecturas[i]) > self.brazo_max_points:  # Control individual de memoria
                self.brazo_lecturas[i].pop(0)   # Eliminar más antiguo del sensor i
                self.brazo_angulos[i].pop(0)    # Eliminar más antiguo del sensor i
        
        # --- ALMACENAR ESTADO DEL SENSOR CAPACITIVO ---
        self.brazo_capacitive_states.append(sensor_estado)  # Estado ON/OFF del capacitivo
        if len(self.brazo_capacitive_states) > self.brazo_max_points:  # Control de memoria
            self.brazo_capacitive_states.pop(0)  # Eliminar estado más antiguo
          # --- ACTUALIZAR ETIQUETAS DE LECTURA ACTUAL CON VERIFICACIÓN DEFENSIVA ---
        try:
            if hasattr(self, 'brazo_labels') and 'pot1' in self.brazo_labels and self.brazo_labels['pot1'] is not None:
                self.brazo_labels['pot1'].setText(f"Potenciómetro 1: Lectura: {lectura1} | Ángulo: {angulo1}°")
            if hasattr(self, 'brazo_labels') and 'pot2' in self.brazo_labels and self.brazo_labels['pot2'] is not None:
                self.brazo_labels['pot2'].setText(f"Potenciómetro 2: Lectura: {lectura2} | Ángulo: {angulo2}°")
            if hasattr(self, 'brazo_labels') and 'pot3' in self.brazo_labels and self.brazo_labels['pot3'] is not None:
                self.brazo_labels['pot3'].setText(f"Potenciómetro 3: Lectura: {lectura3} | Ángulo: {angulo3}°")
            if hasattr(self, 'capacitive_label') and self.capacitive_label is not None:
                self.capacitive_label.setText(f"Sensor Capacitivo: {'Activado' if sensor_estado else 'Desactivado'}")
        except RuntimeError:
            # Widget has been deleted, stop monitoring
            if hasattr(self, 'brazo_is_monitoring'):
                self.brazo_is_monitoring = False
            return
        
        # --- PREPARAR DATOS PARA GRÁFICAS MÚLTIPLES ---
        # Actualizar gráfica de forma optimizada para 3 líneas
        if hasattr(self, 'brazo_lines'):         # Verificar que existen las 3 líneas
            for i, line in enumerate(self.brazo_lines):  # Iterar por cada línea de potenciómetro
                x_data = list(range(len(self.brazo_angulos[i])))  # Índices para eje X
                line.set_data(x_data, self.brazo_angulos[i])  # Actualizar datos de línea i
            
            # --- AJUSTAR LÍMITES DINÁMICOS PARA MÚLTIPLES SENSORES ---
            if len(self.brazo_angulos[0]) > 0:   # Si hay datos en al menos un sensor
                max_len = max(len(angles) for angles in self.brazo_angulos)  # Máximo entre sensores
                self.brazo_ax.set_xlim(0, max(100, max_len))  # Ajustar según sensor con más datos
        
        # --- MARCAR PARA ACTUALIZACIÓN GRÁFICA MULTI-LÍNEA ---
        self.pending_updates = True              # Flag para redibujado pendiente
        self.pending_brazo_data = (lectura1, angulo1, lectura2, angulo2, lectura3, angulo3, sensor_estado)
    
    # =====================================================================================
    # MÉTODO: LIMPIAR GRÁFICA DEL BRAZO ROBÓTICO
    # =====================================================================================
    def clear_brazo_graph(self):
        """
        Limpia todos los datos y gráficas del brazo robótico multi-sensor
        
        Propósito: Resetear visualización completa y datos almacenados de múltiples sensores
        Datos: Borra historial de 3 potenciómetros + sensor capacitivo
        Gráfica: Resetea 3 líneas de datos y límites de ejes
        UI: Restaura 4 etiquetas a estado inicial
        Exportación: Deshabilita botón hasta que haya nuevos datos
        """
        
        # --- LIMPIAR DATOS DE MÚLTIPLES SENSORES ---
        for i in range(3):                       # Iterar por cada potenciómetro
            self.brazo_lecturas[i].clear()       # Borrar lecturas del potenciómetro i
            self.brazo_angulos[i].clear()        # Borrar ángulos del potenciómetro i
        self.brazo_capacitive_states.clear()    # Borrar estados del sensor capacitivo
        
        # --- RESETEAR GRÁFICAS MÚLTIPLES ---
        if hasattr(self, 'brazo_lines'):         # Si existen las líneas de datos
            for line in self.brazo_lines:        # Iterar por cada línea
                line.set_data([], [])            # Limpiar datos de cada línea
            self.brazo_ax.set_xlim(0, 100)       # Restaurar límites iniciales
            self.brazo_canvas.draw()             # Redibujar canvas limpio
        
        # --- RESTAURAR ETIQUETAS MÚLTIPLES ---
        for i in range(1, 4):                    # Iterar por potenciómetros 1, 2, 3
            self.brazo_labels[f'pot{i}'].setText(f"Potenciómetro {i}: Lectura: -- | Ángulo: --°")
        self.capacitive_label.setText("Sensor Capacitivo: --")  # Resetear estado capacitivo
        self.brazo_export_btn.setEnabled(False) # Deshabilitar exportación sin datos    
    # =====================================================================================
    # MÉTODO: EXPORTAR DATOS DEL BRAZO ROBÓTICO A EXCEL
    # =====================================================================================
    def export_brazo_to_excel(self):
        """
        Exporta todos los datos del brazo robótico multi-sensor a archivo Excel complejo
        
        Propósito: Permitir análisis posterior de múltiples sensores y respaldo completo
        Formato: Archivo .xlsx con columnas múltiples y gráficas separadas para cada sensor
        Datos: 3 potenciómetros (lecturas ADC + ángulos) + sensor capacitivo + timestamps
        Gráficas: 3 gráficos de líneas independientes, uno por cada potenciómetro
        Validación: Verifica que existan datos en al menos uno de los sensores
        """
        
        # --- VERIFICAR DATOS DISPONIBLES EN MÚLTIPLES SENSORES ---
        if not any(self.brazo_lecturas):         # Si no hay datos en ningún sensor
            QMessageBox.information(self, "Sin datos", "No hay datos para exportar")
            return                               # Salir sin hacer nada
        
        try:
            # --- GENERAR NOMBRE DE ARCHIVO ÚNICO PARA BRAZO ---
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Formato: YYYYMMDD_HHMMSS
            filename, _ = QFileDialog.getSaveFileName(
                self, "Guardar datos",           # Título del diálogo
                f"SensoraCore_Brazo_{timestamp}.xlsx",  # Nombre específico para brazo
                "Excel files (*.xlsx)"           # Filtro de archivos
            )
            
            if filename:                         # Si usuario seleccionó archivo
                # --- CREAR WORKBOOK Y WORKSHEET PARA MÚLTIPLES SENSORES ---
                wb = openpyxl.Workbook()         # Nuevo libro de Excel
                ws = wb.active                   # Hoja activa
                ws.title = "Datos Brazo Ángulo"  # Título específico para brazo
                
                # --- CREAR HEADERS COMPLEJOS PARA MÚLTIPLES SENSORES ---
                headers = ["Muestra", "Lectura1", "Ángulo1", "Lectura2", "Ángulo2", 
                          "Lectura3", "Ángulo3", "Sensor_Cap", "Timestamp"]  # 9 columnas
                for i, header in enumerate(headers):  # Escribir cada header
                    ws.cell(row=1, column=i+1, value=header)
                
                # --- ESCRIBIR DATOS COMPLEJOS FILA POR FILA ---
                max_len = max(len(angles) for angles in self.brazo_angulos if angles)  # Máximo entre sensores
                for i in range(max_len):         # Iterar por cada muestra
                    row = i + 2                  # Fila actual (empezar en 2)
                    ws.cell(row=row, column=1, value=i+1)  # Número de muestra
                    
                    # --- ESCRIBIR DATOS DE CADA POTENCIÓMETRO ---
                    for j in range(3):           # Iterar por potenciómetros 0, 1, 2
                        if i < len(self.brazo_lecturas[j]):  # Si hay datos para este potenciómetro
                            ws.cell(row=row, column=j*2+2, value=self.brazo_lecturas[j][i])  # Lectura ADC
                            ws.cell(row=row, column=j*2+3, value=self.brazo_angulos[j][i])   # Ángulo calculado
                    
                    # --- ESCRIBIR ESTADO DEL SENSOR CAPACITIVO ---
                    if i < len(self.brazo_capacitive_states):  # Si hay estado capacitivo
                        ws.cell(row=row, column=8, value=self.brazo_capacitive_states[i])  # Estado ON/OFF
                    
                    # --- ESCRIBIR TIMESTAMP ---
                    ws.cell(row=row, column=9, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                
                # --- CREAR GRÁFICAS SEPARADAS PARA CADA POTENCIÓMETRO ---
                colors = ["0000FF", "00FF00", "FF0000"]  # Azul, Verde, Rojo para cada sensor
                for j in range(3):               # Crear gráfico para cada potenciómetro
                    chart = LineChart()          # Nuevo gráfico de líneas
                    chart.title = f"Ángulo Potenciómetro {j+1} vs Tiempo"  # Título específico
                    chart.y_axis.title = "Ángulo (°)"    # Etiqueta eje Y
                    chart.x_axis.title = "Muestra"       # Etiqueta eje X
                    
                    # --- CONFIGURAR DATOS ESPECÍFICOS DEL POTENCIÓMETRO ---
                    data = Reference(ws, min_col=j*2+3, min_row=1, max_row=max_len+1)  # Columna de ángulos
                    categories = Reference(ws, min_col=1, min_row=2, max_row=max_len+1)  # Columna de muestras
                    chart.add_data(data, titles_from_data=True)  # Agregar datos
                    chart.set_categories(categories)  # Establecer categorías
                      # --- POSICIONAR GRÁFICOS EN DIFERENTES UBICACIONES ---
                    # Columnas: K2, S2, AA2 para los tres gráficos
                    chart_positions = ["K2", "S2", "AA2"]
                    ws.add_chart(chart, chart_positions[j])
                
                # --- GUARDAR ARCHIVO COMPLEJO ---
                wb.save(filename)                # Guardar en ubicación seleccionada
                QMessageBox.information(self, "Éxito", f"Datos exportados a {filename}")
                
        except Exception as e:
            # --- MANEJAR ERRORES DE EXPORTACIÓN COMPLEJA ---
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")    
            
    # =====================================================================================
    # SECCIÓN: FUNCIONES DE MONITOREO - SENSOR DE DISTANCIA INFRARROJO (IR)
    # =====================================================================================
    
    # =====================================================================================
    # MÉTODO: ALTERNAR MONITOREO DEL SENSOR IR
    # =====================================================================================
    def toggle_distancia_ir_monitoring(self):
        """
        Alterna entre iniciar y detener el monitoreo del sensor de distancia infrarrojo
        
        Propósito: Control unificado para sensor IR digital de detección de presencia
        Lógica: Verifica estado actual y ejecuta acción opuesta
        Sensor: IR digital en GPIO 14 (solo detección ON/OFF, no distancia exacta)
        Estado: Basado en flag self.distancia_ir_is_monitoring
        """
        
        if not self.distancia_ir_is_monitoring:  # Si no está monitoreando IR
            self.start_distancia_ir_monitoring() # Iniciar monitoreo IR
        else:                                    # Si ya está monitoreando
            self.stop_distancia_ir_monitoring()  # Detener monitoreo IR
    
    # =====================================================================================
    # MÉTODO: INICIAR MONITOREO DEL SENSOR IR
    # =====================================================================================
    def start_distancia_ir_monitoring(self):
        """
        Inicia el monitoreo en tiempo real del sensor de distancia infrarrojo digital
        
        Propósito: Comenzar detección continua de presencia/ausencia con sensor IR
        Thread: Crea DistanciaIRThread para comunicación asíncrona con ESP32
        Datos: Recibe estados digitales ON/OFF del GPIO 14
        Protocolo: "DISTANCIA_IR" - comando específico para sensor infrarrojo
        UI: Actualiza estado visual y botones de control
        Tipo: Digital (solo detección, no medición de distancia exacta)
        """
        
        # --- VERIFICAR CONEXIÓN REQUERIDA ---
        if not self.is_connected:                # Verificar conexión TCP activa
            QMessageBox.warning(self, "Sin conexión", "Debes conectar al ESP32 primero")
            return                               # Salir si no hay conexión
        
        try:
            # --- CREAR Y CONFIGURAR THREAD IR ---
            self.distancia_ir_thread = DistanciaIRThread(self.esp_client.esp32_ip)  # Thread IR
            self.distancia_ir_thread.data_received.connect(self.update_distancia_ir_data)  # Conectar señal
            
            # --- INICIAR MONITOREO IR ---
            self.distancia_ir_thread.start()     # Iniciar thread de comunicación
            self.distancia_ir_is_monitoring = True  # Marcar estado como monitoreando IR
            
            # --- ACTUALIZAR INTERFAZ DE CONTROL IR ---
            self.start_distancia_ir_btn.setText("⏸️ Pausar")  # Cambiar botón a pausar
            self.start_distancia_ir_btn.setStyleSheet("QPushButton { background-color: #ffc107; border-color: #ffc107; color: white; padding: 10px; }")  # Amarillo pausa

            if hasattr(self, 'export_distancia_ir_btn'):  # Si existe botón de exportación
                self.export_distancia_ir_btn.setEnabled(True)  # Habilitar exportación
              # --- INICIAR ACTUALIZACIÓN GRÁFICA IR ---
            self.manage_graph_timer()            # Gestionar timer compartido inteligentemente
            
        except Exception as e:
            # --- MANEJAR ERRORES DE INICIALIZACIÓN IR ---
            QMessageBox.critical(self, "Error", f"Error al iniciar monitoreo: {str(e)}")
    
    # =====================================================================================
    # MÉTODO: DETENER MONITOREO DEL SENSOR IR
    # =====================================================================================
    def stop_distancia_ir_monitoring(self):
        """
        Detiene el monitoreo del sensor de distancia infrarrojo y limpia recursos
        
        Propósito: Parar detección IR y liberar thread
        Thread: Detiene DistanciaIRThread de forma segura
        UI: Restaura botones a estado inicial
        Recursos: Limpia objetos para evitar memory leaks
        """
        
        # --- DETENER THREAD IR ---
        if self.distancia_ir_thread and self.distancia_ir_thread.isRunning():  # Si existe y está corriendo
            self.distancia_ir_thread.stop()     # Detener thread de forma segura
            self.distancia_ir_thread = None     # Limpiar referencia
          # --- ACTUALIZAR ESTADO Y TIMERS ---
        self.distancia_ir_is_monitoring = False # Marcar como no monitoreando IR
        self.manage_graph_timer()                # Gestionar timer compartido inteligentemente
          # Actualizar interfaz        self.start_distancia_ir_btn.setText("▶️ Iniciar Monitoreo")
        self.start_distancia_ir_btn.setStyleSheet("QPushButton { background-color: #e74c3c; border-color: #e74c3c; color: white; padding: 10px; }")

    def update_distancia_ir_data(self, estado_digital):
        """Actualiza los datos del sensor de distancia IR (digital)"""
        try:
            # Verificar que el widget existe antes de actualizarlo
            if not hasattr(self, 'distancia_ir_status'):
                return
            
            # Invertir lógica: 1 = NO DETECTADO, 0 = DETECTADO
            estado_digital = not estado_digital
            estado_texto = "DETECTADO" if estado_digital else "NO DETECTADO"
            color = "#27ae60" if estado_digital else "#e74c3c"
            self.distancia_ir_status.setText(f"Estado: {estado_texto}")
            self.distancia_ir_status.setStyleSheet(f"""
                font-size: 18px;
                font-weight: bold;
                color: {color};
                padding: 15px;
                background-color: #f8f9fa;
                border-radius: 8px;
                border: 2px solid {color};
            """)
            # No necesitamos gráficas para sensores digitales, solo estado actual
            # Marcar para actualización si es necesario
            self.pending_updates = True
            self.pending_distancia_ir_data = estado_digital
        except RuntimeError:
            # Widget fue eliminado durante la actualización
            self.distancia_ir_is_monitoring = False
            return
    
    def clear_distancia_ir_graph(self):
        """Resetea el estado del sensor de distancia IR digital"""
        # Para sensores digitales, no hay datos continuos que exportar
        # Solo exportamos el estado actual si está disponible
        QMessageBox.information(self, "Sensor Digital", 
            "Los sensores digitales no generan datos continuos para exportar.\n\n"
            "El sensor IR solo proporciona estado ON/OFF en tiempo real.\n"
            "Estado actual visible en la interfaz.")
        
        # Resetear estado visual al estado inicial
        self.distancia_ir_status.setText("🔴 SIN DETECCIÓN")
        self.distancia_ir_status.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: #dc3545;
            padding: 20px;
            background-color: #f8f9fa;
            border-radius: 12px;
            border: 3px solid #dc3545;
            text-align: center;
        """)
        return
    
    def export_distancia_ir_to_excel(self):
        """Exporta los datos del sensor de distancia IR a Excel"""
        if not self.distancia_ir_lecturas:
            QMessageBox.information(self, "Sin datos", "No hay datos para exportar")
            return
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename, _ = QFileDialog.getSaveFileName(
                self, "Guardar datos",
                f"SensoraCore_DistanciaIR_{timestamp}.xlsx",
                "Excel files (*.xlsx)"
            )
            
            if filename:
                # Crear workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Datos Distancia IR"
                
                # Headers
                ws['A1'] = "Muestra"
                ws['B1'] = "Lectura ADC"
                ws['C1'] = "Voltaje (V)"
                ws['D1'] = "Distancia (cm)"
                ws['E1'] = "Timestamp"
                
                # Datos
                for i, (lectura, voltaje, distancia) in enumerate(zip(self.distancia_ir_lecturas, self.distancia_ir_voltajes, self.distancia_ir_cm)):
                    ws[f'A{i+2}'] = i+1
                    ws[f'B{i+2}'] = lectura
                    ws[f'C{i+2}'] = voltaje
                    ws[f'D{i+2}'] = distancia
                    ws[f'E{i+2}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Crear gráfica
                chart = LineChart()
                chart.title = "Distancia IR vs Tiempo"
                chart.y_axis.title = "Distancia (cm)"
                chart.x_axis.title = "Muestra"
                
                data = Reference(ws, min_col=4, min_row=1, max_row=len(self.distancia_ir_cm)+1)
                categories = Reference(ws, min_col=1, min_row=2, max_row=len(self.distancia_ir_cm)+1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                ws.add_chart(chart, "G2")
                
                # Guardar
                wb.save(filename)
                QMessageBox.information(self, "Éxito", f"Datos exportados a {filename}")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")    
    # ============================================================================
    # FUNCIONES DE MONITOREO - DISTANCIA CAPACITIVO
    # ============================================================================
    
    def toggle_distancia_cap_monitoring(self):
        """
        Controla el estado de monitoreo del sensor de distancia capacitivo.
        
        FUNCIONALIDAD:
        - Si el monitoreo está inactivo: inicia el monitoreo
        - Si el monitoreo está activo: detiene el monitoreo
        - Proporciona un punto de entrada unificado para el control del sensor capacitivo
        
        SENSOR CAPACITIVO:
        - Tipo: Digital (ON/OFF)
        - GPIO: Pin 35 (configurable en ESP32)
        - Principio: Detección de proximidad por cambio de capacitancia
        - Rango típico: 0-10 cm aproximadamente
        - Sensible a objetos metálicos y no metálicos
        """
        if not self.distancia_cap_is_monitoring:
            self.start_distancia_cap_monitoring()
        else:
            self.stop_distancia_cap_monitoring()
    
    def start_distancia_cap_monitoring(self):
        """
        Inicia el monitoreo del sensor de distancia capacitivo digital.
        
        PROCESO DE INICIALIZACIÓN:
        1. Verificar conexión ESP32
        2. Crear thread de comunicación (DistanciaCapThread)
        3. Configurar callback para recepción de datos
        4. Iniciar thread de monitoreo en segundo plano
        5. Actualizar interfaz gráfica y habilitar controles
        6. Activar timer de actualización de gráficas
        
        COMUNICACIÓN:
        - Protocolo: TCP Socket con ESP32
        - Comando: "DISTANCIA_CAP" al puerto TCP del ESP32
        - Respuesta: Estado digital (0 o 1) cada ~100ms
        - Thread independiente para no bloquear UI
        
        ERROR HANDLING:
        - Valida conexión antes de iniciar
        - Captura excepciones de creación de thread
        - Muestra mensajes de error específicos al usuario
        """
        if not self.is_connected:
            QMessageBox.warning(self, "Sin conexión", "Debes conectar al ESP32 primero")
            return
        
        try:
            # --- CREAR Y CONFIGURAR THREAD DE COMUNICACIÓN ---
            self.distancia_cap_thread = DistanciaCapThread(self.esp_client.esp32_ip)
            self.distancia_cap_thread.data_received.connect(self.update_distancia_cap_data)
            
            # --- INICIAR MONITOREO EN SEGUNDO PLANO ---
            self.distancia_cap_thread.start()
            self.distancia_cap_is_monitoring = True
              # --- ACTUALIZAR INTERFAZ DE USUARIO ---
            self.start_distancia_cap_btn.setText("⏸️ Pausar")
            self.start_distancia_cap_btn.setStyleSheet("QPushButton { background-color: #ffc107; border-color: #ffc107; color: white; padding: 10px; }")
            if hasattr(self, 'export_distancia_cap_btn'):
                self.export_distancia_cap_btn.setEnabled(True)
            
            # --- ACTIVAR TIMER DE ACTUALIZACIÓN GRÁFICA ---
            self.manage_graph_timer()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al iniciar monitoreo: {str(e)}")
    
    def stop_distancia_cap_monitoring(self):
        """
        Detiene el monitoreo del sensor de distancia capacitivo.
        
        PROCESO DE APAGADO:
        1. Verificar existencia y estado del thread
        2. Enviar señal de stop al thread
        3. Esperar finalización del thread y liberar recursos
        4. Actualizar variables de estado interno
        5. Detener timer de actualización de gráficas
        6. Restaurar interfaz gráfica al estado inicial
        
        LIMPIEZA DE RECURSOS:
        - Finaliza thread de comunicación TCP
        - Libera socket de conexión con ESP32
        - Resetea variables de estado de monitoreo
        - Detiene timers de actualización para optimizar CPU
        """
        # --- FINALIZAR THREAD DE COMUNICACIÓN ---
        if self.distancia_cap_thread and self.distancia_cap_thread.isRunning():
            self.distancia_cap_thread.stop()    # Enviar señal de parada
            self.distancia_cap_thread = None    # Liberar referencia del thread
          # --- ACTUALIZAR ESTADO Y TIMERS ---
        self.distancia_cap_is_monitoring = False # Marcar como no monitoreando capacitivo
        self.manage_graph_timer()                 # Gestionar timer compartido inteligentemente
          # --- RESTAURAR INTERFAZ AL ESTADO INICIAL ---
        self.start_distancia_cap_btn.setText("▶️ Iniciar Monitoreo")
        self.start_distancia_cap_btn.setStyleSheet("QPushButton { background-color: #3498db; border-color: #3498db; color: white; padding: 10px; }")
    
    def update_distancia_cap_data(self, estado_digital):
        """
        Actualiza la interfaz con datos del sensor de distancia capacitivo digital.
        
        PARÁMETROS:
        - estado_digital (bool): Estado del sensor (True/False del GPIO)
        
        LÓGICA DE INVERSIÓN:
        - GPIO = 1 (HIGH) → "NO DETECTADO" (estado normal del sensor)
        - GPIO = 0 (LOW)  → "DETECTADO" (objeto cerca del sensor)
        - Se invierte la lógica para que True = detectado
        
        ACTUALIZACIÓN VISUAL:
        - Cambia texto del indicador de estado
        - Modifica colores dinámicamente (verde/rojo)
        - Aplica estilos CSS responsivos
        - Actualiza bordes y fondos según el estado
        
        NOTA: Los sensores digitales no generan gráficas continuas,
        solo muestran el estado actual ON/OFF en tiempo real.        """
        # --- PROCESAR LÓGICA INVERSAS DEL SENSOR ---
        # Inversión necesaria: sensor capacitivo entrega HIGH cuando NO detecta
        try:
            # Verificar que el widget existe antes de actualizarlo
            if not hasattr(self, 'distancia_cap_status'):
                return
                
            estado_digital = not estado_digital
            estado_texto = "🟢 DETECTADO" if estado_digital else "🔴 SIN DETECCIÓN"
            color = "#28a745" if estado_digital else "#dc3545"
            
            # --- ACTUALIZAR INTERFAZ VISUAL ---
            self.distancia_cap_status.setText(estado_texto)
            self.distancia_cap_status.setStyleSheet(f"""
                font-size: 28px;
                font-weight: bold;
                color: {color};
                padding: 20px;
                background-color: #f8f9fa;
                border-radius: 12px;
                border: 3px solid {color};
                text-align: center;
            """)
              # --- MARCAR PARA ACTUALIZACIÓN DE GRÁFICAS ---
            # No se usan gráficas para sensores digitales, solo estado actual
            self.pending_updates = True
            self.pending_distancia_cap_data = estado_digital
        except RuntimeError:
            # Widget fue eliminado durante la actualización
            self.distancia_cap_is_monitoring = False
            return
    # ============================================================================
    # FUNCIONES DE MONITOREO - DISTANCIA ULTRASÓNICO
    # ============================================================================
    
    def toggle_distancia_ultra_monitoring(self):
        """
        Controla el estado de monitoreo del sensor de distancia ultrasónico HC-SR04.
        
        FUNCIONALIDAD:
        - Si el monitoreo está inactivo: inicia el monitoreo
        - Si el monitoreo está activo: detiene el monitoreo
        - Proporciona control unificado para el sensor ultrasónico
        
        SENSOR ULTRASÓNICO HC-SR04:
        - Tipo: Analógico (valores continuos)
        - Rango: 2-400 cm aproximadamente
        - Principio: Tiempo de vuelo de ondas ultrasónicas
        - Precisión: ±3mm bajo condiciones ideales
        - Frecuencia: ~40 kHz
        - Alimentación: 5V (con divisor de voltaje para ESP32)
        """
        if not self.distancia_ultra_is_monitoring:
            self.start_distancia_ultra_monitoring()
        else:
            self.stop_distancia_ultra_monitoring()
    
    def start_distancia_ultra_monitoring(self):
        """
        Inicia el monitoreo del sensor de distancia ultrasónico HC-SR04.
        
        PROCESO DE INICIALIZACIÓN:
        1. Verificar conexión con ESP32
        2. Crear thread de comunicación (DistanciaUltrasonicThread)
        3. Configurar callback para datos analógicos continuos
        4. Iniciar thread para lectura en tiempo real
        5. Habilitar interfaz gráfica y controles de exportación
        6. Activar timers para actualización de gráficas en vivo
        
        COMUNICACIÓN:
        - Protocolo: TCP Socket con ESP32
        - Comando: "DISTANCIA_ULTRA" enviado al ESP32
        - Respuesta: Lectura ADC, voltaje calculado, distancia en cm
        - Frecuencia: ~10 Hz (cada 100ms aproximadamente)
        - Thread dedicado para no bloquear UI principal
        
        CONFIGURACIÓN DEL SENSOR:
        - Trigger Pin: GPIO configurable (ej: GPIO 25)
        - Echo Pin: GPIO configurable (ej: GPIO 26)
        - Conversión: tiempo_echo * 0.034 / 2 = distancia_cm
        """
        if not self.is_connected:
            QMessageBox.warning(self, "Sin conexión", "Debes conectar al ESP32 primero")
            return
        
        try:
            # --- CREAR Y CONFIGURAR THREAD DE COMUNICACIÓN ---
            self.distancia_ultra_thread = DistanciaUltrasonicThread(self.esp_client.esp32_ip)
            self.distancia_ultra_thread.data_received.connect(self.update_distancia_ultra_data)
            
            # --- INICIAR MONITOREO EN SEGUNDO PLANO ---
            self.distancia_ultra_thread.start()
            self.distancia_ultra_is_monitoring = True
              # --- ACTUALIZAR INTERFAZ DE USUARIO ---
            self.start_distancia_ultra_btn.setText("⏸️ Pausar")
            self.start_distancia_ultra_btn.setStyleSheet("QPushButton { background-color: #ffc107; border-color: #ffc107; color: white; padding: 10px; }")
            if hasattr(self, 'export_distancia_ultra_btn'):
                self.export_distancia_ultra_btn.setEnabled(True)
              # --- ACTIVAR TIMER DE ACTUALIZACIÓN GRÁFICA ---
            self.manage_graph_timer()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al iniciar monitoreo: {str(e)}")
    
    def stop_distancia_ultra_monitoring(self):
        """
        Detiene el monitoreo del sensor de distancia ultrasónico.
        
        PROCESO DE APAGADO:
        1. Verificar estado del thread de comunicación
        2. Enviar señal de stop al thread ultrasónico
        3. Esperar finalización y liberar recursos del thread
        4. Actualizar variables de estado interno
        5. Detener timers de actualización de gráficas
        6. Restaurar interfaz gráfica al estado inicial
        
        LIMPIEZA DE RECURSOS:
        - Finaliza thread de comunicación TCP con ESP32
        - Libera memoria de buffer de datos ultrasónicos
        - Resetea banderas de estado de monitoreo
        - Optimiza CPU deteniendo actualizaciones innecesarias
        """
        # --- FINALIZAR THREAD DE COMUNICACIÓN ---
        if self.distancia_ultra_thread and self.distancia_ultra_thread.isRunning():
            self.distancia_ultra_thread.stop()     # Enviar señal de parada
            self.distancia_ultra_thread = None     # Liberar referencia del thread
          # --- ACTUALIZAR ESTADO Y TIMERS ---
        self.distancia_ultra_is_monitoring = False # Marcar como no monitoreando ultrasónico
        self.manage_graph_timer()                   # Gestionar timer compartido inteligentemente
          # --- RESTAURAR INTERFAZ AL ESTADO INICIAL ---
        self.start_distancia_ultra_btn.setText("▶️ Iniciar Monitoreo")
        self.start_distancia_ultra_btn.setStyleSheet("QPushButton { background-color: #17a2b8; border-color: #17a2b8; color: white; padding: 10px; }")
    
    def update_distancia_ultra_data(self, distancia_cm):
        """
        Actualiza datos en tiempo real del sensor de distancia ultrasónico.
        
        PARÁMETROS:
        - distancia_cm (float): Distancia calculada en centímetros
        
        GESTIÓN DE MEMORIA:
        - Mantiene buffer circular de 500 puntos máximo
        - Elimina datos antiguos automáticamente (FIFO)
        - Optimiza memoria para sesiones largas de monitoreo
        
        ACTUALIZACIÓN DE GRÁFICAS:
        - Actualiza línea de gráfica con nuevos datos
        - Ajusta límites del eje X dinámicamente
        - Calcula límites del eje Y con margen automático
        - Optimiza rendimiento con actualización diferida
        
        INTERFAZ EN TIEMPO REAL:
        - Actualiza etiqueta de distancia formateada
        - Proporciona retroalimentación visual continua
        """
        # --- GESTIÓN DE MEMORIA CON BUFFER CIRCULAR ---
        max_points = 500  # Límite para evitar consumo excesivo de memoria
        
        self.distancia_ultra_cm.append(distancia_cm)
        
        # Mantener solo los últimos max_points (comportamiento FIFO)
        if len(self.distancia_ultra_cm) > max_points:
            self.distancia_ultra_cm.pop(0)
          # --- ACTUALIZAR ETIQUETA EN TIEMPO REAL ---
        try:
            # Verificar que el widget existe antes de actualizarlo
            if hasattr(self, 'distancia_ultra_label'):
                self.distancia_ultra_label.setText(f"Distancia: {distancia_cm:.1f} cm")
            
            # --- ACTUALIZAR GRÁFICA DE FORMA OPTIMIZADA ---
            if hasattr(self, 'line_ultra'):
                x_data = list(range(len(self.distancia_ultra_cm)))
                self.line_ultra.set_data(x_data, self.distancia_ultra_cm)
                
                # Ajustar límites del eje X dinámicamente
                if len(self.distancia_ultra_cm) > 0:
                    self.ax_ultra.set_xlim(0, max(100, len(self.distancia_ultra_cm)))
                    
                    # Ajustar límites del eje Y con margen automático
                    min_dist = min(self.distancia_ultra_cm)
                    max_dist = max(self.distancia_ultra_cm)
                    margin = (max_dist - min_dist) * 0.1 if max_dist > min_dist else 10
                    self.ax_ultra.set_ylim(max(0, min_dist - margin), max_dist + margin)
              # --- MARCAR PARA ACTUALIZACIÓN DIFERIDA ---
            self.pending_updates = True
            self.pending_distancia_ultra_data = distancia_cm
        except RuntimeError:
            # Widget fue eliminado durante la actualización
            self.distancia_ultra_is_monitoring = False
            return
    
    def clear_distancia_ultra_graph(self):
        """
        Limpia la gráfica y datos del sensor de distancia ultrasónico.
          OPERACIONES DE LIMPIEZA:
        1. Vaciar lista de datos de distancia
        2. Resetear línea de gráfica a estado vacío
        3. Restaurar límites por defecto de los ejes
        4. Forzar redibujado del canvas de matplotlib
        5. Resetear etiqueta numérica al estado inicial
        6. Mostrar confirmación al usuario
        
        CONFIGURACIÓN POR DEFECTO:
        - Eje X: 0 a 100 muestras
        - Eje Y: 0 a 400 cm (rango típico del HC-SR04)
        - Etiquetas: Mostrar "--" para indicar sin datos
        
        USO: Ideal para limpiar datos antes de una nueva sesión
        de monitoreo o cuando se requiere reiniciar mediciones.
        """        # --- LIMPIAR LISTA DE DATOS ---
        self.distancia_ultra_cm.clear()
        
        # --- RESETEAR GRÁFICA A ESTADO INICIAL ---
        if hasattr(self, 'line_ultra'):
            self.line_ultra.set_data([], [])           # Vaciar datos de la línea
            self.ax_ultra.set_xlim(0, 100)             # Restaurar límite X
            self.ax_ultra.set_ylim(0, 400)             # Restaurar límite Y (rango HC-SR04)
            self.canvas_ultra.draw()                   # Forzar redibujado
        
        # --- RESETEAR ETIQUETA NUMÉRICA ---
        if hasattr(self, 'distancia_ultra_distancia_label'):
            self.distancia_ultra_distancia_label.setText("Distancia: --")
        
        # --- CONFIRMAR OPERACIÓN AL USUARIO ---
        QMessageBox.information(self, "Gráfica Limpia", "Los datos del sensor ultrasónico han sido eliminados")
    
    def export_distancia_ultra_to_excel(self):
        """
        Exporta los datos del sensor de distancia ultrasónico a un archivo Excel.
          CONTENIDO DEL ARCHIVO:
        1. Datos tabulares: Tiempo, Distancia
        2. Gráfica integrada: Distancia vs Tiempo con formato profesional
        3. Estadísticas calculadas: Mín, Máx, Promedio de distancias
        4. Metadatos: Timestamp de exportación y configuración del sensor
        
        FORMATO DE DATOS:
        - Columna A: Tiempo estimado en segundos (basado en frecuencia de muestreo)
        - Columna B: Distancia en centímetros con 1 decimal
        
        CARACTERÍSTICAS:
        - Encabezados con formato bold y centrado
        - Gráfica de líneas insertada automáticamente
        - Cálculos estadísticos incluidos
        - Nombre de archivo con timestamp único
        - Validación de dependencias (openpyxl)
        
        ERROR HANDLING:
        - Verifica existencia de datos antes de exportar
        - Maneja errores de dependencias faltantes
        - Captura excepciones de escritura de archivos
        - Muestra mensajes informativos al usuario
        """        """Exporta los datos del sensor de distancia ultrasónico a Excel"""
        # --- VALIDAR EXISTENCIA DE DATOS ---
        if not self.distancia_ultra_cm:
            QMessageBox.warning(self, "Sin datos", "No hay datos para exportar")
            return
        
        try:
            # --- IMPORTAR DEPENDENCIAS PARA EXCEL ---
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.styles import Font, Alignment
            import os
            from datetime import datetime
            
            # --- CREAR LIBRO DE TRABAJO ---
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Ultrasónico"
              # --- CONFIGURAR ENCABEZADOS CON FORMATO ---
            headers = ["Tiempo (s)", "Distancia (cm)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)                    # Texto en negrita
                cell.alignment = Alignment(horizontal='center') # Centrar texto
            
            # --- ESCRIBIR DATOS TABULARES ---
            for i, distancia in enumerate(self.distancia_ultra_cm):
                ws.cell(row=i+2, column=1, value=i*0.1)  # Tiempo estimado (100ms por muestra)
                ws.cell(row=i+2, column=2, value=distancia)
            
            # --- CREAR GRÁFICA INTEGRADA ---
            chart = LineChart()
            chart.title = "Distancia Ultrasónica vs Tiempo"
            chart.x_axis.title = "Tiempo (s)"
            chart.y_axis.title = "Distancia (cm)"
              # Configurar datos para la gráfica
            data = Reference(ws, min_col=2, min_row=1, max_row=len(self.distancia_ultra_cm)+1)
            chart.add_data(data, titles_from_data=True)
            
            # Agregar gráfica al worksheet
            ws.add_chart(chart, "D2")
            
            # --- AGREGAR ESTADÍSTICAS CALCULADAS ---
            stats_row = len(self.distancia_ultra_cm) + 5
            ws.cell(row=stats_row, column=1, value="Estadísticas:")
            ws.cell(row=stats_row, column=1).font = Font(bold=True)
            
            ws.cell(row=stats_row+1, column=1, value="Distancia mínima:")
            ws.cell(row=stats_row+1, column=2, value=f"{min(self.distancia_ultra_cm):.2f} cm")
            
            ws.cell(row=stats_row+2, column=1, value="Distancia máxima:")
            ws.cell(row=stats_row+2, column=2, value=f"{max(self.distancia_ultra_cm):.2f} cm")
            
            ws.cell(row=stats_row+3, column=1, value="Distancia promedio:")
            ws.cell(row=stats_row+3, column=2, value=f"{sum(self.distancia_ultra_cm)/len(self.distancia_ultra_cm):.2f} cm")
            
            # --- GUARDAR ARCHIVO CON TIMESTAMP ---
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"datos_ultrasonico_{timestamp}.xlsx"
            wb.save(filename)
            
            # --- CONFIRMAR ÉXITO AL USUARIO ---
            QMessageBox.information(self, "Exportación exitosa", 
                                    f"Datos exportados a: {os.path.abspath(filename)}")
            
        except ImportError:
            # --- MANEJAR DEPENDENCIA FALTANTE ---
            QMessageBox.warning(self, "Dependencia faltante", 
                                "Se requiere 'openpyxl' para exportar a Excel.\n"
                                "Instálalo con: pip install openpyxl")
        except Exception as e:        # --- MANEJAR ERRORES GENERALES ---
            QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")
    # ============================================================================
    # MÉTODO: DETENER TODOS LOS THREADS DE MONITOREO
    # ============================================================================
    def stop_all_monitoring_threads(self):
        """
        Detiene todos los threads de monitoreo activos de forma segura
        
        Propósito: Limpiar todos los recursos cuando se cambia de sensor o se cierra la aplicación
        Threads afectados: Ángulo simple, brazo robótico, IR, capacitivo, ultrasónico
        Seguridad: Verifica existencia y estado antes de detener cada thread
        Recursos: Libera memoria y conexiones TCP con ESP32
        """
        
        # --- DETENER THREAD DE ÁNGULO SIMPLE ---
        if hasattr(self, 'angulo_thread') and self.angulo_thread and self.angulo_thread.isRunning():
            self.angulo_thread.stop()
            self.angulo_thread.wait(2000)  # Esperar máximo 2 segundos
            self.angulo_thread = None
            self.is_monitoring = False
        
        # --- DETENER THREAD DE BRAZO ROBÓTICO ---
        if hasattr(self, 'brazo_thread') and self.brazo_thread and self.brazo_thread.isRunning():
            self.brazo_thread.stop()
            self.brazo_thread.wait(2000)  # Esperar máximo 2 segundos
            self.brazo_thread = None
            self.brazo_is_monitoring = False
        
        # --- DETENER THREAD DE DISTANCIA IR ---
        if hasattr(self, 'distancia_ir_thread') and self.distancia_ir_thread and self.distancia_ir_thread.isRunning():
            self.distancia_ir_thread.stop()
            self.distancia_ir_thread.wait(2000)  # Esperar máximo 2 segundos
            self.distancia_ir_thread = None
            self.distancia_ir_is_monitoring = False
        
        # --- DETENER THREAD DE DISTANCIA CAPACITIVO ---
        if hasattr(self, 'distancia_cap_thread') and self.distancia_cap_thread and self.distancia_cap_thread.isRunning():
            self.distancia_cap_thread.stop()
            self.distancia_cap_thread.wait(2000)  # Esperar máximo 2 segundos
            self.distancia_cap_thread = None
            self.distancia_cap_is_monitoring = False
        
        # --- DETENER THREAD DE DISTANCIA ULTRASÓNICO ---
        if hasattr(self, 'distancia_ultra_thread') and self.distancia_ultra_thread and self.distancia_ultra_thread.isRunning():
            self.distancia_ultra_thread.stop()
            self.distancia_ultra_thread.wait(2000)  # Esperar máximo 2 segundos
            self.distancia_ultra_thread = None
            self.distancia_ultra_is_monitoring = False
        
        # --- GESTIONAR TIMER COMPARTIDO ---
        self.manage_graph_timer()  # Detener timer si no hay sensores activos

    # =====================================================================================
    # MÉTODO: REINICIO COMPLETO DE LA APLICACIÓN
    # =====================================================================================
    def restart_application(self):
        """
        Reinicia completamente la aplicación manteniendo la conexión ESP32
        
        Propósito: Limpiar toda la interfaz y datos como si acabara de conectar al ESP32
        Funcionamiento: Detiene todos los threads, limpia datos, reinicia la UI
        Conexión: Mantiene la conexión TCP con el ESP32 sin desconectarla
        Estado: Preserva el estado de conexión pero reinicia todo lo demás
        UI: Vuelve a la pantalla de bienvenida y resetea todos los controles
        """
        
        # --- VERIFICAR CONEXIÓN REQUERIDA ---
        if not self.is_connected:
            QMessageBox.warning(self, "Sin conexión", "No hay conexión ESP32 para reiniciar")
            return
        
        # --- CONFIRMAR REINICIO CON EL USUARIO ---
        reply = QMessageBox.question(self, "Confirmar Reinicio", 
                                   "¿Estás seguro de que quieres reiniciar la interfaz?\n\n"
                                   "Se mantendrá la conexión ESP32 pero se limpiarán todos los datos.",
                                   QMessageBox.Yes | QMessageBox.No,
                                   QMessageBox.No)
        
        if reply != QMessageBox.Yes:
            return  # Usuario canceló el reinicio
        
        # --- PASO 1: DETENER TODOS LOS THREADS ACTIVOS ---
        print("🔄 Iniciando reinicio - Deteniendo threads...")
        self.stop_all_monitoring_threads()      # Detener todos los sensores activos
        
        # --- PASO 2: LIMPIAR TODOS LOS DATOS DE SENSORES ---
        print("🧹 Limpiando datos de sensores...")
        self.clear_all_sensor_data()
        
        # --- PASO 3: RESETEAR ESTADO DE LA INTERFAZ ---
        print("🖥️ Reseteando interfaz de usuario...")
        self.reset_interface_state()
        
        # --- PASO 4: MOSTRAR PANTALLA DE BIENVENIDA ---
        print("🏠 Volviendo a pantalla de bienvenida...")
        self.show_welcome_screen()
        
        # --- CONFIRMACIÓN FINAL ---
        print("✅ Reinicio completado - Conexión ESP32 mantenida")
        QMessageBox.information(self, "Reinicio Completado", 
                              "La interfaz se ha reiniciado exitosamente.\n\n"
                              "La conexión ESP32 se mantiene activa.")

    # =====================================================================================
    # MÉTODO: LIMPIAR TODOS LOS DATOS DE SENSORES
    # =====================================================================================
    def clear_all_sensor_data(self):
        """
        Limpia todos los datos almacenados de todos los sensores
        
        Propósito: Borrar historial de lecturas y resetear gráficas
        Sensores: Ángulo simple, brazo robótico, IR, capacitivo, ultrasónico
        Memoria: Libera listas de datos para optimizar memoria
        """
        
        # --- LIMPIAR DATOS DE ÁNGULO SIMPLE ---
        self.angulos = []                        # Limpiar lista de ángulos
        self.lecturas = []                       # Limpiar lista de lecturas ADC
        
        # --- LIMPIAR DATOS DE BRAZO ROBÓTICO ---
        if hasattr(self, 'brazo_angulos1'):
            self.brazo_angulos1 = []             # Limpiar ángulos potenciómetro 1
        if hasattr(self, 'brazo_angulos2'):
            self.brazo_angulos2 = []             # Limpiar ángulos potenciómetro 2
        if hasattr(self, 'brazo_angulos3'):
            self.brazo_angulos3 = []             # Limpiar ángulos potenciómetro 3
        if hasattr(self, 'brazo_lecturas1'):
            self.brazo_lecturas1 = []            # Limpiar lecturas ADC brazo
        if hasattr(self, 'brazo_lecturas2'):
            self.brazo_lecturas2 = []
        if hasattr(self, 'brazo_lecturas3'):
            self.brazo_lecturas3 = []
        
        # --- LIMPIAR DATOS DE SENSORES DE DISTANCIA ---
        if hasattr(self, 'distancia_ultra_data'):
            self.distancia_ultra_data = []       # Limpiar datos ultrasónicos
        
        # --- LIMPIAR GRÁFICAS SI EXISTEN ---
        if hasattr(self, 'ax') and self.ax:
            self.ax.clear()                      # Limpiar gráfica principal
            self.ax.set_title('Sensor de Ángulo Simple - Tiempo Real', fontsize=14, fontweight='bold')
            self.ax.set_xlabel('Tiempo (s)', fontsize=12)
            self.ax.set_ylabel('Ángulo (grados)', fontsize=12)
            self.ax.grid(True, alpha=0.3)
            self.ax.set_facecolor('#f8f9fa')
            
        if hasattr(self, 'ax_brazo') and self.ax_brazo:
            self.ax_brazo.clear()                # Limpiar gráfica de brazo
            self.ax_brazo.set_title('Brazo Robótico - 3 Sensores + Capacitivo', fontsize=14, fontweight='bold')
            self.ax_brazo.set_xlabel('Tiempo (s)', fontsize=12)
            self.ax_brazo.set_ylabel('Ángulo (grados)', fontsize=12)
            self.ax_brazo.grid(True, alpha=0.3)
            self.ax_brazo.set_facecolor('#f8f9fa')
            
        if hasattr(self, 'ax_ultra') and self.ax_ultra:
            self.ax_ultra.clear()                # Limpiar gráfica ultrasónica
            self.ax_ultra.set_title('Sensor Ultrasónico HC-SR04 - Tiempo Real', fontsize=14, fontweight='bold')
            self.ax_ultra.set_xlabel('Tiempo (s)', fontsize=12)
            self.ax_ultra.set_ylabel('Distancia (cm)', fontsize=12)
            self.ax_ultra.grid(True, alpha=0.3)
            self.ax_ultra.set_facecolor('#f8f9fa')
          # --- REDIBUJAR GRÁFICAS VACÍAS (CON MANEJO SEGURO) ---
        try:
            if hasattr(self, 'canvas') and self.canvas:
                self.canvas.draw()
        except RuntimeError:
            # Canvas ya eliminado, ignorar error
            pass
            
        try:
            if hasattr(self, 'canvas_brazo') and self.canvas_brazo:
                self.canvas_brazo.draw()
        except RuntimeError:
            # Canvas ya eliminado, ignorar error
            pass
            
        try:
            if hasattr(self, 'canvas_ultra') and self.canvas_ultra:
                self.canvas_ultra.draw()
        except RuntimeError:
            # Canvas ya eliminado, ignorar error
            pass

    # =====================================================================================
    # MÉTODO: RESETEAR ESTADO DE LA INTERFAZ
    # =====================================================================================
    def reset_interface_state(self):
        """
        Resetea todos los estados de la interfaz a valores iniciales
        
        Propósito: Volver botones y controles a estado original
        Estado: Mantiene conexión pero resetea flags de monitoreo
        UI: Restaura textos de botones y estilos originales
        """
        
        # --- RESETEAR FLAGS DE MONITOREO ---
        self.is_monitoring = False                       # Ángulo simple no monitoreando
        self.brazo_is_monitoring = False                 # Brazo no monitoreando
        self.distancia_ir_is_monitoring = False          # IR no monitoreando
        self.distancia_cap_is_monitoring = False         # Capacitivo no monitoreando
        self.distancia_ultra_is_monitoring = False       # Ultrasónico no monitoreando
        
        # --- RESETEAR DATOS PENDIENTES ---
        self.pending_updates = False
        self.pending_simple_data = None
        self.pending_brazo_data = None
        self.pending_distancia_ir_data = None
        self.pending_distancia_cap_data = None
        self.pending_distancia_ultra_data = None
        
        # --- RESETEAR BOTONES DE EXPORTACIÓN ---
        # Los botones de exportación se habilitarán automáticamente cuando haya datos nuevos
        
        print("🔧 Estado de interfaz reseteado completamente")

    # =====================================================================================
    # MÉTODO: MOSTRAR PANTALLA DE BIENVENIDA
    # =====================================================================================
    def show_welcome_screen(self):
        """
        Muestra la pantalla de bienvenida inicial
        
        Propósito: Volver al estado inicial como si acabara de conectar
        UI: Oculta detalles de sensores y muestra mensaje de bienvenida
        Estado: Mantiene lista de sensores visible pero sin selección
        """
        
        # --- MOSTRAR PANTALLA DE BIENVENIDA ---
        if hasattr(self, 'welcome_widget') and self.welcome_widget:
            self.welcome_widget.setVisible(True)     # Mostrar mensaje de bienvenida
        
        # --- OCULTAR DETALLES DE SENSORES ---
        if hasattr(self, 'sensor_details') and self.sensor_details:
            self.sensor_details.setVisible(False)   # Ocultar área de detalles
        
        # --- DESELECCIONAR ELEMENTOS DE LA LISTA ---
        if hasattr(self, 'sensors_list') and self.sensors_list:
            self.sensors_list.clearSelection()      # Quitar selección de sensores
        
        print("🏠 Pantalla de bienvenida mostrada")
