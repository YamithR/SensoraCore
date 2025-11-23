from __future__ import annotations

import os
import time
import socket
import tempfile
from datetime import datetime
from typing import Optional, Dict

from PySide6.QtCore import QObject, Signal, Slot
from PySide6.QtWidgets import QWidget, QPushButton, QLabel, QMessageBox, QFileDialog, QLineEdit, QVBoxLayout

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


class _CnyThread(QObject):
    sig_status = Signal(str)
    sig_error = Signal(str)
    sig_sample = Signal(int, float, float)  # adc, volt, ts

    def __init__(self, ip: str, port: int = 8080, interval_ms: int = 200):
        super().__init__()
        self.ip = ip
        self.port = port
        self.interval_ms = max(50, int(interval_ms))
        self._running = False
        self._stopping = False
        self.sock: Optional[socket.socket] = None

    def start(self):
        if self._running:
            return
        from threading import Thread
        self._stopping = False
        self._running = True
        Thread(target=self._run, daemon=True).start()

    def stop(self):
        self._stopping = True
        self._running = False
        try:
            if self.sock:
                try:
                    self.sock.sendall(b"CNY_STOP\n")
                except Exception:
                    pass
                try:
                    self.sock.close()
                except Exception:
                    pass
        finally:
            self.sock = None

    def _run(self):
        try:
            self.sig_status.emit("[CNY] Conectando...")
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5)
            s.connect((self.ip, self.port))
            self.sock = s
            s.sendall(b"MODO:COLOR_CNY\n")
            resp = s.recv(128).decode(errors="ignore").strip()
            if "COLOR_CNY_OK" not in resp:
                self.sig_error.emit("[CNY] ESP32 no acepto COLOR_CNY")
                return
            s.sendall(f"CNY_START:{self.interval_ms}\n".encode())
            self.sig_status.emit("[CNY] Streaming iniciado")
            s.settimeout(1.0)
            buf = ""
            while self._running:
                try:
                    chunk = s.recv(256)
                    if not chunk:
                        break
                    buf += chunk.decode(errors="ignore")
                    while "\n" in buf:
                        line, buf = buf.split("\n", 1)
                        line = line.strip()
                        if not line:
                            continue
                        # CNY_RAW:<n>,VOLTAGE:<v>,REFLECTIVITY:<pct>,SURFACE:<type>
                        try:
                            kv: Dict[str, str] = {}
                            for part in line.replace(";", ",").split(","):
                                if ":" in part:
                                    k, v = part.split(":", 1)
                                    kv[k.strip().upper()] = v.strip()
                            # Soportar ambos formatos
                            adc = int(float(kv.get("CNY_RAW", kv.get("ADC", "0"))))
                            volt = float(kv.get("VOLTAGE", kv.get("VOLT", str((adc/4095.0)*3.3))))
                            self.sig_sample.emit(adc, volt, time.time())
                        except Exception:
                            pass
                except socket.timeout:
                    continue
                except Exception as e:
                    if not self._stopping:
                        self.sig_error.emit(f"[CNY] Error socket: {e}")
                    break
        except Exception as e:
            self.sig_error.emit(f"[CNY] Error de conexion: {e}")
        finally:
            try:
                if self.sock:
                    try:
                        self.sock.sendall(b"CNY_STOP\n")
                    except Exception:
                        pass
                    try:
                        self.sock.close()
                    except Exception:
                        pass
            finally:
                self.sock = None
                self.sig_status.emit("[CNY] Desconectado")


class ColorCNYLogic(QObject):
    def __init__(self, ui_widget: QWidget, main_window):
        super().__init__(ui_widget)
        self.ui = ui_widget
        self.main = main_window

        # Widgets
        self.btn_start = self.ui.findChild(QPushButton, "iniciarBt")
        self.btn_clear = self.ui.findChild(QPushButton, "limpiarBt")
        self.btn_export = self.ui.findChild(QPushButton, "exportarBt")
        self.lbl_adc = self.ui.findChild(QLabel, "analogoDtCny70")
        self.lbl_white = self.ui.findChild(QLabel, "porcentajeDeBlancoDt")
        self.lbl_black = self.ui.findChild(QLabel, "porcentajeDeNegroDt")

        # Plot embed
        self.figure = Figure(figsize=(5, 3), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        grap = self.ui.findChild(QWidget, "grapWid")
        if grap is not None:
            lay = grap.layout() or QVBoxLayout(grap)
            if grap.layout() is None:
                grap.setLayout(lay)
            lay.setContentsMargins(0, 0, 0, 0)
            lay.addWidget(self.canvas)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("CNY70 %Blanco / %Negro")
        self.ax.set_xlabel("Tiempo (s)")
        self.ax.set_ylabel("%")
        # Ventana de tiempo deslizante (segundos) y límites Y fijos en 0..100
        self.window_s = 25.0
        self.ax.set_ylim(0, 100)
        self.line_w, = self.ax.plot([], [], label="%Blanco", color="#007bff")
        self.line_b, = self.ax.plot([], [], label="%Negro", color="#dc3545")
        self.ax.legend(loc="upper right")

        # Estado
        self.interval_ms = 200
        self.thread = None
        self.monitoring = False
        self._t0 = None
        self.adc_min = 4095
        self.adc_max = 0
        self.series = {"t": [], "w": [], "b": [], "adc": [], "volt": []}

        # Init
        self._init_labels()
        if self.btn_start:
            self.btn_start.clicked.connect(self._toggle)
        if self.btn_clear:
            self.btn_clear.clicked.connect(self._clear)
        if self.btn_export:
            self.btn_export.clicked.connect(self._export_excel)

    def _init_labels(self):
        if self.lbl_adc:
            self.lbl_adc.setText("--")
        if self.lbl_white:
            self.lbl_white.setText("--")
        if self.lbl_black:
            self.lbl_black.setText("--")

    def _get_ip(self) -> Optional[str]:
        ip_edit = self.main.findChild(QLineEdit, "ipEdit") if self.main else None
        if not ip_edit:
            QMessageBox.critical(self.ui, "Error", "No se encontró el campo IP (ipEdit)")
            return None
        ip = ip_edit.text().strip()
        if not ip:
            QMessageBox.warning(self.ui, "Advertencia", "Ingrese la IP del ESP32.")
            return None
        return ip

    def _toggle(self):
        if not self.monitoring:
            ip = self._get_ip()
            if not ip:
                return
            self.thread = _CnyThread(ip, interval_ms=self.interval_ms)
            self.thread.sig_status.connect(lambda m: print(m))
            self.thread.sig_error.connect(lambda m: print(m))
            self.thread.sig_sample.connect(self._on_sample)
            self.thread.start()
            self.monitoring = True
            if self.btn_start:
                self.btn_start.setText("Pausar")
            self._t0 = time.time()
        else:
            self._stop()

    def _stop(self):
        try:
            if self.thread:
                self.thread.stop()
        finally:
            self.thread = None
            self.monitoring = False
            if self.btn_start:
                self.btn_start.setText("Iniciar Monitoreo")

    @Slot(int, float, float)
    def _on_sample(self, adc: int, volt: float, ts: float):
        # actualizar min/max dinámico por sesión
        if adc < self.adc_min:
            self.adc_min = adc
        if adc > self.adc_max:
            self.adc_max = adc
        rng = max(1, self.adc_max - self.adc_min)
        white = int(round((adc - self.adc_min) * 100.0 / rng))
        white = max(0, min(100, white))
        black = 100 - white

        if self.lbl_adc:
            self.lbl_adc.setText(str(adc))
        if self.lbl_white:
            self.lbl_white.setText(f"{white}%")
        if self.lbl_black:
            self.lbl_black.setText(f"{black}%")

        if self._t0 is None:
            self._t0 = ts
        t_rel = ts - self._t0
        self.series["t"].append(t_rel)
        self.series["w"].append(white)
        self.series["b"].append(black)
        self.series["adc"].append(adc)
        self.series["volt"].append(volt)

        # actualizar plot
        self.line_w.set_data(self.series["t"], self.series["w"])
        self.line_b.set_data(self.series["t"], self.series["b"])
        # Ajustar ventana deslizante en X y mantener Y en 0..100
        if t_rel <= self.window_s:
            self.ax.set_xlim(0, self.window_s)
        else:
            self.ax.set_xlim(t_rel - self.window_s, t_rel)
        # Y fijo
        self.ax.set_ylim(0, 100)
        self.canvas.draw_idle()

    def _clear(self):
        self.series = {"t": [], "w": [], "b": [], "adc": [], "volt": []}
        self.adc_min = 4095
        self.adc_max = 0
        self._t0 = None
        self._init_labels()
        self.ax.cla()
        self.ax.set_title("CNY70 %Blanco / %Negro")
        self.ax.set_xlabel("Tiempo (s)")
        self.ax.set_ylabel("%")
        self.ax.set_ylim(0, 100)
        # Reiniciar ventana X
        self.ax.set_xlim(0, self.window_s)
        self.line_w, = self.ax.plot([], [], label="%Blanco", color="#007bff")
        self.line_b, = self.ax.plot([], [], label="%Negro", color="#dc3545")
        self.ax.legend(loc="upper right")
        self.canvas.draw_idle()

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.drawing.image import Image as XLImage

            default_name = f"ColorCNY_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path, _ = QFileDialog.getSaveFileName(self.ui, "Exportar a Excel", default_name, "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            headers = ["timestamp_s", "adc", "volt", "pct_blanco", "pct_negro"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.alignment = Alignment(horizontal="center")
            for t, a, v, w, b in zip(self.series["t"], self.series["adc"], self.series["volt"], self.series["w"], self.series["b"]):
                ws.append([t, a, v, w, b])

            meta = wb.create_sheet("Metadatos")
            meta.append(["Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            meta.append(["Intervalo (ms)", self.interval_ms])
            meta.append(["Mapeo", "Dinámico por sesión (min/max)"])

            # Gráfica como imagen temporal
            tmp_png = os.path.join(tempfile.gettempdir(), f"cny_plot_{int(time.time())}.png")
            self.figure.savefig(tmp_png, dpi=150, bbox_inches="tight")
            wsimg = wb.create_sheet("Grafica")
            img = XLImage(tmp_png)
            wsimg.add_image(img, "A1")

            # Auto ancho de columnas
            for sheet in (ws, meta):
                for col in sheet.columns:
                    width = 10
                    for cell in col:
                        width = max(width, len(str(cell.value)) if cell.value is not None else 0)
                    sheet.column_dimensions[col[0].column_letter].width = min(30, width + 2)

            wb.save(path)
            try:
                os.remove(tmp_png)
            except Exception:
                pass
            QMessageBox.information(self.ui, "Exportar", f"Exportado: {path}")
        except ImportError:
            QMessageBox.critical(self.ui, "Exportar", "Falta librería openpyxl.")
        except Exception as e:
            QMessageBox.critical(self.ui, "Exportar", f"Error exportando: {e}")

    def cleanup(self):
        self._stop()
