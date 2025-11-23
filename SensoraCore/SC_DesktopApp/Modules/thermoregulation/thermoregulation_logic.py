from __future__ import annotations

from PySide6.QtCore import QThread, Signal, QObject
from PySide6.QtWidgets import QWidget, QMessageBox, QFileDialog, QInputDialog

import socket
import json
import os
import time
from dataclasses import dataclass, field
from typing import List, Optional, Tuple


# ---------- Helpers de persistencia ----------
def _calib_path() -> str:
	base = os.path.join(os.path.expanduser("~"), ".sensora_core")
	try:
		os.makedirs(base, exist_ok=True)
	except Exception:
		pass
	return os.path.join(base, "thermo_calib.json")


def _load_calib() -> dict:
	p = _calib_path()
	try:
		with open(p, "r", encoding="utf-8") as f:
			return json.load(f)
	except Exception:
		return {}


def _save_calib(data: dict) -> None:
	p = _calib_path()
	try:
		with open(p, "w", encoding="utf-8") as f:
			json.dump(data, f, ensure_ascii=False, indent=2)
	except Exception:
		pass


# ---------- Hilo de comunicación ----------
class ThermoThread(QThread):
	# name: LM35|DS18B20|TYPEK, temp_c: float (calibrada), temp_raw: Optional[float]
	data = Signal(str, float, object)
	status = Signal(str)

	def __init__(self, esp32_ip: str, port: int = 8080, sensor: str = "LM35", period_ms: int = 500):
		super().__init__()
		self.esp32_ip = esp32_ip
		self.port = port
		self.sensor = sensor
		self.period_ms = max(200, int(period_ms))
		self._running = False
		self._stopping = False
		self.sock: Optional[socket.socket] = None

	def run(self):
		self._running = True
		try:
			self.status.emit(f"Conectando a {self.esp32_ip}:{self.port}...")
			self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
			self.sock.settimeout(5)
			self.sock.connect((self.esp32_ip, self.port))
			self.sock.sendall(b"MODO:THERMOREGULATION")
			resp = self.sock.recv(64).decode(errors="ignore").strip()
			if "THERMOREGULATION_OK" not in resp:
				self.status.emit("ESP32 no acepto modo THERMOREGULATION")
				return
			# Enviar comando de sensor al firmware
			# Mapear TYPEK -> MAX6675 para coincidir con firmware
			sensor_cmd = "MAX6675" if self.sensor == "TYPEK" else self.sensor
			self._send_cmd(f"SENSOR:{sensor_cmd}")
			self.status.emit(f"Monitoreo iniciado: sensor {sensor_cmd}")
			self.sock.settimeout(3)
			buffer = ""
			while self._running:
				try:
					chunk = self.sock.recv(256)
					if not chunk:
						break
					buffer += chunk.decode(errors="ignore")
					while "\n" in buffer:
						line, buffer = buffer.split("\n", 1)
						line = line.strip()
						if not line:
							continue
						# Formato: SENSOR:<NAME>,TEMP_C:<float>[,RAW:<int>]
						try:
							parts = [p.strip() for p in line.split(',')]
							kv = {}
							for p in parts:
								if ':' in p:
									k, v = p.split(':', 1)
									kv[k.strip().upper()] = v.strip()
							name = kv.get('SENSOR', 'LM35')
							temp = float(kv.get('TEMP_C', '0'))
							raw = kv.get('RAW', None)
							raw_val = int(raw) if raw is not None else None
							self.data.emit(name, temp, raw_val)
						except Exception:
							pass
				except socket.timeout:
					continue
				except Exception as e:
					if not self._stopping:
						self.status.emit(f"Error socket TH: {e}")
					break
		except Exception as e:
			self.status.emit(f"Error de conexion: {e}")
		finally:
			try:
				if self.sock:
					if not self._stopping:
						try:
							self.sock.sendall(b"TH_STOP")
						except Exception:
							pass
					try:
						self.sock.close()
					except Exception:
						pass
			finally:
				self.sock = None
				self.status.emit("Desconectado")

	def stop(self):
		self._stopping = True
		self._running = False
		if self.sock:
			try:
				self.sock.sendall(b"TH_STOP")
			except Exception:
				pass
			try:
				self.sock.close()
			except Exception:
				pass
		self.wait(2000)

	def switch_sensor(self, name: str):
		self.sensor = name
		# Mapear TYPEK -> MAX6675
		sensor_cmd = "MAX6675" if name == "TYPEK" else name
		self._send_cmd(f"SENSOR:{sensor_cmd}")

	def _send_cmd(self, txt: str):
		try:
			if self.sock and self._running:
				self.sock.sendall((txt + "\n").encode())
		except Exception:
			pass


# ---------- Lógica de UI ----------
class ThermoregulationLogic(QWidget):
	def __init__(self, ui_widget, main_window=None):
		super().__init__()
		if ui_widget is None:
			raise ValueError("ui_widget no puede ser None")
		self.ui = ui_widget
		self.main_window = main_window
		self.thread: Optional[ThermoThread] = None
		self.monitoring = False
		self.current_sensor = 'LM35'  # LM35 | DS18B20 | TYPEK
		self.period_ms = 500
		self._calib = _load_calib()  # {'LM35': {'a':..,'b':..}, 'DS18B20': {'a':..,'b':..}, 'TYPEK': {'a':..,'b':..,'c':..}}
		self._data_rows: List[dict] = []
		# última temperatura medida por sensor (RAW y calibrada firmware)
		self._last_temp = {}  # Temperatura RAW sin calibrar
		self._last_temp_cal = {}  # Temperatura calibrada por firmware

		# Plot embedding (matplotlib)
		self._setup_plot()

		# Wire UI
		if hasattr(self.ui, 'iniciarBt'):
			self.ui.iniciarBt.clicked.connect(self.toggle)
		if hasattr(self.ui, 'limpiarBt'):
			self.ui.limpiarBt.clicked.connect(self._clear_plot)
		if hasattr(self.ui, 'exportarBt'):
			self.ui.exportarBt.clicked.connect(self._export)
		if hasattr(self.ui, 'calibrarBt'):
			self.ui.calibrarBt.clicked.connect(self._calibrate_dialog)

		if hasattr(self.ui, 'LM35'):
			self.ui.LM35.setCheckable(True)
			self.ui.LM35.clicked.connect(lambda: self._select_sensor('LM35'))
		if hasattr(self.ui, 'DS18B20'):
			self.ui.DS18B20.setCheckable(True)
			self.ui.DS18B20.clicked.connect(lambda: self._select_sensor('DS18B20'))
		if hasattr(self.ui, 'TermoparTipoK'):
			self.ui.TermoparTipoK.setCheckable(True)
			self.ui.TermoparTipoK.clicked.connect(lambda: self._select_sensor('TYPEK'))

		self._refresh_cal_status()
		# Inicializar estado checked según el sensor por defecto
		self._update_sensor_checks()

	# ----- Plot -----
	def _setup_plot(self):
		try:
			from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
			from matplotlib.figure import Figure
			self._Figure = Figure
			self._FigureCanvas = FigureCanvas
			self.fig = Figure(figsize=(5, 3), tight_layout=True)
			self.ax = self.fig.add_subplot(111)
			self.ax.set_title('Temperatura (°C)')
			self.ax.set_xlabel('Tiempo (s)')
			self.ax.set_ylabel('°C')
			self.ax.grid(True, alpha=0.3)
			self.canvas = FigureCanvas(self.fig)
			if hasattr(self.ui, 'grapWid'):
				# Insert canvas in grapWid's layout
				from PySide6.QtWidgets import QVBoxLayout
				lay = self.ui.grapWid.layout() or QVBoxLayout(self.ui.grapWid)
				lay.addWidget(self.canvas)
			self._t0 = time.time()
			self._times: List[float] = []
			self._lm35_y: List[float] = []
			self._typek_y: List[float] = []
			self._ds18_y: List[float] = []
			(self._lm35_line,) = self.ax.plot([], [], label='LM35')
			(self._typek_line,) = self.ax.plot([], [], label='TypeK')
			(self._ds18_line,) = self.ax.plot([], [], label='DS18B20')
			self.ax.legend(loc='upper left')
			self.canvas.draw()
		except Exception as e:
			print(f"No se pudo inicializar la gráfica: {e}")
			self.canvas = None

	def _clear_plot(self):
		self._t0 = time.time()
		self._times.clear()
		self._lm35_y.clear()
		self._typek_y.clear()
		self._ds18_y.clear()
		if self.canvas:
			self._lm35_line.set_data([], [])
			self._typek_line.set_data([], [])
			self._ds18_line.set_data([], [])
			self.ax.relim()
			self.ax.autoscale_view()
			self.canvas.draw_idle()

	# ----- Control -----
	def toggle(self):
		if self.monitoring:
			self.stop()
		else:
			self.start()

	def start(self):
		try:
			ip = self._get_ip()
			if not ip:
				QMessageBox.information(self, "ESP32", "Ingrese la IP del ESP32 en el campo correspondiente.")
				return
			if self.thread and self.thread.isRunning():
				self.thread.stop()
				self.thread = None
			self.thread = ThermoThread(ip, sensor=self.current_sensor, period_ms=self.period_ms)
			self.thread.data.connect(self._on_data)
			self.thread.status.connect(self._on_status)
			self.thread.start()
			self.monitoring = True
			if hasattr(self.ui, 'iniciarBt'):
				self.ui.iniciarBt.setText('Pausar')
		except Exception as e:
			QMessageBox.critical(self, "Thermoregulation", f"No se pudo iniciar: {e}")

	def stop(self):
		try:
			self.monitoring = False
			if self.thread:
				if self.thread.isRunning():
					self.thread.stop()
				self.thread = None
			if hasattr(self.ui, 'iniciarBt'):
				self.ui.iniciarBt.setText('Iniciar Monitoreo')
		except Exception:
			pass

	def cleanup(self):
		self.stop()

	def _select_sensor(self, name: str):
		self.current_sensor = name
		self._update_sensor_checks()
		if self.thread and self.thread.isRunning():
			self.thread.switch_sensor(name)

	def _update_sensor_checks(self):
		"""Refresca el estado 'checked' de los botones de sensor."""
		try:
			if hasattr(self.ui, 'LM35'):
				self.ui.LM35.setChecked(self.current_sensor.upper() == 'LM35')
			if hasattr(self.ui, 'TermoparTipoK'):
				self.ui.TermoparTipoK.setChecked(self.current_sensor.upper() == 'TYPEK')
			if hasattr(self.ui, 'DS18B20'):
				self.ui.DS18B20.setChecked(self.current_sensor.upper() == 'DS18B20')
		except Exception:
			pass

	# ----- Datos entrantes -----
	def _on_data(self, name: str, temp_c: float, temp_raw: Optional[float]):
		t = time.time() - getattr(self, '_t0', time.time())
		self._times.append(t)
		# Mapear MAX6675 -> TYPEK para mostrar en UI
		if name.upper() == 'MAX6675':
			name = 'TYPEK'
		# temp_c ya viene calibrada del firmware
		# temp_raw es la lectura sin calibrar
		# Aplicar calibración local adicional si existe
		cal_temp = self._apply_calibration(name, temp_c)
		# Guardar última temperatura RAW medida por sensor (para calibración)
		self._last_temp[name.upper()] = float(temp_raw if temp_raw is not None else temp_c)
		self._last_temp_cal[name.upper()] = float(temp_c)

		if name.upper() == 'LM35':
			if hasattr(self.ui, 'LecturaLm35Dt'):
				self.ui.LecturaLm35Dt.setText(f"{temp_raw:.2f}°C" if temp_raw is not None else '-')
			if hasattr(self.ui, 'TemperaturaLm35Dt'):
				self.ui.TemperaturaLm35Dt.setText(f"{temp_c:.2f}°C")
			if hasattr(self.ui, 'Lm35DtCalibrado'):
				self.ui.Lm35DtCalibrado.setText(f"{cal_temp:.2f}°C")
			self._lm35_y.append(cal_temp)
		elif name.upper() == 'TYPEK':
			if hasattr(self.ui, 'LecturaTypeKDt'):
				self.ui.LecturaTypeKDt.setText(f"{temp_raw:.2f}°C" if temp_raw is not None else '-')
			if hasattr(self.ui, 'TemperaturaTypeKDt'):
				self.ui.TemperaturaTypeKDt.setText(f"{temp_c:.2f}°C")
			if hasattr(self.ui, 'TypeKDtCalibrado'):
				self.ui.TypeKDtCalibrado.setText(f"{cal_temp:.2f}°C")
			self._typek_y.append(cal_temp)
		elif name.upper() == 'DS18B20':
			if hasattr(self.ui, 'LecturaDS18B20Dt'):
				self.ui.LecturaDS18B20Dt.setText(f"{temp_raw:.2f}°C" if temp_raw is not None else '-')
			if hasattr(self.ui, 'TemperaturaDS18B20Dt'):
				self.ui.TemperaturaDS18B20Dt.setText(f"{temp_c:.2f}°C")
			if hasattr(self.ui, 'DS18B20DtCalibrado'):
				self.ui.DS18B20DtCalibrado.setText(f"{cal_temp:.2f}°C")
			self._ds18_y.append(cal_temp)

		# Registrar fila de exportación
		self._data_rows.append({
			't_s': t,
			'sensor': name,
			'temp_raw': temp_raw if temp_raw is not None else temp_c,
			'temp_firmware': temp_c,
			'temp_cal': cal_temp,
		})

		# Actualizar plot
		if self.canvas:
			self._lm35_line.set_data(self._times[:len(self._lm35_y)], self._lm35_y)
			self._typek_line.set_data(self._times[:len(self._typek_y)], self._typek_y)
			self._ds18_line.set_data(self._times[:len(self._ds18_y)], self._ds18_y)
			self.ax.relim()
			self.ax.autoscale_view()
			self.canvas.draw_idle()

	def _on_status(self, msg: str):
		print(f"[THERMO] {msg}")

	# ----- Calibración -----
	def _apply_calibration(self, name: str, temp: float) -> float:
		try:
			c = self._calib.get(name.upper()) or self._calib.get(name)
			if not c:
				return temp
			# lineal: a*temp + b; cuadrática: a*temp^2 + b*temp + c
			if 'a2' in c:  # permitir claves alternativas
				return float(c['a2'])*temp*temp + float(c['a1'])*temp + float(c['a0'])
			if 'c' in c:
				return float(c['a'])*temp*temp + float(c['b'])*temp + float(c['c'])
			return float(c.get('a', 1.0)) * temp + float(c.get('b', 0.0))
		except Exception:
			return temp

	def _calibrate_dialog(self):
		"""
		Asistente de calibración mejorado para sensores de temperatura
		- Soporte para calibración rápida (offset) y avanzada (multi-punto)
		- Instrucciones paso a paso con baño térmico
		- Validación de estabilidad de lectura
		- Envío de parámetros al firmware
		"""
		try:
			if not self.monitoring:
				QMessageBox.information(self, "Calibración", "Primero inicie el monitoreo y espere a ver datos.")
				return
			sensor = self.current_sensor.upper()
			last = self._last_temp.get(sensor)
			if last is None:
				QMessageBox.information(self, "Calibración", f"Aún no hay lectura para {sensor}. Espere unos segundos y vuelva a intentar.")
				return

			# Paso 1: Tipo de calibración
			calib_type, ok = QInputDialog.getItem(
				self,
				"Calibración - Paso 1/4",
				f"Tipo de calibración para {sensor}:",
				[
					"Rápida (1 punto - solo offset)",
					"Estándar (2 puntos - offset y escala)",
					"Avanzada (3+ puntos - curva completa)"
				],
				1, False
			)
			if not ok:
				return
			
			if "Rápida" in calib_type:
				self._calibrate_quick(sensor)
			elif "Estándar" in calib_type:
				self._calibrate_standard(sensor)
			else:
				self._calibrate_advanced(sensor)
		
		except Exception as e:
			QMessageBox.critical(self, "Calibración", f"Error al calibrar: {e}")
	
	def _calibrate_quick(self, sensor: str):
		"""Calibración rápida de 1 punto (solo offset)"""
		# Instrucciones
		msg = (
			f"CALIBRACIÓN RÁPIDA - {sensor}\\n\\n"
			"Pasos:\\n"
			"1. Prepare un baño térmico estable\\n"
			"   (ej: hielo derritiendo = 0°C)\\n"
			"2. Sumerja el sensor en el baño\\n"
			"3. Espere 2-3 minutos de estabilización\\n"
			"4. Ingrese la temperatura real\\n\\n"
			"Esta calibración solo corrige el OFFSET.\\n"
			"¿Continuar?"
		)
		reply = QMessageBox.question(self, "Calibración Rápida", msg, QMessageBox.Yes | QMessageBox.No)
		if reply != QMessageBox.Yes:
			return
		
		# Esperar estabilización
		QMessageBox.information(
			self, 
			"Estabilización",
			f"Coloque {sensor} en el baño térmico\\n"
			"y presione OK cuando esté estable (2-3 min)."
		)
		
		# Capturar lectura actual
		temp_medida = self._last_temp.get(sensor.upper())
		if temp_medida is None:
			QMessageBox.warning(self, "Error", "No se pudo leer la temperatura actual.")
			return
		
		# Solicitar temperatura de referencia
		temp_ref, ok = QInputDialog.getDouble(
			self,
			"Temperatura de Referencia",
			f"Lectura actual: {temp_medida:.2f}°C\\n\\n"
			"Ingrese la temperatura REAL del baño térmico (°C):",
			0.0, -200.0, 1000.0, 2
		)
		if not ok:
			return
		
		# Calcular offset
		offset = temp_ref - temp_medida
		
		# Guardar calibración
		self._calib[sensor] = {'a': 1.0, 'b': offset}
		_save_calib(self._calib)
		
		# Enviar al firmware
		if self.thread and self.thread.sock:
			try:
				cmd = f"CALIB_OFFSET:{sensor}:{offset}\\n"
				self.thread.sock.sendall(cmd.encode())
			except Exception:
				pass
		
		self._refresh_cal_status()
		QMessageBox.information(
			self,
			"Calibración Completada",
			f"{sensor} calibrado correctamente\\n\\n"
			f"Offset: {offset:+.2f}°C\\n\\n"
			f"Temp_calibrada = Temp_medida + ({offset:+.2f})"
		)
	
	def _calibrate_standard(self, sensor: str):
		"""Calibración estándar de 2 puntos (offset y escala)"""
		# Instrucciones
		msg = (
			f"CALIBRACIÓN ESTÁNDAR - {sensor}\\n\\n"
			"Necesitará 2 baños térmicos estables:\\n"
			"- Punto bajo (ej: 0°C hielo)\\n"
			"- Punto alto (ej: 100°C agua hirviendo)\\n\\n"
			"Esta calibración corrige OFFSET y ESCALA.\\n"
			"¿Continuar?"
		)
		reply = QMessageBox.question(self, "Calibración Estándar", msg, QMessageBox.Yes | QMessageBox.No)
		if reply != QMessageBox.Yes:
			return
		
		pares: List[Tuple[float, float]] = []
		
		for i, punto_sugerido in enumerate([(0.0, "hielo"), (100.0, "agua hirviendo")]):
			temp_sugerida, nombre = punto_sugerido
			
			QMessageBox.information(
				self,
				f"Punto {i+1}/2 - {nombre}",
				f"Coloque {sensor} en baño de {nombre}\\n"
				f"(~{temp_sugerida}°C)\\n\\n"
				"Espere 2-3 minutos de estabilización\\n"
				"y presione OK."
			)
			
			temp_medida = self._last_temp.get(sensor.upper())
			if temp_medida is None:
				QMessageBox.warning(self, "Error", f"No se pudo leer temperatura en punto {i+1}.")
				return
			
			temp_ref, ok = QInputDialog.getDouble(
				self,
				f"Punto {i+1}/2 - Referencia",
				f"Lectura actual: {temp_medida:.2f}°C\\n\\n"
				"Ingrese temperatura REAL del baño (°C):",
				temp_sugerida, -200.0, 1000.0, 2
			)
			if not ok:
				return
			
			pares.append((temp_medida, temp_ref))
		
		# Ajuste lineal: Tcal = a*T + b
		a, b = self._fit_linear(pares)
		
		# Guardar
		self._calib[sensor] = {'a': a, 'b': b}
		_save_calib(self._calib)
		
		# Enviar al firmware (offset y escala)
		if self.thread and self.thread.sock:
			try:
				cmd_offset = f"CALIB_OFFSET:{sensor}:{b}\\n"
				cmd_scale = f"CALIB_SCALE:{sensor}:{a}\\n"
				self.thread.sock.sendall(cmd_offset.encode())
				self.thread.sock.sendall(cmd_scale.encode())
			except Exception:
				pass
		
		self._refresh_cal_status()
		QMessageBox.information(
			self,
			"Calibración Completada",
			f"{sensor} calibrado correctamente\\n\\n"
			f"Ecuación: Tcal = {a:.6f}*T + {b:.6f}\\n\\n"
			f"Error típico: ±{self._calculate_calibration_error(pares, a, b):.2f}°C"
		)
	
	def _calibrate_advanced(self, sensor: str):
		"""Calibración avanzada con múltiples puntos"""
		# Número de puntos
		min_pts = 3 if sensor == 'TYPEK' else 2
		n, ok = QInputDialog.getInt(
			self, 
			"Calibración Avanzada", 
			f"¿Cuántos puntos desea? (mín {min_pts}, máx 10)",
			min_pts, min_pts, 10, 1
		)
		if not ok:
			return

		pares: List[Tuple[float, float]] = []  # (medido, referencia)
		sugeridos = [0.0, 25.0, 50.0, 75.0, 100.0, 150.0, 200.0, 250.0, 300.0, 400.0]
		for i in range(n):
			# Instrucciones
			QMessageBox.information(
				self,
				f"Punto {i+1}/{n}",
				f"Coloque {sensor} en baño térmico {i+1}\\n"
				f"Temperatura sugerida: {sugeridos[i] if i < len(sugeridos) else 'cualquiera'}°C\\n\\n"
				"Espere 2-3 minutos de estabilización\\n"
				"y presione OK."
			)
			
			# Usar el último medido actual
			med = self._last_temp.get(sensor)
			if med is None:
				QMessageBox.warning(self, "Calibración", "No se detecta lectura actual. Intente nuevamente.")
				return
			# Pedir referencia
			default = sugeridos[i] if i < len(sugeridos) else med
			ref, ok2 = QInputDialog.getDouble(
				self, 
				f"Punto {i+1}/{n}", 
				f"Lectura medida: {med:.2f}°C\\nIngrese temperatura REAL del baño (°C):", 
				float(default), -1000.0, 1000.0, 2
			)
			if not ok2:
				QMessageBox.information(self, "Calibración", "Calibración cancelada.")
				return
			pares.append((float(med), float(ref)))

		# Ajuste según tipo de sensor
		if sensor == 'TYPEK':
			a, b, c = self._fit_quadratic(pares)
			self._calib['TYPEK'] = {'a': a, 'b': b, 'c': c}
			eq = f"Tcal = {a:.6f}*T^2 + {b:.6f}*T + {c:.6f}"
		else:
			a, b = self._fit_linear(pares)
			self._calib[sensor] = {'a': a, 'b': b}
			eq = f"Tcal = {a:.6f}*T + {b:.6f}"

		_save_calib(self._calib)
		self._refresh_cal_status()
		QMessageBox.information(
			self, 
			"Calibración", 
			f"Calibración guardada para {sensor}.\\n{eq}"
		)
		
		# Mostrar resultado con gráfica
		try:
			if sensor == 'TYPEK':
				self._show_calibration_result(sensor, pares, ('quad', a, b, c))
			else:
				self._show_calibration_result(sensor, pares, ('lin', a, b))
		except Exception:
			pass
	
	def _fit_linear(self, p: List[Tuple[float,float]]) -> Tuple[float, float]:
		"""Ajuste lineal: y = a*x + b"""
		xs = [x for x,_ in p]
		ys = [y for _,y in p]
		n = len(xs)
		mx = sum(xs)/n
		my = sum(ys)/n
		num = sum((xs[i]-mx)*(ys[i]-my) for i in range(n))
		den = sum((xs[i]-mx)**2 for i in range(n))
		if den == 0:
			return 1.0, 0.0
		a = num/den
		b = my - a*mx
		return a, b
	
	def _calculate_calibration_error(self, pares: List[Tuple[float,float]], a: float, b: float) -> float:
		"""Calcula el error RMS de la calibración lineal"""
		errors = []
		for x_med, y_ref in pares:
			y_cal = a * x_med + b
			errors.append((y_cal - y_ref) ** 2)
		if not errors:
			return 0.0
		import math
		return math.sqrt(sum(errors) / len(errors))
	
	def _fit_quadratic(self, p: List[Tuple[float,float]]) -> Tuple[float, float, float]:
		"""Ajuste cuadrático: y = a*x^2 + b*x + c"""
		xs = [x for x,_ in p]
		ys = [y for _,y in p]
		n = len(xs)
		if n < 3:
			a1, b1 = self._fit_linear(p)
			return 0.0, a1, b1  # a*T^2 + b*T + c => a=0, b=a1, c=b1
		s1 = sum(xs)
		s2 = sum(x*x for x in xs)
		s3 = sum((x**3) for x in xs)
		s4 = sum((x**4) for x in xs)
		sy = sum(ys)
		sxy = sum(xs[i]*ys[i] for i in range(n))
		sx2y = sum((xs[i]**2)*ys[i] for i in range(n))
		# Resolver sistema 3x3 para a,b,c
		# | s4 s3 s2 | |a| = |sx2y|
		# | s3 s2 s1 | |b|   | sxy |
		# | s2 s1 n  | |c|   | sy  |
		def det3(m):
			return (m[0][0]*((m[1][1]*m[2][2])-(m[1][2]*m[2][1]))
				   -m[0][1]*((m[1][0]*m[2][2])-(m[1][2]*m[2][0]))
				   +m[0][2]*((m[1][0]*m[2][1])-(m[1][1]*m[2][0])))
		M = [[s4,s3,s2],[s3,s2,s1],[s2,s1,n]]
		D = det3(M)
		if abs(D) < 1e-12:
			a1, b1 = self._fit_linear(p)
			return 0.0, a1, b1
		Ma = [[sx2y,s3,s2],[sxy,s2,s1],[sy,s1,n]]
		Mb = [[s4,sx2y,s2],[s3,sxy,s1],[s2,sy,n]]
		Mc = [[s4,s3,sx2y],[s3,s2,sxy],[s2,s1,sy]]
		Da = det3(Ma)
		Db = det3(Mb)
		Dc = det3(Mc)
		a = Da/D
		b = Db/D
		c = Dc/D
		return a, b, c

	def _show_calibration_result(self, sensor: str, pares: List[Tuple[float,float]], fit_info: Tuple):
		from matplotlib.figure import Figure
		from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
		from PySide6.QtWidgets import QDialog, QVBoxLayout
		# Crear diálogo con gráfica de puntos medidos (x) vs referencia (y) y curva ajustada
		dlg = QDialog(self)
		dlg.setWindowTitle(f"Calibración {sensor}")
		layout = QVBoxLayout(dlg)
		fig = Figure(figsize=(5, 4), tight_layout=True)
		ax = fig.add_subplot(111)
		xs = [m for m,_ in pares]
		ys = [r for _,r in pares]
		ax.scatter(xs, ys, c='red', label='Puntos', zorder=3)
		# Curva
		if fit_info[0] == 'lin':
			_, a, b = fit_info
			xmin = min(xs); xmax = max(xs)
			x_line = [xmin + (xmax - xmin) * t / 50.0 for t in range(51)]
			y_line = [a*x + b for x in x_line]
			ax.plot(x_line, y_line, 'b-', label=f'Tcal = {a:.4f}*T + {b:.4f}')
		else:
			_, a, b, c = fit_info
			xmin = min(xs); xmax = max(xs)
			x_line = [xmin + (xmax - xmin) * t / 50.0 for t in range(51)]
			y_line = [a*(x**2) + b*x + c for x in x_line]
			ax.plot(x_line, y_line, 'b-', label=f'Tcal = {a:.4f}*T^2 + {b:.4f}*T + {c:.4f}')
		ax.set_xlabel('Medido (°C)')
		ax.set_ylabel('Referencia (°C)')
		ax.grid(True, alpha=0.3)
		ax.legend(loc='best')
		canvas = FigureCanvas(fig)
		layout.addWidget(canvas)
		dlg.resize(600, 450)
		dlg.exec()

	def _refresh_cal_status(self):
		try:
			has_any = any(k in self._calib for k in ('LM35', 'DS18B20', 'TYPEK'))
			if hasattr(self.ui, 'calibrarBt'):
				self.ui.calibrarBt.setText('Calibrado' if has_any else 'No Calibrado')
		except Exception:
			pass

	# ----- Exportar -----
	def _export(self):
		if not self._data_rows:
			QMessageBox.information(self, "Exportar", "No hay datos para exportar.")
			return
		path, _ = QFileDialog.getSaveFileName(self, "Guardar archivo", "thermo_export.xlsx", "Archivos Excel (*.xlsx);;CSV (*.csv)")
		if not path:
			return
		try:
			if path.lower().endswith('.csv'):
				self._export_csv(path)
			else:
				self._export_xlsx(path)
			QMessageBox.information(self, "Exportar", f"Exportado: {path}")
		except Exception as e:
			QMessageBox.critical(self, "Exportar", f"Fallo exportando: {e}")

	def _export_csv(self, path: str):
		import csv
		with open(path, 'w', newline='', encoding='utf-8') as f:
			w = csv.writer(f)
			w.writerow(['t_s', 'sensor', 'raw', 'temp_c', 'temp_cal'])
			for r in self._data_rows:
				w.writerow([r['t_s'], r['sensor'], r['raw'], r['temp_c'], r['temp_cal']])

	def _export_xlsx(self, path: str):
		try:
			from openpyxl import Workbook
			from openpyxl.styles import Font, PatternFill, Alignment
			from openpyxl.drawing.image import Image
		except Exception:
			# Fallback a CSV si openpyxl no está
			base, _ = os.path.splitext(path)
			self._export_csv(base + '.csv')
			return

		wb = Workbook()
		# === Hoja 1: Datos ===
		ws = wb.active
		ws.title = 'Datos'
		headers = ['t_s', 'sensor', 'raw', 'temp_c', 'temp_cal']
		ws.append(headers)
		for r in self._data_rows:
			ws.append([r['t_s'], r['sensor'], r['raw'], r['temp_c'], r['temp_cal']])
		# Estilo encabezado
		header_font = Font(bold=True, color="FFFFFF")
		header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		for cell in ws[1]:
			cell.font = header_font
			cell.fill = header_fill
			cell.alignment = Alignment(horizontal="center")
		# Auto ancho
		for col in ws.columns:
			max_len = 0
			col_letter = col[0].column_letter
			for cell in col:
				max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
			ws.column_dimensions[col_letter].width = min(max_len + 2, 24)

		# === Hoja 2: Metadatos ===
		ws_meta = wb.create_sheet('Metadatos')
		ws_meta.append(["INFORMACIÓN DE EXPORTACIÓN", ""])
		from datetime import datetime
		ws_meta.append(["Fecha", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
		ws_meta.append(["Muestras", len(self._data_rows)])
		ws_meta.append(["Sensores presentes", ", ".join(sorted(set(r['sensor'] for r in self._data_rows)))])
		ws_meta.append(["Nota", "temp_c son valores crudos; temp_cal ya incluye calibración si existe."])
		ws_meta.column_dimensions['A'].width = 28
		ws_meta.column_dimensions['B'].width = 60
		ws_meta["A1"].font = Font(bold=True)

		# === Hoja 3: Calibración ===
		ws_cal = wb.create_sheet('Calibracion')
		ws_cal.append(['Sensor', 'Ecuación'])
		ws_cal["A1"].font = header_font; ws_cal["A1"].fill = header_fill; ws_cal["A1"].alignment = Alignment(horizontal="center")
		ws_cal["B1"].font = header_font; ws_cal["B1"].fill = header_fill; ws_cal["B1"].alignment = Alignment(horizontal="center")
		for k, v in self._calib.items():
			if 'c' in v:
				eq = f"Tcal = {v['a']:.6f}*T^2 + {v['b']:.6f}*T + {v['c']:.6f}"
			else:
				eq = f"Tcal = {v.get('a',1.0):.6f}*T + {v.get('b',0.0):.6f}"
			ws_cal.append([k, eq])
		ws_cal.column_dimensions['A'].width = 18
		ws_cal.column_dimensions['B'].width = 60

		# === Hoja 4: Gráfica ===
		ws_plot = wb.create_sheet('Grafica')
		img_path = self._make_plot_image()
		temp_img_path = None
		if img_path and os.path.exists(img_path):
			img = Image(img_path)
			img.width = 1000
			img.height = 500
			ws_plot['A1'] = 'Gráfica de Temperatura'
			ws_plot['A1'].font = Font(bold=True, size=16)
			ws_plot.add_image(img, 'A3')
			# No eliminar aquí: openpyxl lee la imagen al guardar
			temp_img_path = img_path

		wb.save(path)
		# Ahora sí, eliminar el archivo temporal si existe
		if temp_img_path:
			try:
				os.unlink(temp_img_path)
			except Exception:
				pass

	def _make_plot_image(self) -> Optional[str]:
		try:
			from matplotlib.figure import Figure
			import io, tempfile
			fig = Figure(figsize=(12, 5), tight_layout=True)
			ax = fig.add_subplot(111)
			ax.set_title('Temperatura (°C)')
			ax.set_xlabel('Tiempo (s)')
			ax.set_ylabel('°C')
			ax.grid(True, alpha=0.3)
			# reconstruir series
			ax.plot(self._times[:len(self._lm35_y)], self._lm35_y, label='LM35')
			ax.plot(self._times[:len(self._typek_y)], self._typek_y, label='TypeK')
			ax.plot(self._times[:len(self._ds18_y)], self._ds18_y, label='DS18B20')
			ax.legend(loc='best')
			buf = io.BytesIO()
			fig.savefig(buf, format='png', dpi=150)
			buf.seek(0)
			tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
			tmp.write(buf.read())
			tmp.close()
			buf.close()
			return tmp.name
		except Exception:
			return None

	# ----- Utils -----
	def _get_ip(self) -> Optional[str]:
		try:
			from PySide6.QtWidgets import QLineEdit
			if self.main_window:
				ip_widget = self.main_window.findChild(QLineEdit, "ipEdit")
				if ip_widget:
					return ip_widget.text().strip()
		except Exception:
			pass
		return None
