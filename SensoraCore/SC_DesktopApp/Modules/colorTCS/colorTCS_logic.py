from __future__ import annotations

import os
import time
import socket
import json
import tempfile
from datetime import datetime
from typing import Optional, Dict

from PySide6.QtCore import QObject, Signal, Slot
from PySide6.QtWidgets import QWidget, QPushButton, QLabel, QMessageBox, QFileDialog, QLineEdit, QVBoxLayout

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure


class _TcsThread(QObject):
    sig_status = Signal(str)
    sig_error = Signal(str)
    sig_sample = Signal(float, float, float, float)  # r_hz, g_hz, b_hz, ts

    def __init__(self, ip: str, port: int = 8080, interval_ms: int = 300):
        super().__init__()
        self.ip = ip
        self.port = port
        self.interval_ms = max(150, int(interval_ms))
        self._running = False
        self._stopping = False
        self.sock: Optional[socket.socket] = None

    def start(self):
        if self._running:
            return
        from threading import Thread
        self._stopping = False
        self._running = True
        Thread(target=self._run, daemon=True).start()

    def stop(self):
        self._stopping = True
        self._running = False
        try:
            if self.sock:
                try:
                    self.sock.sendall(b"TCS_STOP\n")
                except Exception:
                    pass
                try:
                    self.sock.close()
                except Exception:
                    pass
        finally:
            self.sock = None

    def _run(self):
        try:
            self.sig_status.emit("[TCS] Conectando...")
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5)
            s.connect((self.ip, self.port))
            self.sock = s
            s.sendall(b"MODO:COLOR_TCS\n")
            resp = s.recv(128).decode(errors="ignore").strip()
            if "COLOR_TCS_OK" not in resp:
                self.sig_error.emit("[TCS] ESP32 no acepto COLOR_TCS")
                return
            s.sendall(f"TCS_START:{self.interval_ms}\n".encode())
            self.sig_status.emit("[TCS] Streaming iniciado")
            s.settimeout(1.0)
            buf = ""
            while self._running:
                try:
                    chunk = s.recv(256)
                    if not chunk:
                        break
                    buf += chunk.decode(errors="ignore")
                    while "\n" in buf:
                        line, buf = buf.split("\n", 1)
                        line = line.strip()
                        if not line:
                            continue
                        # TCS_RED:<rgb>,TCS_GREEN:<rgb>,TCS_BLUE:<rgb>,FREQ_R:<hz>,FREQ_G:<hz>,FREQ_B:<hz>
                        try:
                            kv: Dict[str, str] = {}
                            for part in line.replace(";", ",").split(","):
                                if ":" in part:
                                    k, v = part.split(":", 1)
                                    kv[k.strip().upper()] = v.strip()
                            # Priorizar valores RGB calibrados del firmware
                            if "TCS_RED" in kv:
                                # Nuevo formato con RGB ya calibrado
                                r_rgb = float(kv.get("TCS_RED", "0"))
                                g_rgb = float(kv.get("TCS_GREEN", "0"))
                                b_rgb = float(kv.get("TCS_BLUE", "0"))
                                # Obtener frecuencias si existen (para referencia)
                                r_hz = float(kv.get("FREQ_R", "0"))
                                g_hz = float(kv.get("FREQ_G", "0"))
                                b_hz = float(kv.get("FREQ_B", "0"))
                                # Emitir frecuencias para mantener compatibilidad con gráficas
                                self.sig_sample.emit(r_hz if r_hz > 0 else r_rgb, 
                                                   g_hz if g_hz > 0 else g_rgb, 
                                                   b_hz if b_hz > 0 else b_rgb, 
                                                   time.time())
                            else:
                                # Formato antiguo con solo frecuencias
                                r = float(kv.get("R", "0"))
                                g = float(kv.get("G", "0"))
                                b = float(kv.get("B", "0"))
                                self.sig_sample.emit(r, g, b, time.time())
                        except Exception:
                            pass
                except socket.timeout:
                    continue
                except Exception as e:
                    if not self._stopping:
                        self.sig_error.emit(f"[TCS] Error socket: {e}")
                    break
        except Exception as e:
            self.sig_error.emit(f"[TCS] Error de conexion: {e}")
        finally:
            try:
                if self.sock:
                    try:
                        self.sock.sendall(b"TCS_STOP\n")
                    except Exception:
                        pass
                    try:
                        self.sock.close()
                    except Exception:
                        pass
            finally:
                self.sock = None
                self.sig_status.emit("[TCS] Desconectado")


class ColorTCSLogic(QObject):
    def __init__(self, ui_widget: QWidget, main_window):
        super().__init__(ui_widget)
        self.ui = ui_widget
        self.main = main_window

        # Widgets
        self.btn_start = self.ui.findChild(QPushButton, "iniciarBt")
        self.btn_clear = self.ui.findChild(QPushButton, "limpiarBt")
        self.btn_export = self.ui.findChild(QPushButton, "exportarBt")
        self.btn_cal = self.ui.findChild(QPushButton, "calibrarBt")
        self.lbl_r = self.ui.findChild(QLabel, "lecturaRojoDt")
        self.lbl_g = self.ui.findChild(QLabel, "lecturaVerdeDt")
        self.lbl_b = self.ui.findChild(QLabel, "lecturaAzulDt")
        self.lbl_rc = self.ui.findChild(QLabel, "lecturaRojoDtCalibrado")
        self.lbl_gc = self.ui.findChild(QLabel, "lecturaVerdeDtCalibrado")
        self.lbl_bc = self.ui.findChild(QLabel, "lecturaAzulDtCalibrado")

        # Plot embed
        self.figure = Figure(figsize=(5, 3), dpi=100)
        self.canvas = FigureCanvas(self.figure)
        grap = self.ui.findChild(QWidget, "grapWid")
        if grap is not None:
            lay = grap.layout() or QVBoxLayout(grap)
            if grap.layout() is None:
                grap.setLayout(lay)
            lay.setContentsMargins(0, 0, 0, 0)
            lay.addWidget(self.canvas)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title("TCS3200 RGB (calibrado 0-255)")
        self.ax.set_xlabel("Tiempo (s)")
        self.ax.set_ylabel("Nivel (0-255)")
        self.window_s = 25.0
        self.ax.set_ylim(0, 255)
        self.line_r, = self.ax.plot([], [], label="R", color="#ff0000")
        self.line_g, = self.ax.plot([], [], label="G", color="#00aa00")
        self.line_b, = self.ax.plot([], [], label="B", color="#0000ff")
        self.ax.legend(loc="upper right")

        # Estado
        self.interval_ms = 300
        self.thread: Optional[_TcsThread] = None
        self.monitoring = False
        self._t0: Optional[float] = None
        self.series = {"t": [], "r": [], "g": [], "b": [], "r_raw": [], "g_raw": [], "b_raw": []}

        # Calibración
        # Calibración por pasos: none -> await_white -> await_black -> await_red -> await_green -> await_blue -> done
        self.cal_state = "none"
        self.cal = {
            "white": {"r": None, "g": None, "b": None},
            "black": {"r": None, "g": None, "b": None},
            "red":   {"r": None, "g": None, "b": None},
            "green": {"r": None, "g": None, "b": None},
            "blue":  {"r": None, "g": None, "b": None},
        }
        self._load_calibration()

        # Init
        self._init_labels()
        if self.btn_start:
            self.btn_start.clicked.connect(self._toggle)
        if self.btn_clear:
            self.btn_clear.clicked.connect(self._clear)
        if self.btn_export:
            self.btn_export.clicked.connect(self._export_excel)
        if self.btn_cal:
            self.btn_cal.clicked.connect(self._toggle_calibration)

    # ---- Helpers ----
    def _init_labels(self):
        if self.lbl_r:
            self.lbl_r.setText("--")
        if self.lbl_g:
            self.lbl_g.setText("--")
        if self.lbl_b:
            self.lbl_b.setText("--")
        if self.lbl_rc:
            self.lbl_rc.setText("--")
        if self.lbl_gc:
            self.lbl_gc.setText("--")
        if self.lbl_bc:
            self.lbl_bc.setText("--")
        self._update_cal_button()

    def _update_cal_button(self):
        if not self.btn_cal:
            return
        if self._is_calibrated():
            self.btn_cal.setText("Calibrado")
        elif self.cal_state == "await_white":
            self.btn_cal.setText("Capturar Blanco")
        elif self.cal_state == "await_black":
            self.btn_cal.setText("Capturar Negro")
        elif self.cal_state == "await_red":
            self.btn_cal.setText("Capturar Rojo")
        elif self.cal_state == "await_green":
            self.btn_cal.setText("Capturar Verde")
        elif self.cal_state == "await_blue":
            self.btn_cal.setText("Capturar Azul")
        else:
            self.btn_cal.setText("No Calibrado")

    def _cal_file(self) -> str:
        base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "colorTCS_calibration.json")

    def _load_calibration(self):
        try:
            with open(self._cal_file(), "r", encoding="utf-8") as f:
                data = json.load(f)
            self.cal = data
            # Si el archivo tenía todas las referencias, marcar done
            self.cal_state = "done" if self._is_calibrated() else "none"
        except Exception:
            self.cal_state = "none"

    def _save_calibration(self):
        try:
            with open(self._cal_file(), "w", encoding="utf-8") as f:
                json.dump(self.cal, f, indent=2)
        except Exception:
            pass

    def _is_calibrated(self) -> bool:
        try:
            # Requerimos white, black, red, green, blue todos presentes
            return all(
                self.cal[ref][ch] is not None
                for ref in ("white", "black", "red", "green", "blue")
                for ch in ("r", "g", "b")
            )
        except Exception:
            return False

    def _get_ip(self) -> Optional[str]:
        ip_edit = self.main.findChild(QLineEdit, "ipEdit") if self.main else None
        if not ip_edit:
            QMessageBox.critical(self.ui, "Error", "No se encontró el campo IP (ipEdit)")
            return None
        ip = ip_edit.text().strip()
        if not ip:
            QMessageBox.warning(self.ui, "Advertencia", "Ingrese la IP del ESP32.")
            return None
        return ip

    # ---- Actions ----
    def _toggle(self):
        if not self.monitoring:
            ip = self._get_ip()
            if not ip:
                return
            self.thread = _TcsThread(ip, interval_ms=self.interval_ms)
            self.thread.sig_status.connect(lambda m: print(m))
            self.thread.sig_error.connect(lambda m: print(m))
            self.thread.sig_sample.connect(self._on_sample)
            self.thread.start()
            self.monitoring = True
            if self.btn_start:
                self.btn_start.setText("Pausar")
            self._t0 = time.time()
        else:
            self._stop()

    def _stop(self):
        try:
            if self.thread:
                self.thread.stop()
        finally:
            self.thread = None
            self.monitoring = False
            if self.btn_start:
                self.btn_start.setText("Iniciar Monitoreo")

    def _toggle_calibration(self):
        if self._is_calibrated():
            # Reiniciar calibración completa
            self.cal = {
                "white": {"r": None, "g": None, "b": None},
                "black": {"r": None, "g": None, "b": None},
                "red":   {"r": None, "g": None, "b": None},
                "green": {"r": None, "g": None, "b": None},
                "blue":  {"r": None, "g": None, "b": None},
            }
            self.cal_state = "await_white"
            QMessageBox.information(self.ui, "Calibración", "Coloque un blanco de referencia y presione de nuevo para avanzar.")
        else:
            if self.cal_state == "none":
                self.cal_state = "await_white"
                QMessageBox.information(self.ui, "Calibración", "Coloque un blanco de referencia y presione nuevamente para avanzar.")
            elif self.cal_state == "await_white":
                # Los valores se actualizan en tiempo real mientras está en este estado
                self.cal_state = "await_black"
                QMessageBox.information(self.ui, "Calibración", "Blanco listo. Coloque negro y presione nuevamente.")
            elif self.cal_state == "await_black":
                self.cal_state = "await_red"
                QMessageBox.information(self.ui, "Calibración", "Negro listo. Coloque ROJO y presione nuevamente.")
            elif self.cal_state == "await_red":
                self.cal_state = "await_green"
                QMessageBox.information(self.ui, "Calibración", "Rojo listo. Coloque VERDE y presione nuevamente.")
            elif self.cal_state == "await_green":
                self.cal_state = "await_blue"
                QMessageBox.information(self.ui, "Calibración", "Verde listo. Coloque AZUL y presione nuevamente.")
            elif self.cal_state == "await_blue":
                self.cal_state = "done"
                self._save_calibration()
                QMessageBox.information(self.ui, "Calibración", "Calibración completa y guardada.")
        self._update_cal_button()

    @Slot(float, float, float, float)
    def _on_sample(self, r_hz: float, g_hz: float, b_hz: float, ts: float):
        # Actualizar etiquetas crudas
        if self.lbl_r:
            self.lbl_r.setText(f"{r_hz:.0f}")
        if self.lbl_g:
            self.lbl_g.setText(f"{g_hz:.0f}")
        if self.lbl_b:
            self.lbl_b.setText(f"{b_hz:.0f}")

        # Captura de calibración si está en progreso
        if self.cal_state == "await_white":
            self.cal["white"] = {"r": r_hz, "g": g_hz, "b": b_hz}
        elif self.cal_state == "await_black":
            self.cal["black"] = {"r": r_hz, "g": g_hz, "b": b_hz}
        elif self.cal_state == "await_red":
            self.cal["red"] = {"r": r_hz, "g": g_hz, "b": b_hz}
        elif self.cal_state == "await_green":
            self.cal["green"] = {"r": r_hz, "g": g_hz, "b": b_hz}
        elif self.cal_state == "await_blue":
            self.cal["blue"] = {"r": r_hz, "g": g_hz, "b": b_hz}

        # Convertir a 0..255 si calibrado; si no, escalar dinámico por sesión
        r_val, g_val, b_val = self._to_255(r_hz, g_hz, b_hz)

        if self.lbl_rc:
            self.lbl_rc.setText(str(int(r_val)))
        if self.lbl_gc:
            self.lbl_gc.setText(str(int(g_val)))
        if self.lbl_bc:
            self.lbl_bc.setText(str(int(b_val)))

        # Series
        if self._t0 is None:
            self._t0 = ts
        t_rel = ts - self._t0
        self.series["t"].append(t_rel)
        self.series["r_raw"].append(r_hz)
        self.series["g_raw"].append(g_hz)
        self.series["b_raw"].append(b_hz)
        self.series["r"].append(r_val)
        self.series["g"].append(g_val)
        self.series["b"].append(b_val)

        # Plot (solo calibrado 0..255 para claridad)
        self.line_r.set_data(self.series["t"], self.series["r"])
        self.line_g.set_data(self.series["t"], self.series["g"])
        self.line_b.set_data(self.series["t"], self.series["b"])
        # Ventana X deslizante 25s y Y fija 0..255
        if t_rel <= self.window_s:
            self.ax.set_xlim(0, self.window_s)
        else:
            self.ax.set_xlim(t_rel - self.window_s, t_rel)
        self.ax.set_ylim(0, 255)
        self.canvas.draw_idle()

    def _to_255(self, r_hz: float, g_hz: float, b_hz: float):
        if self._is_calibrated():
            # 1) Escala base por canal usando [black, white]
            def base_scale(v, w, k):
                try:
                    if w is None or k is None or w == k:
                        return 0.0
                    lo, hi = (k, w) if k < w else (w, k)
                    x = (v - lo) / (hi - lo)
                    return max(0.0, min(1.0, x))
                except Exception:
                    return 0.0
            r_s = base_scale(r_hz, self.cal["white"]["r"], self.cal["black"]["r"])
            g_s = base_scale(g_hz, self.cal["white"]["g"], self.cal["black"]["g"])
            b_s = base_scale(b_hz, self.cal["white"]["b"], self.cal["black"]["b"])

            # 2) Ganancia por canal a partir de referencias de color puro
            eps = 1e-6
            r_on_red = base_scale(self.cal["red"]["r"], self.cal["white"]["r"], self.cal["black"]["r"]) if self.cal["red"]["r"] is not None else 1.0
            g_on_green = base_scale(self.cal["green"]["g"], self.cal["white"]["g"], self.cal["black"]["g"]) if self.cal["green"]["g"] is not None else 1.0
            b_on_blue = base_scale(self.cal["blue"]["b"], self.cal["white"]["b"], self.cal["black"]["b"]) if self.cal["blue"]["b"] is not None else 1.0
            gain_r = 1.0 / max(eps, r_on_red)
            gain_g = 1.0 / max(eps, g_on_green)
            gain_b = 1.0 / max(eps, b_on_blue)

            r = max(0.0, min(255.0, 255.0 * r_s * gain_r))
            g = max(0.0, min(255.0, 255.0 * g_s * gain_g))
            b = max(0.0, min(255.0, 255.0 * b_s * gain_b))
            return r, g, b
        else:
            # Mapeo dinámico por sesión usando min/max acumulados por canal
            # Guardar mínimos/máximos como atributos
            if not hasattr(self, "_dyn"):
                self._dyn = {
                    "r": {"min": r_hz, "max": r_hz},
                    "g": {"min": g_hz, "max": g_hz},
                    "b": {"min": b_hz, "max": b_hz},
                }
            # actualizar
            for ch, v in ("r", r_hz), ("g", g_hz), ("b", b_hz):
                self._dyn[ch]["min"] = min(self._dyn[ch]["min"], v)
                self._dyn[ch]["max"] = max(self._dyn[ch]["max"], v)
            def dyn_map(v, ch):
                lo = self._dyn[ch]["min"]
                hi = self._dyn[ch]["max"]
                if hi <= lo:
                    return 0.0
                x = (v - lo) / (hi - lo)
                x = max(0.0, min(1.0, x))
                return x * 255.0
            return dyn_map(r_hz, "r"), dyn_map(g_hz, "g"), dyn_map(b_hz, "b")

    def _clear(self):
        self.series = {"t": [], "r": [], "g": [], "b": [], "r_raw": [], "g_raw": [], "b_raw": []}
        self._t0 = None
        self._init_labels()
        if hasattr(self, "_dyn"):
            delattr(self, "_dyn")
        self.ax.cla()
        self.ax.set_title("TCS3200 RGB (calibrado 0-255)")
        self.ax.set_xlabel("Tiempo (s)")
        self.ax.set_ylabel("Nivel (0-255)")
        self.ax.set_ylim(0, 255)
        self.ax.set_xlim(0, self.window_s)
        self.line_r, = self.ax.plot([], [], label="R", color="#ff0000")
        self.line_g, = self.ax.plot([], [], label="G", color="#00aa00")
        self.line_b, = self.ax.plot([], [], label="B", color="#0000ff")
        self.ax.legend(loc="upper right")
        self.canvas.draw_idle()

    def _export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.drawing.image import Image as XLImage

            default_name = f"ColorTCS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path, _ = QFileDialog.getSaveFileName(self.ui, "Exportar a Excel", default_name, "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            headers = ["timestamp_s", "R_hz", "G_hz", "B_hz", "R_0_255", "G_0_255", "B_0_255"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="FFEEDD")
                c.alignment = Alignment(horizontal="center")
            for t, rr, gg, bb, r, g, b in zip(
                self.series["t"], self.series["r_raw"], self.series["g_raw"], self.series["b_raw"],
                self.series["r"], self.series["g"], self.series["b"]
            ):
                ws.append([t, rr, gg, bb, r, g, b])

            meta = wb.create_sheet("Metadatos")
            meta.append(["Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            meta.append(["Intervalo (ms)", self.interval_ms])
            meta.append(["Calibrado", "Sí" if self._is_calibrated() else "No"])
            if self._is_calibrated():
                meta.append(["White R,G,B", f"{self.cal['white']['r']},{self.cal['white']['g']},{self.cal['white']['b']}"])
                meta.append(["Black R,G,B", f"{self.cal['black']['r']},{self.cal['black']['g']},{self.cal['black']['b']}"])
                meta.append(["Red Ref R,G,B",   f"{self.cal['red']['r']},{self.cal['red']['g']},{self.cal['red']['b']}"])
                meta.append(["Green Ref R,G,B", f"{self.cal['green']['r']},{self.cal['green']['g']},{self.cal['green']['b']}"])
                meta.append(["Blue Ref R,G,B",  f"{self.cal['blue']['r']},{self.cal['blue']['g']},{self.cal['blue']['b']}"])

            tmp_png = os.path.join(tempfile.gettempdir(), f"tcs_plot_{int(time.time())}.png")
            self.figure.savefig(tmp_png, dpi=150, bbox_inches="tight")
            wsimg = wb.create_sheet("Grafica")
            img = XLImage(tmp_png)
            wsimg.add_image(img, "A1")

            for sheet in (ws, meta):
                for col in sheet.columns:
                    width = 10
                    for cell in col:
                        width = max(width, len(str(cell.value)) if cell.value is not None else 0)
                    sheet.column_dimensions[col[0].column_letter].width = min(30, width + 2)

            wb.save(path)
            try:
                os.remove(tmp_png)
            except Exception:
                pass
            QMessageBox.information(self.ui, "Exportar", f"Exportado: {path}")
        except ImportError:
            QMessageBox.critical(self.ui, "Exportar", "Falta librería openpyxl.")
        except Exception as e:
            QMessageBox.critical(self.ui, "Exportar", f"Error exportando: {e}")

    def cleanup(self):
        self._stop()
