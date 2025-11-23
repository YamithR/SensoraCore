from __future__ import annotations

import os
import time
import socket
import json
from typing import Optional

from PySide6.QtCore import QObject, Signal, Slot
from PySide6.QtWidgets import QWidget, QPushButton, QLabel, QLCDNumber, QProgressBar, QMessageBox, QFileDialog, QLineEdit


class _BrightnessThread(QObject):
    sig_status = Signal(str)
    sig_error = Signal(str)
    sig_sample = Signal(int, float, float)  # adc, volt, ts

    def __init__(self, ip: str, port: int = 8080, interval_ms: int = 200):
        super().__init__()
        self.ip = ip
        self.port = port
        self.interval_ms = max(50, int(interval_ms))
        self._running = False
        self._stopping = False
        self.sock: Optional[socket.socket] = None

    def start(self):
        if self._running:
            return
        from threading import Thread
        self._stopping = False
        self._running = True
        Thread(target=self._run, daemon=True).start()

    def stop(self):
        self._stopping = True
        self._running = False
        try:
            if self.sock:
                try:
                    self.sock.sendall(b"BR_STOP\n")
                except Exception:
                    pass
                try:
                    self.sock.close()
                except Exception:
                    pass
        finally:
            self.sock = None

    def _run(self):
        try:
            self.sig_status.emit("[BR] Conectando...")
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(5)
            s.connect((self.ip, self.port))
            self.sock = s
            s.sendall(b"MODO:BRIGHTNESS\n")
            resp = s.recv(128).decode(errors="ignore").strip()
            if "BRIGHTNESS_OK" not in resp:
                self.sig_error.emit("[BR] ESP32 no acepto BRIGHTNESS")
                return
            s.sendall(f"BR_START:{self.interval_ms}\n".encode())
            self.sig_status.emit("[BR] Streaming iniciado")
            s.settimeout(1.0)
            buf = ""
            while self._running:
                try:
                    chunk = s.recv(256)
                    if not chunk:
                        break
                    buf += chunk.decode(errors="ignore")
                    while "\n" in buf:
                        line, buf = buf.split("\n", 1)
                        line = line.strip()
                        if not line:
                            continue
                        # LDR_RAW:<n>,VOLTAGE:<v>,BRIGHTNESS:<pct>
                        try:
                            kv = {}
                            for part in line.replace(";", ",").split(","):
                                if ":" in part:
                                    k, v = part.split(":", 1)
                                    kv[k.strip().upper()] = v.strip()
                            # Soportar ambos formatos: nuevo y antiguo
                            adc = int(float(kv.get("LDR_RAW", kv.get("ADC", "0"))))
                            volt = float(kv.get("VOLTAGE", kv.get("VOLT", str((adc/4095.0)*3.3))))
                            self.sig_sample.emit(adc, volt, time.time())
                        except Exception:
                            pass
                except socket.timeout:
                    continue
                except Exception as e:
                    if not self._stopping:
                        self.sig_error.emit(f"[BR] Error socket: {e}")
                    break
        except Exception as e:
            self.sig_error.emit(f"[BR] Error de conexion: {e}")
        finally:
            try:
                if self.sock:
                    try:
                        self.sock.sendall(b"BR_STOP\n")
                    except Exception:
                        pass
                    try:
                        self.sock.close()
                    except Exception:
                        pass
            finally:
                self.sock = None
                self.sig_status.emit("[BR] Desconectado")


class BrightnessLogic(QObject):
    def __init__(self, ui_widget: QWidget, main_window):
        super().__init__(ui_widget)
        self.ui = ui_widget
        self.main = main_window

        # Widgets
        self.btn_start: QPushButton = self.ui.findChild(QPushButton, "iniciarBt")
        self.btn_clear: QPushButton = self.ui.findChild(QPushButton, "limpiarBt")
        self.btn_export: QPushButton = self.ui.findChild(QPushButton, "exportarBt")
        self.btn_cal: QPushButton = self.ui.findChild(QPushButton, "calibrarBt")
        self.lbl_adc: QLabel = self.ui.findChild(QLabel, "LecturaLDRDt")
        self.lcd_pct: QLCDNumber = self.ui.findChild(QLCDNumber, "PorcentajeDeLuz")
        self.bar_pct: QProgressBar = self.ui.findChild(QProgressBar, "BarraIndicadorDeLuz")

        # Estado
        self.interval_ms = 200
        self.thread: Optional[_BrightnessThread] = None
        self.monitoring = False
        self._t0: Optional[float] = None
        self._series = {"t": [], "adc": [], "volt": [], "pct": []}

        # Calibración
        # Estados: none -> await_dark -> await_bright -> done
        self.cal_state = "none"
        self.cal_dark_adc: Optional[int] = None  # ADC en oscuridad total
        self.cal_bright_adc: Optional[int] = None  # ADC en brillo máximo
        self._load_calibration()

        # Init
        self._init_widgets()
        if self.btn_start:
            self.btn_start.clicked.connect(self._toggle)
        if self.btn_clear:
            self.btn_clear.clicked.connect(self._clear)
        if self.btn_export:
            self.btn_export.clicked.connect(self._export_excel)
        if self.btn_cal:
            self.btn_cal.clicked.connect(self._toggle_calibration)

    def _init_widgets(self):
        if self.bar_pct:
            self.bar_pct.setRange(0, 100)
            self.bar_pct.setValue(0)
        if self.lcd_pct:
            self.lcd_pct.display(0)
        if self.lbl_adc:
            self.lbl_adc.setText("--")
        self._update_cal_button()

    def _update_cal_button(self):
        """Actualiza el texto del botón de calibración según el estado"""
        if not self.btn_cal:
            return
        
        base_style = """
            QPushButton {
                padding: 8px;
                margin-top: 5px;
                border-radius: 4px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                opacity: 0.9;
            }
        """
        
        if self._is_calibrated():
            self.btn_cal.setText("✓ Calibrado")
            self.btn_cal.setStyleSheet(base_style + """
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: 2px solid #1e7e34;
                }
            """)
        elif self.cal_state == "await_dark":
            self.btn_cal.setText("Capturar Oscuridad")
            self.btn_cal.setStyleSheet(base_style + """
                QPushButton {
                    background-color: #17a2b8;
                    color: white;
                    border: 2px solid #117a8b;
                }
            """)
        elif self.cal_state == "await_bright":
            self.btn_cal.setText("Capturar Brillo Máximo")
            self.btn_cal.setStyleSheet(base_style + """
                QPushButton {
                    background-color: #ffc107;
                    color: black;
                    border: 2px solid #d39e00;
                }
            """)
        else:
            self.btn_cal.setText("No Calibrado")
            self.btn_cal.setStyleSheet(base_style + """
                QPushButton {
                    background-color: #6c757d;
                    color: white;
                    border: 2px solid #545b62;
                }
            """)

    def _cal_file(self) -> str:
        """Ruta del archivo de calibración"""
        base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, "brightness_calibration.json")

    def _load_calibration(self):
        """Carga la calibración desde archivo si existe"""
        try:
            with open(self._cal_file(), "r", encoding="utf-8") as f:
                data = json.load(f)
            self.cal_dark_adc = data.get("dark_adc")
            self.cal_bright_adc = data.get("bright_adc")
            if self._is_calibrated():
                self.cal_state = "done"
                print(f"[Brightness] Calibración cargada: Oscuridad={self.cal_dark_adc}, Brillo={self.cal_bright_adc}")
            else:
                self.cal_state = "none"
        except Exception:
            self.cal_state = "none"
            self.cal_dark_adc = None
            self.cal_bright_adc = None

    def _save_calibration(self):
        """Guarda la calibración en archivo"""
        try:
            data = {
                "dark_adc": self.cal_dark_adc,
                "bright_adc": self.cal_bright_adc
            }
            with open(self._cal_file(), "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            print(f"[Brightness] Calibración guardada: Oscuridad={self.cal_dark_adc}, Brillo={self.cal_bright_adc}")
        except Exception as e:
            print(f"[Brightness] Error guardando calibración: {e}")

    def _is_calibrated(self) -> bool:
        """Verifica si la calibración está completa"""
        return (self.cal_dark_adc is not None and 
                self.cal_bright_adc is not None and 
                self.cal_bright_adc > self.cal_dark_adc)

    def _get_ip(self) -> Optional[str]:
        ip_edit = self.main.findChild(QLineEdit, "ipEdit") if self.main else None
        if not ip_edit:
            QMessageBox.critical(self.ui, "Error", "No se encontró el campo IP (ipEdit)")
            return None
        ip = ip_edit.text().strip()
        if not ip:
            QMessageBox.warning(self.ui, "Advertencia", "Ingrese la IP del ESP32.")
            return None
        return ip

    def _toggle(self):
        if not self.monitoring:
            ip = self._get_ip()
            if not ip:
                return
            self.thread = _BrightnessThread(ip, interval_ms=self.interval_ms)
            self.thread.sig_status.connect(lambda m: print(m))
            self.thread.sig_error.connect(lambda m: print(m))
            self.thread.sig_sample.connect(self._on_sample)
            self.thread.start()
            self.monitoring = True
            if self.btn_start:
                self.btn_start.setText("Pausar")
            self._t0 = time.time()
            
            # Enviar calibración al ESP32 si está calibrado
            if self._is_calibrated():
                self._send_calibration_to_esp32()
        else:
            self._stop()
    
    def _send_calibration_to_esp32(self):
        """Envía los valores de calibración al ESP32"""
        if self.thread and self.thread.sock and self._is_calibrated():
            try:
                cal_cmd = f"CAL:MIN={self.cal_dark_adc},MAX={self.cal_bright_adc}\n"
                self.thread.sock.sendall(cal_cmd.encode())
                print(f"[Brightness] Calibración enviada al ESP32: MIN={self.cal_dark_adc}, MAX={self.cal_bright_adc}")
                time.sleep(0.1)  # Dar tiempo al ESP32 para procesar
            except Exception as e:
                print(f"[Brightness] Error enviando calibración: {e}")

    def _stop(self):
        try:
            if self.thread:
                self.thread.stop()
        finally:
            self.thread = None
            self.monitoring = False
            if self.btn_start:
                self.btn_start.setText("Iniciar Monitoreo")

    def _toggle_calibration(self):
        """Maneja el proceso de calibración paso a paso"""
        if self._is_calibrated():
            # Reiniciar calibración
            reply = QMessageBox.question(
                self.ui, 
                "Reiniciar Calibración",
                "¿Desea reiniciar la calibración?\n\nSe perderán los valores actuales.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.cal_dark_adc = None
                self.cal_bright_adc = None
                self.cal_state = "await_dark"
                QMessageBox.information(
                    self.ui, 
                    "Calibración - Paso 1",
                    "PASO 1: Cubra completamente el sensor LDR (oscuridad total)\n\n"
                    "Espere a que la lectura se estabilice y presione el botón nuevamente."
                )
        else:
            if self.cal_state == "none":
                # Iniciar calibración
                self.cal_state = "await_dark"
                QMessageBox.information(
                    self.ui, 
                    "Calibración - Paso 1",
                    "PASO 1: Cubra completamente el sensor LDR (oscuridad total)\n\n"
                    "Espere a que la lectura se estabilice y presione el botón nuevamente."
                )
            elif self.cal_state == "await_dark":
                # Capturar valor de oscuridad
                if self.lbl_adc and self.lbl_adc.text() != "--":
                    try:
                        self.cal_dark_adc = int(self.lbl_adc.text())
                        self.cal_state = "await_bright"
                        QMessageBox.information(
                            self.ui, 
                            "Calibración - Paso 2",
                            f"✓ Oscuridad capturada: ADC = {self.cal_dark_adc}\n\n"
                            "PASO 2: Ilumine el sensor LDR al máximo (luz brillante)\n\n"
                            "Espere a que la lectura se estabilice y presione el botón nuevamente."
                        )
                    except Exception:
                        QMessageBox.warning(self.ui, "Error", "No se pudo leer el valor ADC actual.")
                else:
                    QMessageBox.warning(self.ui, "Error", "Inicie el monitoreo primero para obtener lecturas.")
            elif self.cal_state == "await_bright":
                # Capturar valor de brillo máximo
                if self.lbl_adc and self.lbl_adc.text() != "--":
                    try:
                        self.cal_bright_adc = int(self.lbl_adc.text())
                        if self.cal_bright_adc <= self.cal_dark_adc:
                            QMessageBox.warning(
                                self.ui, 
                                "Error de Calibración",
                                f"El valor de brillo ({self.cal_bright_adc}) debe ser mayor que el de oscuridad ({self.cal_dark_adc}).\n\n"
                                "Asegúrese de iluminar correctamente el sensor."
                            )
                            return
                        self.cal_state = "done"
                        self._save_calibration()
                        # Enviar calibración al ESP32
                        self._send_calibration_to_esp32()
                        QMessageBox.information(
                            self.ui, 
                            "Calibración Completa",
                            f"✓ Calibración completada y guardada!\n\n"
                            f"Oscuridad: ADC = {self.cal_dark_adc}\n"
                            f"Brillo máximo: ADC = {self.cal_bright_adc}\n\n"
                            f"Todas las mediciones futuras se escalarán automáticamente usando estos valores.\n\n"
                            f"El estado 'LDR [CAL]' se mostrará en la pantalla OLED del ESP32."
                        )
                    except Exception:
                        QMessageBox.warning(self.ui, "Error", "No se pudo leer el valor ADC actual.")
                else:
                    QMessageBox.warning(self.ui, "Error", "Inicie el monitoreo primero para obtener lecturas.")
        
        self._update_cal_button()

    @Slot(int, float, float)
    def _on_sample(self, adc: int, volt: float, ts: float):
        # Actualizar captura durante calibración
        if self.cal_state == "await_dark":
            # Guardar temporalmente para captura
            pass
        elif self.cal_state == "await_bright":
            # Guardar temporalmente para captura
            pass
        
        # Calcular porcentaje con calibración si está disponible
        if self._is_calibrated():
            # Usar valores calibrados para reescalar
            adc_range = self.cal_bright_adc - self.cal_dark_adc
            adc_normalized = adc - self.cal_dark_adc
            pct = int(round(max(0.0, min(100.0, (adc_normalized * 100.0 / adc_range)))))
        else:
            # Sin calibración, usar rango completo del ADC
            pct = int(round(max(0.0, min(100.0, (adc * 100.0 / 4095.0)))))
        
        if self.lbl_adc:
            self.lbl_adc.setText(str(adc))
        if self.lcd_pct:
            self.lcd_pct.display(pct)
        if self.bar_pct:
            self.bar_pct.setValue(pct)

        if self._t0 is None:
            self._t0 = ts
        self._series["t"].append(ts - self._t0)
        self._series["adc"].append(adc)
        self._series["volt"].append(volt)
        self._series["pct"].append(pct)

    def _clear(self):
        for k in self._series:
            self._series[k].clear()
        self._init_widgets()

    def _export_excel(self):
        # Sin gráfica, solo datos + metadatos
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            path, _ = QFileDialog.getSaveFileName(self.ui, "Exportar a Excel", None, "Excel (*.xlsx)")
            if not path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            headers = ["t_s", "adc", "volt", "percent"]
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDEEFF")
                c.alignment = Alignment(horizontal="center")
            for t, a, v, p in zip(self._series["t"], self._series["adc"], self._series["volt"], self._series["pct"]):
                ws.append([t, a, v, p])

            meta = wb.create_sheet("Metadatos")
            from datetime import datetime
            meta.append(["Fecha", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            meta.append(["Intervalo (ms)", self.interval_ms])
            meta.append(["Descripcion", "Lectura de LDR → % de luz (0..100)"])
            meta.append(["Calibrado", "Sí" if self._is_calibrated() else "No"])
            if self._is_calibrated():
                meta.append(["ADC Oscuridad", self.cal_dark_adc])
                meta.append(["ADC Brillo Máximo", self.cal_bright_adc])
                meta.append(["Rango ADC", self.cal_bright_adc - self.cal_dark_adc])

            # auto ancho simple
            for sheet in (ws, meta):
                for col in sheet.columns:
                    width = 10
                    for cell in col:
                        width = max(width, len(str(cell.value)) if cell.value is not None else 0)
                    sheet.column_dimensions[col[0].column_letter].width = min(30, width + 2)

            wb.save(path)
            QMessageBox.information(self.ui, "Exportar", f"Exportado: {path}")
        except ImportError:
            QMessageBox.critical(self.ui, "Exportar", "Falta librería openpyxl.")
        except Exception as e:
            QMessageBox.critical(self.ui, "Exportar", f"Error exportando: {e}")

    def cleanup(self):
        self._stop()
