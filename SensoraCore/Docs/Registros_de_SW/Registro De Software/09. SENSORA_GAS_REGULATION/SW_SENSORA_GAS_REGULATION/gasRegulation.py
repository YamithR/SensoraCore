from __future__ import annotations

import os
import json
import time
import math
import socket
from dataclasses import dataclass
from typing import Optional, Dict, List

from PySide6.QtCore import QThread, Signal
from PySide6.QtWidgets import QWidget, QMessageBox, QFileDialog


# ===== Persistencia =====
def _calib_path() -> str:
    base = os.path.join(os.path.expanduser("~"), ".sensora_core")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, "gas_calib.json")


def _load_calib() -> Dict[str, dict]:
    p = _calib_path()
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_calib(d: Dict[str, dict]) -> None:
    p = _calib_path()
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(d, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ===== Hilo TCP =====
class GasThread(QThread):
    data = Signal(str, int, float)  # name, adc, volt
    status = Signal(str)

    def __init__(self, esp32_ip: str, port: int, sensor: str, period_ms: int = 500):
        super().__init__()
        self.esp32_ip = esp32_ip
        self.port = port
        self.sensor = sensor  # "MQ2" | "MQ3"
        self.period_ms = max(200, int(period_ms))
        self._running = False
        self._stopping = False
        self.sock: Optional[socket.socket] = None

    def run(self):
        self._running = True
        try:
            self.status.emit(f"Conectando a {self.esp32_ip}:{self.port}...")
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.sock.settimeout(5)
            self.sock.connect((self.esp32_ip, self.port))
            self.sock.sendall(b"MODO:GAS_REGULATION\n")
            resp = self.sock.recv(128).decode(errors="ignore").strip()
            if "GAS_REGULATION_OK" not in resp:
                self.status.emit("ESP32 no acepto modo GAS_REGULATION")
                return
            # set sensor
            self._send(f"GR_SET:{self.sensor}\n")
            # start
            self._send(f"GR_START:{self.period_ms}\n")
            self.status.emit("Monitoreo de gas iniciado")
            self.sock.settimeout(2)
            buf = ""
            while self._running:
                try:
                    chunk = self.sock.recv(256)
                    if not chunk:
                        break
                    buf += chunk.decode(errors="ignore")
                    while "\n" in buf:
                        line, buf = buf.split("\n", 1)
                        line = line.strip()
                        if not line:
                            continue
                        # Esperado: SENSOR:MQ2,ADC:2048,VOLT:1.65
                        try:
                            parts = [p.strip() for p in line.replace(";", ",").split(",")]
                            kv = {}
                            for p in parts:
                                if ":" in p:
                                    k, v = p.split(":", 1)
                                    kv[k.strip().upper()] = v.strip()
                            name = kv.get("SENSOR", self.sensor)
                            adc_s = kv.get("ADC")
                            volt_s = kv.get("VOLT")
                            adc = int(float(adc_s)) if adc_s is not None else 0
                            if volt_s is not None:
                                volt = float(volt_s)
                            else:
                                # fallback si firmware no manda VOLT
                                volt = (adc / 4095.0) * 3.3
                            self.data.emit(name, adc, volt)
                        except Exception:
                            pass
                except socket.timeout:
                    continue
                except Exception as e:
                    if not self._stopping:
                        self.status.emit(f"Error socket GAS: {e}")
                    break
        except Exception as e:
            self.status.emit(f"Error de conexion: {e}")
        finally:
            try:
                if self.sock:
                    if not self._stopping:
                        try:
                            self.sock.sendall(b"GR_STOP\n")
                        except Exception:
                            pass
                    try:
                        self.sock.close()
                    except Exception:
                        pass
            finally:
                self.sock = None
                self.status.emit("Desconectado")

    def stop(self):
        self._stopping = True
        self._running = False
        if self.sock:
            try:
                self.sock.sendall(b"GR_STOP\n")
            except Exception:
                pass
            try:
                self.sock.close()
            except Exception:
                pass
        self.wait(2000)

    def switch_sensor(self, name: str):
        self.sensor = name
        self._send(f"GR_SET:{self.sensor}\n")

    def _send(self, txt: str):
        try:
            if self.sock and self._running:
                self.sock.sendall(txt.encode())
        except Exception:
            pass


# ===== Lógica de UI =====
@dataclass
class GasCal:
    RL: float = 10000.0  # ohm
    Ro: Optional[float] = None
    m: Optional[float] = None
    b: Optional[float] = None


class GasRegulationLogic(QWidget):
    def __init__(self, ui_widget, main_window=None):
        super().__init__()
        if ui_widget is None:
            raise ValueError("ui_widget no puede ser None")
        self.ui = ui_widget
        self.main_window = main_window

        # Estado
        self.current_sensor = "MQ2"  # MQ2 | MQ3
        self.period_ms = 500
        self._calib = _load_calib()  # por sensor
        # series
        self._t0 = time.time()
        self._times: List[float] = []
        self._ppm_mq2: List[float] = []
        self._ppm_mq3: List[float] = []
        self._data_rows: List[dict] = []

        # Plot
        self._setup_plot()

        # UI wiring
        if hasattr(self.ui, 'iniciarBt'):
            self.ui.iniciarBt.clicked.connect(self.toggle)
        if hasattr(self.ui, 'limpiarBt'):
            self.ui.limpiarBt.clicked.connect(self._clear_plot)
        if hasattr(self.ui, 'exportarBt'):
            self.ui.exportarBt.clicked.connect(self._export_excel)
        if hasattr(self.ui, 'calibrarBt'):
            self.ui.calibrarBt.clicked.connect(self._calibrate_flow)

        if hasattr(self.ui, 'MQ2'):
            self.ui.MQ2.setCheckable(True)
            self.ui.MQ2.clicked.connect(lambda: self._select_sensor('MQ2'))
        if hasattr(self.ui, 'MQ3'):
            self.ui.MQ3.setCheckable(True)
            self.ui.MQ3.clicked.connect(lambda: self._select_sensor('MQ3'))

        self._refresh_checks()
        self._refresh_cal_status()

        self.thread: Optional[GasThread] = None
        self.monitoring = False

    # ---- Plot ----
    def _setup_plot(self):
        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            self._Figure = Figure
            self._FigureCanvas = FigureCanvas
            self.fig = Figure(figsize=(5, 3), tight_layout=True)
            self.ax = self.fig.add_subplot(111)
            self.ax.set_title('Concentración (PPM)')
            self.ax.set_xlabel('Tiempo (s)')
            self.ax.set_ylabel('PPM')
            self.ax.grid(True, alpha=0.3)
            self.canvas = FigureCanvas(self.fig)
            if hasattr(self.ui, 'grapWid'):
                from PySide6.QtWidgets import QVBoxLayout
                lay = self.ui.grapWid.layout() or QVBoxLayout(self.ui.grapWid)
                lay.addWidget(self.canvas)
            (self._mq2_line,) = self.ax.plot([], [], label='MQ2')
            (self._mq3_line,) = self.ax.plot([], [], label='MQ3')
            self.ax.legend(loc='upper right')
            self.canvas.draw()
        except Exception as e:
            print(f"No se pudo inicializar la gráfica GAS: {e}")
            self.canvas = None

    def _clear_plot(self):
        self._t0 = time.time()
        self._times.clear()
        self._ppm_mq2.clear()
        self._ppm_mq3.clear()
        self._data_rows.clear()
        if self.canvas:
            self._mq2_line.set_data([], [])
            self._mq3_line.set_data([], [])
            self.ax.relim(); self.ax.autoscale_view(); self.canvas.draw_idle()

    # ---- Control ----
    def toggle(self):
        if self.monitoring:
            self.stop()
        else:
            self.start()

    def start(self):
        try:
            ip = self._get_ip()
            if not ip:
                QMessageBox.information(self, "ESP32", "Ingrese la IP del ESP32 en el campo correspondiente.")
                return
            if self.thread and self.thread.isRunning():
                self.thread.stop(); self.thread = None
            self.thread = GasThread(ip, 8080, self.current_sensor, self.period_ms)
            self.thread.data.connect(self._on_data)
            self.thread.status.connect(lambda m: print(f"[GAS] {m}"))
            self.thread.start()
            self.monitoring = True
            if hasattr(self.ui, 'iniciarBt'):
                self.ui.iniciarBt.setText('Pausar')
        except Exception as e:
            QMessageBox.critical(self, "GasRegulation", f"No se pudo iniciar: {e}")

    def stop(self):
        try:
            self.monitoring = False
            if self.thread:
                if self.thread.isRunning():
                    self.thread.stop()
                self.thread = None
            if hasattr(self.ui, 'iniciarBt'):
                self.ui.iniciarBt.setText('Iniciar Monitoreo')
        except Exception:
            pass

    def cleanup(self):
        self.stop()

    def _select_sensor(self, name: str):
        self.current_sensor = name
        self._refresh_checks()
        if self.thread and self.thread.isRunning():
            self.thread.switch_sensor(name)

    def _refresh_checks(self):
        try:
            if hasattr(self.ui, 'MQ2'):
                self.ui.MQ2.setChecked(self.current_sensor == 'MQ2')
            if hasattr(self.ui, 'MQ3'):
                self.ui.MQ3.setChecked(self.current_sensor == 'MQ3')
        except Exception:
            pass

    # ---- Datos entrantes ----
    def _on_data(self, name: str, adc: int, volt: float):
        t = time.time() - getattr(self, '_t0', time.time())
        self._times.append(t)

        # Mostrar lecturas crudas por sensor
        if name.upper() == 'MQ2':
            if hasattr(self.ui, 'LecturaMQ2Dt'):
                self.ui.LecturaMQ2Dt.setText(str(adc))
        elif name.upper() == 'MQ3':
            if hasattr(self.ui, 'LecturaMQ3Dt'):
                self.ui.LecturaMQ3Dt.setText(str(adc))

        # Calcular PPM si hay calibración
        ppm = self._compute_ppm(name.upper(), adc, volt)
        if name.upper() == 'MQ2':
            if hasattr(self.ui, 'ppmMQ2Dt'):
                self.ui.ppmMQ2Dt.setText(f"{ppm:.2f}" if ppm == ppm else "-")
            if hasattr(self.ui, 'ppmMQ2DtCalibrado'):
                self.ui.ppmMQ2DtCalibrado.setText(f"{ppm:.2f}" if ppm == ppm else "-")
            self._ppm_mq2.append(ppm if ppm == ppm else 0.0)
        elif name.upper() == 'MQ3':
            if hasattr(self.ui, 'ppmMQ3Dt'):
                self.ui.ppmMQ3Dt.setText(f"{ppm:.2f}" if ppm == ppm else "-")
            if hasattr(self.ui, 'ppmMQ3DtCalibrado'):
                self.ui.ppmMQ3DtCalibrado.setText(f"{ppm:.2f}" if ppm == ppm else "-")
            self._ppm_mq3.append(ppm if ppm == ppm else 0.0)

        # Registrar para exportación
        self._data_rows.append({
            't_s': t,
            'sensor': name,
            'adc': adc,
            'volt': volt,
            'ppm': ppm if ppm == ppm else None,
        })

        # Actualizar plot
        if self.canvas:
            self._mq2_line.set_data(self._times[:len(self._ppm_mq2)], self._ppm_mq2)
            self._mq3_line.set_data(self._times[:len(self._ppm_mq3)], self._ppm_mq3)
            self.ax.relim(); self.ax.autoscale_view(); self.canvas.draw_idle()

    def _apply_cal(self, sensor: str) -> GasCal:
        c = self._calib.get(sensor.upper()) or {}
        return GasCal(
            RL=float(c.get('RL', 10000.0)),
            Ro=(float(c['Ro']) if 'Ro' in c else None),
            m=(float(c['m']) if 'm' in c else None),
            b=(float(c['b']) if 'b' in c else None),
        )

    def _compute_ppm(self, sensor: str, adc: int, volt: float) -> float:
        try:
            cal = self._apply_cal(sensor)
            RL = cal.RL if cal.RL else 10000.0
            Vcc = 3.3
            Rs = RL * (Vcc - volt) / max(1e-6, volt)
            if not cal.Ro or not cal.m is not None or not cal.b is not None or Rs <= 0:
                return float('nan')
            ratio = Rs / cal.Ro
            if ratio <= 0:
                return float('nan')
            return 10 ** (cal.m * math.log10(ratio) + cal.b)
        except Exception:
            return float('nan')

    def _refresh_cal_status(self):
        try:
            c = self._apply_cal(self.current_sensor)
            has = (c.Ro is not None) and (c.m is not None) and (c.b is not None)
            if hasattr(self.ui, 'calibrarBt'):
                self.ui.calibrarBt.setText('Calibrado' if has else 'No Calibrado')
            if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
                self.ui.calBox.setVisible(has)
        except Exception:
            pass

    # ---- Calibración en app ----
    def _calibrate_flow(self):
        try:
            from PySide6.QtWidgets import QInputDialog
            # Paso 1: Ro (Ω) en aire limpio
            # Tomar pista de últimos valores si existen
            last_ro = 10000.0
            ro_val, ok = QInputDialog.getDouble(self, "Calibración (Ro)", "Ro (Ω) en aire limpio:", float(last_ro), 1.0, 1e9, 0)
            if not ok:
                return
            # Paso 2: puntos para m, b en log-log
            n, ok = QInputDialog.getInt(self, "Puntos de calibración", "Cantidad de puntos (>=2):", 2, 2, 10, 1)
            if not ok:
                # guarda solo Ro
                cur = self._calib.get(self.current_sensor, {})
                cur.update({'Ro': ro_val})
                self._calib[self.current_sensor] = cur
                _save_calib(self._calib)
                self._refresh_cal_status()
                return

            xs: List[float] = []
            ys: List[float] = []
            for i in range(n):
                # Pedir Rs actual y ppm de referencia
                rs, ok1 = QInputDialog.getDouble(self, f"Punto {i+1}/{n}", "Rs actual (Ω):", 10000.0, 1.0, 1e12, 0)
                if not ok1:
                    return
                ppm_ref, ok2 = QInputDialog.getDouble(self, f"Punto {i+1}/{n}", "PPM de referencia:", 100.0, 0.0, 1e9, 2)
                if not ok2:
                    return
                xs.append(math.log10(max(1e-12, rs / ro_val)))
                ys.append(math.log10(max(1e-12, ppm_ref)))

            # Ajuste lineal
            m, b = self._linreg(xs, ys)
            self._calib[self.current_sensor] = {'RL': 10000.0, 'Ro': ro_val, 'm': m, 'b': b}
            _save_calib(self._calib)
            self._refresh_cal_status()
            QMessageBox.information(self, "Calibración", f"Guardado: log10(PPM) = {m:.4f}*log10(Rs/Ro) + {b:.4f}")
        except Exception as e:
            QMessageBox.critical(self, "Calibración", f"Error al calibrar: {e}")

    def _linreg(self, xs: List[float], ys: List[float]):
        n = len(xs)
        if n < 2:
            return 1.0, 0.0
        sx = sum(xs); sy = sum(ys)
        sxx = sum(x*x for x in xs)
        sxy = sum(xs[i]*ys[i] for i in range(n))
        den = n*sxx - sx*sx
        if abs(den) < 1e-12:
            return 1.0, 0.0
        m = (n*sxy - sx*sy)/den
        b = (sy - m*sx)/n
        return m, b

    # ---- Exportación Excel (no CSV) ----
    def _export_excel(self):
        try:
            path, _ = QFileDialog.getSaveFileName(self, "Guardar archivo", f"Gas_{time.strftime('%Y%m%d_%H%M%S')}.xlsx", "Archivos Excel (*.xlsx)")
            if not path:
                return
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.drawing.image import Image

            wb = Workbook()
            # Datos
            ws = wb.active; ws.title = 'Datos'
            headers = ['t_s', 'sensor', 'adc', 'volt', 'ppm']
            ws.append(headers)
            for r in self._data_rows:
                ws.append([r['t_s'], r['sensor'], r['adc'], r['volt'], r['ppm'] if r['ppm'] is not None else ''])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid"); cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = min(24, max(10, *(len(str(c.value)) if c.value is not None else 0 for c in col)) + 2)

            # Metadatos
            from datetime import datetime
            meta = wb.create_sheet('Metadatos')
            meta.append(["Fecha", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
            meta.append(["Sensores", "MQ2 (humo), MQ3 (alcohol)"])
            meta.append(["Intervalo (ms)", self.period_ms])

            # Calibración
            cal = wb.create_sheet('Calibracion')
            cal.append(['Sensor', 'RL(Ω)', 'Ro(Ω)', 'm', 'b', 'Ecuación'])
            for s in ('MQ2','MQ3'):
                c = self._calib.get(s, {})
                eq = '-' if 'Ro' not in c or 'm' not in c or 'b' not in c else f"log10(PPM) = {c['m']:.4f}*log10(Rs/Ro) + {c['b']:.4f}"
                cal.append([s, c.get('RL', 10000.0), c.get('Ro', '-'), c.get('m', '-'), c.get('b', '-'), eq])
            cal.column_dimensions['A'].width = 10
            cal.column_dimensions['B'].width = 12
            cal.column_dimensions['C'].width = 14
            cal.column_dimensions['D'].width = 10
            cal.column_dimensions['E'].width = 10
            cal.column_dimensions['F'].width = 44

            # Hoja Gráfica
            ws_plot = wb.create_sheet('Grafica')
            img_path = self._make_plot_image()
            temp_img_path = None
            if img_path and os.path.exists(img_path):
                img = Image(img_path); img.width = 1000; img.height = 500
                ws_plot['A1'] = 'Concentración (PPM)'; ws_plot['A1'].font = Font(bold=True, size=16)
                ws_plot.add_image(img, 'A3')
                temp_img_path = img_path

            wb.save(path)
            if temp_img_path:
                try:
                    os.unlink(temp_img_path)
                except Exception:
                    pass
            QMessageBox.information(self, "Exportar", f"Exportado: {path}")
        except ImportError:
            QMessageBox.critical(self, "Exportar", "No se encontró openpyxl. Instala la librería para exportar a Excel.")
        except Exception as e:
            QMessageBox.critical(self, "Exportar", f"Fallo exportando: {e}")

    def _make_plot_image(self) -> Optional[str]:
        try:
            from matplotlib.figure import Figure
            import io, tempfile
            fig = Figure(figsize=(12, 5), tight_layout=True)
            ax = fig.add_subplot(111)
            ax.set_title('PPM en vivo')
            ax.set_xlabel('Tiempo (s)')
            ax.set_ylabel('PPM')
            ax.grid(True, alpha=0.3)
            ax.plot(self._times[:len(self._ppm_mq2)], self._ppm_mq2, label='MQ2')
            ax.plot(self._times[:len(self._ppm_mq3)], self._ppm_mq3, label='MQ3')
            ax.legend(loc='best')
            buf = io.BytesIO(); fig.savefig(buf, format='png', dpi=150); buf.seek(0)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            tmp.write(buf.read()); tmp.close(); buf.close()
            return tmp.name
        except Exception:
            return None

    # ---- Utils ----
    def _get_ip(self) -> Optional[str]:
        try:
            from PySide6.QtWidgets import QLineEdit
            if self.main_window:
                ip_widget = self.main_window.findChild(QLineEdit, "ipEdit")
                if ip_widget:
                    return ip_widget.text().strip()
        except Exception:
            pass
        return None
