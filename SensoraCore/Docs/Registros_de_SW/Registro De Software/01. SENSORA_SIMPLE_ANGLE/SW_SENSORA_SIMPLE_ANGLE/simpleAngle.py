from PySide6.QtCore import QTimer, QThread, Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QLineEdit, QSizePolicy, QInputDialog, QMessageBox, QDialog, QFileDialog
from .simpleAngle_ui import Ui_simpleAngle
import socket
import time
import os
import json
import numpy as np
from datetime import datetime
import io
import base64
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

class SimpleAngleThread(QThread):
    data_received = Signal(int, int)  # Señal para recibir datos (lectura AD, ángulo)
    connection_status = Signal(str)  # Señal para mostrar estado de conexión

    def __init__(self, esp32_ip, port=8080):
        super().__init__()
        self.esp32_ip = esp32_ip
        self.port = port
        self.running = False
        self.sock = None
        self.last_data_time = 0  # Para controlar la frecuencia de datos

    def run(self):
        self.running = True
        try:
            print(f"Intentando conectar al ESP32 en {self.esp32_ip}:{self.port}")
            self.connection_status.emit(f"Conectando a {self.esp32_ip}:{self.port}...")
            
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.sock.settimeout(5)  # Aumentar timeout a 5 segundos
            
            # Intentar conectar
            self.sock.connect((self.esp32_ip, self.port))
            print("Conexión establecida, enviando comando MODO:ANGULO_SIMPLE")
            self.connection_status.emit("Conectado, iniciando monitoreo...")
            
            # Enviar comando y esperar respuesta
            self.sock.sendall(b'MODO:ANGULO_SIMPLE')
            
            # Esperar confirmación del ESP32
            response = self.sock.recv(64).decode('utf-8', errors='ignore').strip()
            print(f"Respuesta del ESP32: {response}")
            
            if 'ANGULO_SIMPLE_OK' in response:
                self.connection_status.emit("Monitoreo iniciado correctamente")
                self.sock.settimeout(2)  # Timeout de 2 segundos para muestreo de 0.5s
                
                while self.running:
                    try:
                        data = self.sock.recv(64)
                        if not data:
                            print("No se recibieron datos del ESP32")
                            break

                        msg = data.decode(errors='ignore').strip()
                        print(f"Datos recibidos: {msg}")  # Debug
                        
                        # Procesar solo la primera línea válida para evitar múltiples emisiones
                        lines = [line for line in msg.split('\n') if line.strip()]
                        for line in lines:
                            if line.startswith('POT:'):
                                try:
                                    # Control de tiempo para asegurar muestreo cada 0.5 segundos
                                    current_time = time.time()
                                    if current_time - self.last_data_time >= 0.4:  # Mínimo 0.4s entre muestras
                                        parts = line.replace('POT:', '').split(',ANG:')
                                        if len(parts) == 2:
                                            lectura = int(parts[0])
                                            angulo = int(parts[1])
                                            self.data_received.emit(lectura, angulo)
                                            self.last_data_time = current_time
                                            print(f"Datos procesados - Lectura: {lectura}, Ángulo: {angulo}")
                                            break  # Solo procesar la primera línea válida
                                    else:
                                        print(f"Datos ignorados - muy rápidos (delta: {current_time - self.last_data_time:.2f}s)")
                                except ValueError as ve:
                                    print(f"Error procesando datos: {ve}")
                                    
                    except socket.timeout:
                        print("Timeout esperando datos del ESP32 (normal con muestreo de 0.5s)")
                        continue
                    except Exception as e:
                        print(f"Error recibiendo datos: {e}")
                        break
            else:
                print(f"Respuesta inesperada del ESP32: {response}")
                self.connection_status.emit(f"Error: Respuesta inesperada del ESP32")
                
        except socket.timeout as e:
            print(f"Timeout de conexión al ESP32: {e}")
            self.connection_status.emit("Error: Timeout de conexión al ESP32")
        except ConnectionRefusedError as e:
            print(f"Conexión rechazada por el ESP32: {e}")
            self.connection_status.emit("Error: Conexión rechazada. Verificar que el ESP32 esté encendido")
        except Exception as e:
            print(f"Error en SimpleAngleThread: {e}")
            self.connection_status.emit(f"Error de conexión: {e}")
        finally:
            if self.sock:
                try:
                    self.sock.sendall(b'STOP')
                    print("Comando STOP enviado al ESP32")
                except:
                    pass
                self.sock.close()
                print("Socket cerrado")
            self.connection_status.emit("Desconectado")

    def stop(self):
        print("Deteniendo SimpleAngleThread...")
        self.running = False
        
        # Cerrar socket si está abierto para forzar la salida del loop
        if self.sock:
            try:
                self.sock.shutdown(socket.SHUT_RDWR)
            except:
                pass
            try:
                self.sock.close()
            except:
                pass
            self.sock = None
        
        # Esperar a que termine el hilo
        self.wait(3000)  # Timeout de 3 segundos

class SimpleAngleLogic(QWidget):
    def __init__(self, ui_widget, main_window=None):
        super().__init__()
        # Asegurar que el widget de calibración no esté visible hasta que se realice la calibración
        if hasattr(ui_widget, 'calBox') and ui_widget.calBox is not None:
            ui_widget.calBox.setVisible(False)
        try:
            print("Inicializando SimpleAngleLogic...")

            # Verificar que el widget UI es válido
            if ui_widget is None:
                raise ValueError("ui_widget no puede ser None")

            # Usar el widget de interfaz proporcionado
            self.ui = ui_widget
            self.main_window = main_window  # Referencia a la ventana principal
            self.connections = []  # Lista para rastrear conexiones de señales

            # Conectar botones de la interfaz
            if hasattr(self.ui, 'iniciar'):
                self.ui.iniciar.clicked.connect(self.toggle_angulo_monitoring)

            if hasattr(self.ui, 'calibrar'):
                self.ui.calibrar.clicked.connect(self.calibrar_sensor)

            if hasattr(self.ui, 'limpiar'):
                self.ui.limpiar.clicked.connect(self.limpiar_grafica)

            if hasattr(self.ui, 'exportar'):
                self.ui.exportar.clicked.connect(self.exportar_datos)
                # Inicialmente deshabilitado hasta que haya datos
                self.ui.exportar.setEnabled(False)

            # Inicializar variables
            self.timer = QTimer(self)  # Asignar parent para manejo automático de memoria
            self.timer.timeout.connect(self.actualizar_datos)
            self.monitoreo_activo = False
            self.lecturas = []  # Datos analógicos (0-4095)
            self.angulos = []   # Datos de ángulo (-135 a +135)
            self.tiempo = []    # Tiempo en segundos
            self.muestra_actual = 0  # Contador de muestras
            self.MAX_MUESTRAS = 25   # Límite de muestras en la gráfica (reducido de 50 a 25)
            # Para calibración
            self.last_lectura = None
            self.last_angulo = None
            self.calibrated = False
            self.cal_m = 1.0
            self.cal_b = 0.0
            self.cal_points = []  # lista de (lectura, angulo)
            # Intentar cargar calibración previa
            try:
                self._load_calibration()
            except Exception:
                pass

            # Configurar gráfica con dos ejes Y
            self.figure = Figure(figsize=(8, 5), tight_layout=True)
            self.canvas = FigureCanvas(self.figure)
            
            # Crear subplot con dos ejes Y
            self.ax1 = self.figure.add_subplot(111)
            self.ax2 = self.ax1.twinx()  # Segundo eje Y compartiendo el eje X
            
            # Configurar el primer eje (lecturas analógicas)
            self.ax1.set_xlabel("Tiempo (segundos)", fontsize=10)
            self.ax1.set_ylabel("Lectura Analógica (AD)", color='blue', fontsize=10)
            self.ax1.tick_params(axis='y', labelcolor='blue', labelsize=8)
            self.ax1.tick_params(axis='x', labelsize=8)
            self.ax1.set_ylim(0, 4095)  # Rango del AD
            self.ax1.grid(True, alpha=0.3)
            
            # Configurar el segundo eje (ángulos)
            self.ax2.set_ylabel("Ángulo (grados)", color='red', fontsize=10)
            self.ax2.tick_params(axis='y', labelcolor='red', labelsize=8)
            self.ax2.set_ylim(-135, 135)  # Rango del ángulo
            
            # Configurar título
            self.figure.suptitle("Sensor de Ángulo Simple - Tiempo Real", fontsize=12, fontweight='bold')
            
            # Líneas de la gráfica (original y calibrada)
            self.line1, = self.ax1.plot([], [], 'b-', label='Lectura AD', linewidth=2, alpha=0.8)
            # Ángulo original en rojo, línea discontinua (evita definir linestyle dos veces)
            self.line2, = self.ax2.plot([], [], 'r--', label='Ángulo Original', linewidth=1.5, alpha=0.6)
            self.line3, = self.ax2.plot([], [], 'g-', label='Ángulo Calibrado', linewidth=2, alpha=0.8)
            
            # Leyenda compacta
            lines1, labels1 = self.ax1.get_legend_handles_labels()
            lines2, labels2 = self.ax2.get_legend_handles_labels()
            self.ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize=9)
            
            # Ajustar márgenes para que se vea bien en el widget
            self.figure.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.15)

            # Variables para almacenar datos originales separados
            self.angulos_originales = []  # Ángulos sin calibración

            # Asegurar que grapWid tenga un layout y configurar el canvas
            if not self.ui.grapWid.layout():
                layout = QVBoxLayout()
                layout.setContentsMargins(5, 5, 5, 5)  # Pequeños márgenes
                layout.setSpacing(0)  # Sin espaciado entre widgets
                self.ui.grapWid.setLayout(layout)

            # Limpiar layout existente
            layout = self.ui.grapWid.layout()
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()

            # Configurar el canvas para que se adapte al widget
            self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            layout.addWidget(self.canvas)

            print("SimpleAngleLogic inicializado correctamente.")
        except Exception as e:
            print(f"Error al inicializar SimpleAngleLogic: {e}")
            raise

    def cleanup(self):
        """
        Limpia las conexiones y recursos antes de eliminar el objeto.
        Detiene cualquier proceso activo con el ESP32.
        """
        try:
            print("Limpiando SimpleAngleLogic...")
            
            # Detener el monitoreo si está activo
            if self.monitoreo_activo:
                print("Deteniendo monitoreo activo...")
                self.stop_angulo_monitoring()
            
            # Detener timer si está activo
            if hasattr(self, 'timer') and self.timer.isActive():
                print("Deteniendo timer...")
                self.timer.stop()
            
            # Detener y limpiar el hilo del ESP32 si existe
            if hasattr(self, 'thread_SIMPLE_ANGLE') and self.thread_SIMPLE_ANGLE is not None:
                print("Limpiando hilo de comunicación ESP32...")
                if self.thread_SIMPLE_ANGLE.isRunning():
                    self.thread_SIMPLE_ANGLE.stop()
                    # Esperar un breve momento para que el hilo termine
                    self.thread_SIMPLE_ANGLE.wait(3000)  # Timeout de 3 segundos
                self.thread_SIMPLE_ANGLE = None
            
            # Limpiar las listas de datos
            if hasattr(self, 'lecturas'):
                self.lecturas.clear()
            if hasattr(self, 'angulos'):
                self.angulos.clear()
            if hasattr(self, 'angulos_originales'):
                self.angulos_originales.clear()
            if hasattr(self, 'tiempo'):
                self.tiempo.clear()
            
            # Resetear variables de estado
            self.monitoreo_activo = False
            self.last_lectura = None
            self.last_angulo = None
            
            # Desconectar señales manualmente si es necesario
            self.connections.clear()
            
            print("Limpieza de SimpleAngleLogic completada.")
        except Exception as e:
            print(f"Error durante la limpieza: {e}")

    def closeEvent(self, event):
        """
        Maneja el evento de cierre del widget para limpiar recursos.
        """
        print("Cerrando widget SimpleAngle...")
        self.cleanup()
        super().closeEvent(event) if hasattr(super(), 'closeEvent') else None


    def calibrar_sensor(self):
        """
        Flujo interactivo de calibración por referencia con número personalizable de puntos.
        Se piden N puntos de referencia (ángulos conocidos). El usuario posiciona el potenciómetro,
        pulsa OK en el diálogo y se captura la lectura AD actual. Se calcula una regresión lineal
        (ángulo = m * lectura + b) y se guarda en disco. Al final muestra una gráfica de calibración.
        """
        try:
            # Si no hay monitoreo activo, pedir al usuario que lo inicie
            if not self.monitoreo_activo:
                QMessageBox.information(self, "Calibración", "Para calibrar, primero inicie el monitoreo y asegúrese de que llegan datos.")
                return

            # Si ya está calibrado, preguntar si desea recalibrar o resetear
            if self.calibrated:
                resp = QMessageBox.question(self, "Calibración", "Ya existe una calibración. ¿Desea reemplazarla?", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
                if resp != QMessageBox.Yes:
                    return

            # Preguntar cuántos puntos de calibración desea
            num_points, ok = QInputDialog.getInt(self, "Calibración", "¿Cuántos puntos de calibración desea usar?\n(Mínimo: 2, Máximo: 10)", value=3)
            if not ok:
                QMessageBox.information(self, "Calibración", "Calibración cancelada por el usuario.")
                return
                
            # Validar rango manualmente
            if num_points < 2:
                QMessageBox.warning(self, "Calibración", "Mínimo 2 puntos de calibración requeridos.")
                return
            elif num_points > 10:
                QMessageBox.warning(self, "Calibración", "Máximo 10 puntos de calibración permitidos.")
                return

            puntos = []
            # Sugerencias de ángulos de referencia basadas en la tabla (ángulo simple)
            sugeridos = [-90.0, -45.0, 0.0, 45.0, 90.0, -120.0, 120.0, -135.0, 135.0, -60.0]
            
            for i in range(1, num_points + 1):
                # Pedir al usuario el ángulo de referencia
                default = sugeridos[i-1] if i-1 < len(sugeridos) else 0.0
                angle, ok = QInputDialog.getDouble(self, f"Calibración - Punto {i} de {num_points}", 
                                                 f"Mueva el potenciómetro a la posición deseada y escriba el ángulo de referencia (grados):\n\nPunto {i}/{num_points}", 
                                                 value=default, decimals=1)
                if not ok:
                    QMessageBox.information(self, "Calibración", "Calibración cancelada por el usuario.")
                    return

                # Capturar la última lectura conocida
                if self.last_lectura is None:
                    QMessageBox.warning(self, "Calibración", "No se detectan lecturas. Asegúrese de que el monitoreo está activo y que el ESP32 envía datos.")
                    return

                lectura = int(self.last_lectura)
                puntos.append((lectura, float(angle)))
                print(f"Punto {i}: Lectura AD = {lectura}, Ángulo = {angle}°")

            # Calcular regresión lineal simple (y = m*x + b) con los puntos
            xs = [p[0] for p in puntos]
            ys = [p[1] for p in puntos]
            n = len(xs)
            mean_x = sum(xs) / n
            mean_y = sum(ys) / n
            num = sum((xs[i] - mean_x) * (ys[i] - mean_y) for i in range(n))
            den = sum((xs[i] - mean_x) ** 2 for i in range(n))
            if den == 0:
                QMessageBox.warning(self, "Calibración", "No se puede calcular calibración: varianza en lecturas = 0")
                return

            m = num / den
            b = mean_y - m * mean_x

            # Calcular R² (coeficiente de determinación)
            ss_res = sum((ys[i] - (m * xs[i] + b)) ** 2 for i in range(n))
            ss_tot = sum((ys[i] - mean_y) ** 2 for i in range(n))
            r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 1.0

            # Guardar parámetros
            self.cal_m = m
            self.cal_b = b
            self.cal_points = puntos
            self.calibrated = True
            self._save_calibration()

            # Actualizar UI
            if hasattr(self.ui, 'calibrar') and self.ui.calibrar is not None:
                self.ui.calibrar.setText("Calibrado")
                # color neutro/verde suave (evita saturar con colores intensos)
                self.ui.calibrar.setStyleSheet(
                    "font-size: 14px;"
                    "font-weight: bold;"
                    "color: rgb(20, 80, 20);"
                    "padding: 6px;"
                    "background-color: rgb(220, 255, 220);"
                    "border-radius: 4px;"
                    "border: 1px solid rgb(120, 180, 120);"
                )
            if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
                self.ui.calBox.setVisible(True)

            # Mostrar gráfica de calibración
            self._show_calibration_graph(puntos, m, b, r_squared)

            QMessageBox.information(self, "Calibración", f"Calibración completada con {num_points} puntos.\n\nEcuación: y = {m:.6f}x + {b:.3f}\nR² = {r_squared:.4f}")

        except Exception as e:
            print(f"Error en calibrar_sensor: {e}")
            QMessageBox.critical(self, "Error", f"Error durante la calibración: {e}")

    def apply_calibration(self, lectura, raw_angle=None):
        """
        Devuelve el ángulo calibrado usando la regresión calculada y recorta al rango físico (-135..135).
        Si no hay calibración, devuelve raw_angle (si se proporcionó) o None.
        """
        if not self.calibrated:
            return raw_angle
        ang = self.cal_m * lectura + self.cal_b
        # Clamp al rango físico del sensor (-135 a 135) según la tabla de sensores
        ang = max(-135.0, min(135.0, ang))
        return ang

    def _show_calibration_graph(self, puntos, m, b, r_squared):
        """
        Muestra una ventana con la gráfica de calibración, puntos capturados, línea de tendencia y ecuación.
        """
        try:
            # Crear ventana de diálogo
            dialog = QDialog(self)
            dialog.setWindowTitle("Resultados de Calibración")
            dialog.setModal(True)
            dialog.resize(600, 500)
            
            # Layout del diálogo
            layout = QVBoxLayout(dialog)
            
            # Crear figura de matplotlib
            fig = Figure(figsize=(8, 6), dpi=100)
            canvas = FigureCanvas(fig)
            
            # Crear subplot
            ax = fig.add_subplot(111)
            
            # Extraer datos
            xs = [p[0] for p in puntos]
            ys = [p[1] for p in puntos]
            
            # Rango de lecturas para la línea de tendencia
            x_min, x_max = min(xs) if xs else 0, max(xs) if xs else 4095
            x_range = np.linspace(x_min - 100, x_max + 100, 100)
            y_trend = m * x_range + b
            
            # Gráfica de puntos de calibración
            ax.scatter(xs, ys, color='red', s=100, alpha=0.8, label=f'Puntos de calibración ({len(puntos)})', zorder=3)
            
            # Línea de tendencia
            ax.plot(x_range, y_trend, 'b-', linewidth=2, alpha=0.7, label='Línea de tendencia', zorder=2)
            
            # Configuración de la gráfica
            ax.set_xlabel('Lectura Analógica (AD)', fontsize=12)
            ax.set_ylabel('Ángulo de Referencia (°)', fontsize=12)
            ax.set_title(f'Calibración del Sensor de Ángulo\ny = {m:.6f}x + {b:.3f}  (R² = {r_squared:.4f})', fontsize=14, fontweight='bold')
            ax.grid(True, alpha=0.3)
            ax.legend(fontsize=10)
            
            # Configurar límites
            ax.set_xlim(0, 4095)
            ax.set_ylim(-150, 150)
            
            # Añadir texto con la ecuación
            textstr = f'Ecuación: y = {m:.6f}x + {b:.3f}\nR² = {r_squared:.4f}\nPuntos: {len(puntos)}'
            props = dict(boxstyle='round', facecolor='wheat', alpha=0.8)
            ax.text(0.02, 0.98, textstr, transform=ax.transAxes, fontsize=10,
                   verticalalignment='top', bbox=props)
            
            fig.tight_layout()
            
            # Añadir canvas al layout
            layout.addWidget(canvas)
            
            # Mostrar diálogo
            dialog.exec()
            
        except Exception as e:
            print(f"Error mostrando gráfica de calibración: {e}")
            QMessageBox.warning(self, "Error", f"No se pudo mostrar la gráfica de calibración: {e}")

    def limpiar_grafica(self):
        try:
            # Lógica para limpiar la gráfica
            self.lecturas.clear()
            self.angulos.clear()
            self.angulos_originales.clear()
            self.tiempo.clear()
            self.muestra_actual = 0
            self.update_graph()

            if hasattr(self.ui, 'analogoDt'):
                self.ui.analogoDt.setText("--")

            if hasattr(self.ui, 'anguloDt'):
                self.ui.anguloDt.setText("--")
                
            # Limpiar también labels de calibración
            if hasattr(self.ui, 'analogoDtCal'):
                self.ui.analogoDtCal.setText("--")

            if hasattr(self.ui, 'anguloDtCal'):
                self.ui.anguloDtCal.setText("--")

            # Deshabilitar botón de exportar cuando no hay datos
            if hasattr(self.ui, 'exportar'):
                self.ui.exportar.setEnabled(False)

            print("Gráfica limpiada correctamente.")
        except RuntimeError as e:
            print(f"Error en limpiar_grafica: {e}")

    def exportar_datos(self):
        """
        Exporta los datos de la gráfica a un archivo Excel (.xlsx) con formato profesional.
        Incluye datos originales, calibrados (si aplica), metadatos y parámetros de calibración.
        """
        try:
            # Verificar que hay datos para exportar
            if len(self.lecturas) == 0:
                QMessageBox.information(self, "Exportar", "No hay datos para exportar. Inicie el monitoreo primero.")
                return

            # Diálogo para guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"SensorAngulo_Simple_{timestamp}.xlsx"
            
            filename, _ = QFileDialog.getSaveFileName(
                self, 
                "Exportar datos a Excel", 
                default_filename,
                "Archivos Excel (*.xlsx);;Archivos CSV (*.csv);;Todos los archivos (*.*)"
            )
            
            if not filename:
                return  # Usuario canceló
            
            # Determinar formato basado en extensión
            export_format = 'xlsx' if filename.lower().endswith('.xlsx') else 'csv'
            
            if export_format == 'xlsx':
                self._export_to_excel(filename)
            else:
                self._export_to_csv(filename)
                
            QMessageBox.information(self, "Exportar", f"Datos exportados exitosamente a:\n{filename}")
            
        except Exception as e:
            print(f"Error en exportar_datos: {e}")
            QMessageBox.critical(self, "Error", f"Error al exportar datos: {e}")

    def _export_to_excel(self, filename):
        """
        Exporta los datos a Excel con formato profesional y múltiples hojas.
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Crear workbook
            wb = Workbook()
            
            # === HOJA 1: DATOS PRINCIPALES ===
            ws_data = wb.active
            ws_data.title = "Datos del Sensor"
            
            # Preparar datos principales
            data = {
                'Tiempo (s)': self.tiempo.copy(),
                'Lectura_AD': self.lecturas.copy(),
                'Angulo_Original (°)': self.angulos_originales.copy() if self.angulos_originales else [0] * len(self.tiempo)
            }
            
            # Añadir datos calibrados si existen
            if self.calibrated and len(self.angulos) > 0:
                # Recalcular ángulos calibrados para asegurar precisión
                angulos_cal = [self.apply_calibration(lectura) for lectura in self.lecturas]
                data['Angulo_Calibrado (°)'] = angulos_cal
                data['Diferencia (°)'] = [cal - orig for cal, orig in zip(angulos_cal, data['Angulo_Original (°)'])]
            
            # Crear DataFrame
            df = pd.DataFrame(data)
            
            # Escribir datos
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)
            
            # Formatear encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws_data[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Autoajustar columnas
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # === HOJA 2: METADATOS Y CONFIGURACIÓN ===
            ws_meta = wb.create_sheet("Metadatos")
            
            # Información general
            metadata = [
                ["INFORMACIÓN DEL EXPERIMENTO", ""],
                ["Fecha de exportación", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["Sensor", "Ángulo Simple (Potenciómetro)"],
                ["Muestras totales", len(self.lecturas)],
                ["Intervalo de muestreo", "0.5 segundos"],
                ["Tiempo total", f"{len(self.tiempo) * 0.5:.1f} segundos"],
                ["", ""],
                ["RANGO DE DATOS", ""],
                ["Lectura AD mínima", min(self.lecturas) if self.lecturas else 0],
                ["Lectura AD máxima", max(self.lecturas) if self.lecturas else 0],
                ["Ángulo original mínimo (°)", min(self.angulos_originales) if self.angulos_originales else 0],
                ["Ángulo original máximo (°)", max(self.angulos_originales) if self.angulos_originales else 0],
            ]
            
            # Información de calibración
            if self.calibrated:
                metadata.extend([
                    ["", ""],
                    ["CALIBRACIÓN APLICADA", ""],
                    ["Estado", "Calibrado"],
                    ["Ecuación", f"y = {self.cal_m:.6f}x + {self.cal_b:.3f}"],
                    ["Pendiente (m)", f"{self.cal_m:.6f}"],
                    ["Intersección (b)", f"{self.cal_b:.3f}"],
                    ["Puntos de calibración", len(self.cal_points) if self.cal_points else 0],
                ])
                
                if self.angulos:
                    metadata.extend([
                        ["Ángulo calibrado mínimo (°)", f"{min(self.angulos):.2f}"],
                        ["Ángulo calibrado máximo (°)", f"{max(self.angulos):.2f}"],
                    ])
            else:
                metadata.extend([
                    ["", ""],
                    ["CALIBRACIÓN", ""],
                    ["Estado", "No calibrado"],
                ])
            
            # Escribir metadatos
            for row in metadata:
                ws_meta.append(row)
            
            # Formatear metadatos
            for row in ws_meta.iter_rows():
                if row[0].value and str(row[0].value).isupper():  # Títulos en mayúsculas
                    row[0].font = Font(bold=True, size=12)
                    row[0].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            ws_meta.column_dimensions['A'].width = 25
            ws_meta.column_dimensions['B'].width = 20
            
            # === HOJA 3: PUNTOS DE CALIBRACIÓN (si existen) ===
            if self.calibrated and self.cal_points:
                ws_cal = wb.create_sheet("Puntos Calibración")
                
                # Encabezados
                ws_cal.append(["Punto", "Lectura AD", "Ángulo Referencia (°)", "Ángulo Calculado (°)", "Error (°)"])
                
                # Datos de calibración
                for i, (lectura, angulo_ref) in enumerate(self.cal_points, 1):
                    angulo_calc = self.cal_m * lectura + self.cal_b
                    error = angulo_calc - angulo_ref
                    ws_cal.append([i, lectura, angulo_ref, round(angulo_calc, 3), round(error, 3)])
                
                # Formatear
                for cell in ws_cal[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Autoajustar columnas
                for column in ws_cal.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 18)
                    ws_cal.column_dimensions[column_letter].width = adjusted_width
            
            # === HOJA 4: GRÁFICA ===
            ws_graph = wb.create_sheet("Gráfica")
            
            # Crear y guardar la gráfica
            graph_image = self._create_export_graph()
            if graph_image:
                from openpyxl.drawing.image import Image
                
                # Crear imagen para Excel
                img = Image(graph_image)
                # Ajustar tamaño (opcional)
                img.width = 800
                img.height = 600
                
                # Agregar título
                ws_graph['A1'] = "GRÁFICA DEL SENSOR DE ÁNGULO SIMPLE"
                ws_graph['A1'].font = Font(bold=True, size=16, color="2F5597")
                ws_graph['A1'].alignment = Alignment(horizontal="center")
                
                # Información de la gráfica
                ws_graph['A3'] = f"Generada: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws_graph['A4'] = f"Muestras mostradas: {len(self.lecturas)}"
                if self.calibrated:
                    ws_graph['A5'] = f"Calibración aplicada: y = {self.cal_m:.6f}x + {self.cal_b:.3f}"
                else:
                    ws_graph['A5'] = "Sin calibración aplicada"
                
                # Insertar imagen debajo del texto
                ws_graph.add_image(img, 'A7')
                
                # Ajustar columnas
                ws_graph.column_dimensions['A'].width = 50
            
            # Guardar archivo
            wb.save(filename)
            
            # Limpiar archivo temporal de la gráfica
            if graph_image:
                try:
                    os.unlink(graph_image)  # Eliminar archivo temporal
                except:
                    pass  # Ignorar errores de limpieza
            print(f"Datos exportados a Excel: {filename}")
            
        except ImportError:
            # Fallback a CSV si no está disponible pandas/openpyxl
            QMessageBox.warning(self, "Exportar", "Las librerías de Excel no están disponibles. Exportando a CSV...")
            csv_filename = filename.replace('.xlsx', '.csv')
            self._export_to_csv(csv_filename)
        except Exception as e:
            raise Exception(f"Error al crear archivo Excel: {e}")

    def _export_to_csv(self, filename):
        """
        Exporta los datos a CSV como alternativa cuando Excel no está disponible.
        """
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                # Encabezado del archivo
                csvfile.write(f"# Datos del Sensor de Ángulo Simple\n")
                csvfile.write(f"# Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                csvfile.write(f"# Muestras: {len(self.lecturas)}\n")
                csvfile.write(f"# Intervalo: 0.5 segundos\n")
                
                if self.calibrated:
                    csvfile.write(f"# Calibración: y = {self.cal_m:.6f}x + {self.cal_b:.3f}\n")
                else:
                    csvfile.write(f"# Calibración: No aplicada\n")
                    
                csvfile.write("#\n")
                csvfile.write("# NOTA: Para ver la gráfica completa, use exportación a Excel (.xlsx)\n")
                csvfile.write("#\n")
                
                # Encabezados de columnas
                headers = ["Tiempo_s", "Lectura_AD", "Angulo_Original_grados"]
                if self.calibrated and len(self.angulos) > 0:
                    headers.extend(["Angulo_Calibrado_grados", "Diferencia_grados"])
                
                csvfile.write(",".join(headers) + "\n")
                
                # Datos
                for i in range(len(self.tiempo)):
                    row = [
                        f"{self.tiempo[i]:.1f}",
                        str(self.lecturas[i]),
                        str(self.angulos_originales[i] if i < len(self.angulos_originales) else 0)
                    ]
                    
                    if self.calibrated and i < len(self.angulos):
                        angulo_cal = self.apply_calibration(self.lecturas[i])
                        diferencia = angulo_cal - (self.angulos_originales[i] if i < len(self.angulos_originales) else 0)
                        row.extend([f"{angulo_cal:.2f}", f"{diferencia:.2f}"])
                    
                    csvfile.write(",".join(row) + "\n")
            
            print(f"Datos exportados a CSV: {filename}")
            
        except Exception as e:
            raise Exception(f"Error al crear archivo CSV: {e}")

    def _create_export_graph(self):
        """
        Crea una gráfica de alta calidad para exportar a Excel.
        Retorna el path de la imagen temporal o None si hay error.
        """
        try:
            # Crear figura de alta resolución para exportación
            fig = Figure(figsize=(12, 8), dpi=150, facecolor='white')
            
            # Crear subplot con dos ejes Y (igual que la gráfica principal)
            ax1 = fig.add_subplot(111)
            ax2 = ax1.twinx()
            
            # Configurar el primer eje (lecturas analógicas)
            ax1.set_xlabel("Tiempo (segundos)", fontsize=14, fontweight='bold')
            ax1.set_ylabel("Lectura Analógica (AD)", color='blue', fontsize=14, fontweight='bold')
            ax1.tick_params(axis='y', labelcolor='blue', labelsize=12)
            ax1.tick_params(axis='x', labelsize=12)
            ax1.set_ylim(0, 4095)
            ax1.grid(True, alpha=0.3, linestyle='--')
            
            # Configurar el segundo eje (ángulos)
            ax2.set_ylabel("Ángulo (grados)", color='red', fontsize=14, fontweight='bold')
            ax2.tick_params(axis='y', labelcolor='red', labelsize=12)
            ax2.set_ylim(-135, 135)
            
            # Configurar título
            title = "Sensor de Ángulo Simple - Datos Exportados"
            if self.calibrated:
                title += f"\nCalibración: y = {self.cal_m:.6f}x + {self.cal_b:.3f}"
            fig.suptitle(title, fontsize=16, fontweight='bold', y=0.95)
            
            # Plotear datos si existen
            if len(self.tiempo) > 0:
                # Línea de lecturas AD
                line1 = ax1.plot(self.tiempo, self.lecturas, 'b-', label='Lectura AD', 
                                linewidth=2.5, alpha=0.8, marker='o', markersize=3)
                
                # Línea de ángulos originales
                if len(self.angulos_originales) > 0:
                    line2 = ax2.plot(self.tiempo, self.angulos_originales, 'r--', 
                                    label='Ángulo Original', linewidth=2, alpha=0.7,
                                    marker='s', markersize=3)
                
                # Línea de ángulos calibrados (si existe calibración)
                if self.calibrated and len(self.angulos) > 0:
                    line3 = ax2.plot(self.tiempo, self.angulos, 'g-', 
                                    label='Ángulo Calibrado', linewidth=2.5, alpha=0.9,
                                    marker='^', markersize=3)
                
                # Configurar límites del eje X
                if len(self.tiempo) > 1:
                    x_min = min(self.tiempo)
                    x_max = max(self.tiempo)
                    margen = (x_max - x_min) * 0.05 if x_max > x_min else 1
                    ax1.set_xlim(x_min - margen, x_max + margen)
                else:
                    ax1.set_xlim(0, 1)
            
            # Crear leyenda combinada
            lines1, labels1 = ax1.get_legend_handles_labels()
            lines2, labels2 = ax2.get_legend_handles_labels()
            ax1.legend(lines1 + lines2, labels1 + labels2, 
                      loc='upper left', fontsize=12, framealpha=0.9)
            
            # Añadir información adicional
            info_text = f"Muestras: {len(self.lecturas)} | Intervalo: 0.5s | "
            info_text += f"Tiempo total: {len(self.tiempo) * 0.5:.1f}s"
            fig.text(0.02, 0.02, info_text, fontsize=10, style='italic', alpha=0.7)
            
            # Ajustar layout
            fig.tight_layout()
            
            # Guardar en memoria como imagen
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            img_buffer.seek(0)
            
            # Crear archivo temporal
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            temp_file.write(img_buffer.getvalue())
            temp_file.close()
            
            # Limpiar
            img_buffer.close()
            fig.clear()
            
            return temp_file.name
            
        except Exception as e:
            print(f"Error creando gráfica de exportación: {e}")
            return None

    def actualizar_datos(self):
        try:
            # Lógica para actualizar los datos en la interfaz
            if hasattr(self.ui, 'analogoDt') and self.ui.analogoDt is not None:
                self.ui.analogoDt.setText("123")  # Ejemplo de datos
            if hasattr(self.ui, 'anguloDt') and self.ui.anguloDt is not None:
                self.ui.anguloDt.setText("45°")  # Ejemplo de datos
        except RuntimeError as e:
            print(f"Error en actualizar_datos: {e}")
            # Si hay error, detener el monitoreo
            self.monitoreo_activo = False
            self.timer.stop()

    def toggle_angulo_monitoring(self):
        """
        Alterna entre iniciar y detener el monitoreo del sensor de ángulo simple.
        """
        if not self.monitoreo_activo:
            self.start_angulo_monitoring()
        else:
            self.stop_angulo_monitoring()

    def start_angulo_monitoring(self):
        """
        Inicia el monitoreo en tiempo real del sensor de ángulo simple.
        """
        try:
            # Detener hilo anterior si existe
            if hasattr(self, 'thread_SIMPLE_ANGLE') and self.thread_SIMPLE_ANGLE is not None:
                if self.thread_SIMPLE_ANGLE.isRunning():
                    self.thread_SIMPLE_ANGLE.stop()
                    self.thread_SIMPLE_ANGLE = None

            # Obtener la IP del ESP32 desde la ventana principal
            esp32_ip = None
            if self.main_window:
                ip_widget = self.main_window.findChild(QLineEdit, "ipEdit")
                if ip_widget:
                    esp32_ip = ip_widget.text().strip()
            
            if not esp32_ip:
                esp32_ip = "192.168.1.100"  # IP por defecto
                print("No se pudo obtener la IP desde main.py, usando IP por defecto")
            
            print(f"Iniciando conexión con ESP32 en IP: {esp32_ip}")
            
            self.thread_SIMPLE_ANGLE = SimpleAngleThread(esp32_ip)
            self.thread_SIMPLE_ANGLE.data_received.connect(self.update_angulo_data)
            self.thread_SIMPLE_ANGLE.connection_status.connect(self.update_connection_status)

            self.thread_SIMPLE_ANGLE.start()
            self.monitoreo_activo = True

            # Actualizar el botón de inicio específico de simpleAngle
            if hasattr(self.ui, 'iniciar'):
                self.ui.iniciar.setText("Pausar")
                self.ui.iniciar.setStyleSheet(
                    "background-color: #e53935;"   # Rojo principal
                    "border-color: #b71c1c;"       # Rojo oscuro para el borde
                    "color: white;"
                    "padding: 8px;"
                    "font-size: 14px;"
                    "font-weight: bold;"
                    "border-radius: 4px;"
                    "border: 1px solid #b71c1c;"
                    "margin-top: 5px;"
                )
                
        except Exception as e:
            print(f"Error al iniciar monitoreo: {e}")
            self.monitoreo_activo = False
            if hasattr(self.ui, 'iniciar'):
                self.ui.iniciar.setText("Iniciar Monitoreo")
                self.ui.iniciar.setStyleSheet(
                    "background-color: #2196f3;"   # Azul principal
                    "border-color: #1976d2;"       # Azul oscuro para el borde
                    "color: white;"
                    "padding: 8px;"
                    "font-size: 14px;"
                    "font-weight: bold;"
                    "border-radius: 4px;"
                    "border: 1px solid #1976d2;"
                    "margin-top: 5px;"
                )
    def update_connection_status(self, status):
        """
        Actualiza el estado de conexión en la interfaz.
        """
        print(f"Estado de conexión: {status}")
        # Aquí puedes agregar un label en la UI para mostrar el estado si lo deseas

    def stop_angulo_monitoring(self):
        """
        Detiene el monitoreo del sensor de ángulo simple y limpia recursos.
        Asegura que se envíe comando STOP al ESP32 y se cierre la conexión.
        """
        try:
            print("Deteniendo monitoreo del sensor...")
            
            # Marcar como no activo inmediatamente
            self.monitoreo_activo = False
            
            if hasattr(self, 'thread_SIMPLE_ANGLE') and self.thread_SIMPLE_ANGLE is not None:
                if self.thread_SIMPLE_ANGLE.isRunning():
                    print("Deteniendo hilo de monitoreo...")
                    
                    # Enviar comando STOP al ESP32 antes de cerrar
                    try:
                        if hasattr(self.thread_SIMPLE_ANGLE, 'sock') and self.thread_SIMPLE_ANGLE.sock:
                            self.thread_SIMPLE_ANGLE.sock.sendall(b'STOP')
                            print("Comando STOP enviado al ESP32")
                    except Exception as e:
                        print(f"Error enviando comando STOP: {e}")
                    
                    # Detener el hilo de manera segura
                    self.thread_SIMPLE_ANGLE.stop()
                    
                    # Esperar a que termine con timeout
                    if not self.thread_SIMPLE_ANGLE.wait(5000):  # 5 segundos de timeout
                        print("Advertencia: El hilo no terminó en el tiempo esperado")
                        # Forzar terminación si es necesario
                        try:
                            self.thread_SIMPLE_ANGLE.terminate()
                            self.thread_SIMPLE_ANGLE.wait(2000)  # 2 segundos adicionales
                        except:
                            pass
                
                # Limpiar referencia
                self.thread_SIMPLE_ANGLE = None
                print("Hilo de monitoreo detenido y limpiado")

            # Restaurar el botón de inicio específico de simpleAngle
            if hasattr(self.ui, 'iniciar'):
                self.ui.iniciar.setText("Iniciar Monitoreo")
                self.ui.iniciar.setStyleSheet(
                    "background-color: #2196f3;"   # Azul principal
                    "border-color: #1976d2;"       # Azul oscuro para el borde
                    "color: white;"
                    "padding: 8px;"
                    "font-size: 14px;"
                    "font-weight: bold;"
                    "border-radius: 4px;"
                    "border: 1px solid #1976d2;"
                    "margin-top: 5px;"
                )

            # Limpiar datos de la última lectura
            self.last_lectura = None
            self.last_angulo = None

            print("Monitoreo detenido correctamente")
        except Exception as e:
            print(f"Error al detener monitoreo: {e}")
            # Asegurar que el estado se marque como no activo aunque haya errores
            self.monitoreo_activo = False

    # ----------------------
    # Calibración: persistencia
    # ----------------------
    def _cal_file_path(self):
        # Guardar junto al módulo (por simplicidad)
        base = os.path.dirname(__file__)
        return os.path.join(base, 'simpleAngle_calibration.json')

    def _save_calibration(self):
        path = self._cal_file_path()
        payload = {
            'm': self.cal_m,
            'b': self.cal_b,
            'points': self.cal_points
        }
        try:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(payload, f, indent=2)
            print(f"Calibración guardada en {path}")
        except Exception as e:
            print(f"No se pudo guardar calibración: {e}")

    def _load_calibration(self):
        path = self._cal_file_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            self.cal_m = float(payload.get('m', 1.0))
            self.cal_b = float(payload.get('b', 0.0))
            self.cal_points = payload.get('points', [])
            self.calibrated = True
            # Update UI markers if available
            if hasattr(self.ui, 'calibrar') and self.ui.calibrar is not None:
                self.ui.calibrar.setText("Calibrado")
            if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
                self.ui.calBox.setVisible(True)
            print(f"Calibración cargada desde {path}: m={self.cal_m}, b={self.cal_b}")
        except Exception as e:
            print(f"No se pudo cargar calibración: {e}")

    def _reset_calibration(self):
        self.cal_m = 1.0
        self.cal_b = 0.0
        self.cal_points = []
        self.calibrated = False
        path = self._cal_file_path()
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass
        # Reset UI
        if hasattr(self.ui, 'calibrar') and self.ui.calibrar is not None:
            self.ui.calibrar.setText("No Calibrado")
            try:
                self.ui.calibrar.setStyleSheet("")
            except Exception:
                pass
        if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
            self.ui.calBox.setVisible(False)

    def update_angulo_data(self, lectura, angulo):
        """
        Procesa y actualiza los datos recibidos del sensor de ángulo simple.
        Implementa ventana deslizante de 25 muestras con muestreo cada 0.5 segundos.
        Almacena tanto datos originales como calibrados por separado.
        """
        try:
            # Incrementar contador de muestras
            self.muestra_actual += 1

            # Guardar última lectura cruda para calibración
            self.last_lectura = lectura
            self.last_angulo = angulo

            # Aplicar calibración si está disponible (ángulo_cal = m * lectura + b)
            if self.calibrated:
                angulo_cal = self.apply_calibration(lectura, angulo)
            else:
                angulo_cal = angulo

            # Calcular tiempo en segundos (cada muestra es cada 0.5 segundos)
            tiempo_actual = self.muestra_actual * 0.5

            # Agregar nuevos datos
            self.lecturas.append(lectura)
            # Almacenar ambos: ángulo original y calibrado
            self.angulos_originales.append(angulo)
            self.angulos.append(int(round(angulo_cal)))
            self.tiempo.append(tiempo_actual)

            # Implementar ventana deslizante de 25 muestras
            if len(self.lecturas) > self.MAX_MUESTRAS:
                self.lecturas.pop(0)  # Eliminar el más antiguo
                self.angulos_originales.pop(0)  # Eliminar el más antiguo
                self.angulos.pop(0)   # Eliminar el más antiguo
                self.tiempo.pop(0)    # Eliminar el más antiguo

            # Actualizar displays de la interfaz
            if hasattr(self.ui, 'analogoDt'):
                self.ui.analogoDt.setText(str(lectura))

            if hasattr(self.ui, 'anguloDt'):
                # mostrar ángulo original (sin calibración)
                self.ui.anguloDt.setText(f"{angulo}°")
                
            # Actualizar labels de calibración si existe calibración
            if self.calibrated:
                if hasattr(self.ui, 'analogoDtCal'):
                    self.ui.analogoDtCal.setText(str(lectura))
                if hasattr(self.ui, 'anguloDtCal'):
                    self.ui.anguloDtCal.setText(f"{angulo_cal:.1f}°")

            # Actualizar gráfica
            self.update_graph()
            
            # Habilitar botón de exportar si hay datos
            if hasattr(self.ui, 'exportar'):
                self.ui.exportar.setEnabled(len(self.lecturas) > 0)
            
            print(f"Datos actualizados - Muestra: {self.muestra_actual}, Tiempo: {tiempo_actual:.1f}s, AD: {lectura}, Ángulo: {angulo}°, Muestras en gráfica: {len(self.lecturas)}")
            
        except RuntimeError as e:
            print(f"Error al actualizar datos: {e}")

    def update_graph(self):
        """
        Actualiza la gráfica con los datos actuales.
        Muestra lectura analógica, ángulo original y ángulo calibrado en ejes Y separados.
        """
        try:
            if len(self.tiempo) == 0:
                # Si no hay datos, limpiar las líneas
                self.line1.set_data([], [])
                self.line2.set_data([], [])
                self.line3.set_data([], [])
            else:
                # Actualizar datos de las líneas
                self.line1.set_data(self.tiempo, self.lecturas)
                
                # Mostrar datos originales y calibrados
                if len(self.angulos_originales) > 0:
                    self.line2.set_data(self.tiempo, self.angulos_originales)
                
                if self.calibrated and len(self.angulos) > 0:
                    self.line3.set_data(self.tiempo, self.angulos)
                    self.line3.set_visible(True)
                else:
                    self.line3.set_visible(False)
                
                # Ajustar rango del eje X dinámicamente
                if len(self.tiempo) > 1:
                    x_min = min(self.tiempo)
                    x_max = max(self.tiempo)
                    # Agregar un pequeño margen
                    margen = (x_max - x_min) * 0.05 if x_max > x_min else 1
                    self.ax1.set_xlim(x_min - margen, x_max + margen)
                else:
                    self.ax1.set_xlim(0, 1)
            
            # Redibujar la gráfica
            self.canvas.draw()
            
        except Exception as e:
            print(f"Error al actualizar gráfica: {e}")




