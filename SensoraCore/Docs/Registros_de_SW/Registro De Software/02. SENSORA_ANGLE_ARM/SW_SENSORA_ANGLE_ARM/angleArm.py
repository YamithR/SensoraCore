from PySide6.QtCore import QTimer, QThread, Signal
from PySide6.QtWidgets import QWidget, QVBoxLayout, QSizePolicy, QInputDialog, QMessageBox, QDialog, QFileDialog
from .angleArm_ui import Ui_angleArm  # Import sólo para tipos/nombres (la UI real viene de QUiLoader)
import socket
import time
import os
import json
import io
from datetime import datetime
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas


class AngleArmThread(QThread):
    data_received = Signal(int, int, int, int, int, int, bool)  # lect1,ang1, lect2,ang2, lect3,ang3, sensor
    connection_status = Signal(str)

    def __init__(self, esp32_ip, port=8080):
        super().__init__()
        self.esp32_ip = esp32_ip
        self.port = port
        self.running = False
        self.sock = None

    def run(self):
        self.running = True
        try:
            self.connection_status.emit(f"Conectando a {self.esp32_ip}:{self.port}...")
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.sock.settimeout(5)
            self.sock.connect((self.esp32_ip, self.port))
            # Iniciar modo continuo
            self.sock.sendall(b'MODO:BRAZO_ANGULO')
            resp = self.sock.recv(64).decode('utf-8', errors='ignore').strip()
            if 'BRAZO_ANGULO_OK' not in resp:
                self.connection_status.emit("Error: ESP32 no acepto modo BRAZO_ANGULO")
                return
            self.connection_status.emit("Monitoreo BRAZO_ANGULO iniciado")
            self.sock.settimeout(3)

            buffer = ""
            while self.running:
                try:
                    chunk = self.sock.recv(256)
                    if not chunk:
                        break
                    buffer += chunk.decode('utf-8', errors='ignore')
                    # Procesar por líneas
                    while '\n' in buffer:
                        line, buffer = buffer.split('\n', 1)
                        line = line.strip()
                        if not line:
                            continue
                        # Esperado: POT1:x,ANG1:y,POT2:x,ANG2:y,POT3:x,ANG3:y,SENSOR:True/False
                        try:
                            parts = line.split(',')
                            kv = {}
                            for p in parts:
                                if ':' in p:
                                    k, v = p.split(':', 1)
                                    kv[k.strip()] = v.strip()
                            l1 = int(kv.get('POT1', '0'))
                            a1 = int(kv.get('ANG1', '0'))
                            l2 = int(kv.get('POT2', '0'))
                            a2 = int(kv.get('ANG2', '0'))
                            l3 = int(kv.get('POT3', '0'))
                            a3 = int(kv.get('ANG3', '0'))
                            s = kv.get('SENSOR', 'False').lower() in ('true', '1', 'on')
                            self.data_received.emit(l1, a1, l2, a2, l3, a3, s)
                        except Exception as pe:
                            # Ignorar líneas malformadas
                            print(f"Parse error AngleArmThread: {pe} line={line}")
                except socket.timeout:
                    continue
                except Exception as e:
                    print(f"Error de socket AngleArmThread: {e}")
                    break
        except Exception as e:
            self.connection_status.emit(f"Error de conexión: {e}")
        finally:
            try:
                if self.sock:
                    try:
                        self.sock.sendall(b'STOP')
                    except Exception:
                        pass
                    try:
                        self.sock.close()
                    except Exception:
                        pass
            finally:
                self.sock = None
                self.connection_status.emit("Desconectado")

    def stop(self):
        self.running = False
        if self.sock:
            try:
                self.sock.shutdown(socket.SHUT_RDWR)
            except Exception:
                pass
            try:
                self.sock.close()
            except Exception:
                pass
        self.wait(2000)


class AngleArmLogic(QWidget):
    """
    Lógica del módulo Brazo Robótico (3 potenciómetros + sensor capacitivo digital).
    - Recibe stream del ESP32 (MODO:BRAZO_ANGULO)
    - Muestra lecturas AD y ángulos por potenciómetro
    - Calibración lineal independiente por canal (y = m*x + b)
    - Persistencia JSON por módulo
    """

    def __init__(self, ui_widget, main_window=None):
        super().__init__()
        if ui_widget is None:
            raise ValueError("ui_widget no puede ser None")
        self.ui = ui_widget
        self.main_window = main_window

        # Conectar botones
        if hasattr(self.ui, 'iniciarBt'):
            self.ui.iniciarBt.clicked.connect(self.toggle_monitoring)
        if hasattr(self.ui, 'limpiarBt'):
            self.ui.limpiarBt.clicked.connect(self.limpiar)
        if hasattr(self.ui, 'calibrarBt'):
            self.ui.calibrarBt.clicked.connect(self.calibrar)
        if hasattr(self.ui, 'exportarBt'):
            self.ui.exportarBt.clicked.connect(self.exportar)

        # Estado
        self.thread = None
        self.monitoreo_activo = False
        self.last_l = [None, None, None]
        self.last_a = [None, None, None]
        self.sensor_on = False

        # Calibración por canal
        self.cal = {
            0: {'m': 1.0, 'b': 0.0, 'ok': False, 'points': []},
            1: {'m': 1.0, 'b': 0.0, 'ok': False, 'points': []},
            2: {'m': 1.0, 'b': 0.0, 'ok': False, 'points': []},
        }
        self._load_calibration()

        # Gráfica básica: 2 ejes Y, 3 líneas por eje (AD y Ángulo Calibrado)
        self.figure = Figure(figsize=(8, 5), tight_layout=True)
        self.canvas = FigureCanvas(self.figure)
        self.ax1 = self.figure.add_subplot(111)  # AD
        self.ax2 = self.ax1.twinx()              # Ángulo
        self.ax1.set_xlabel("Tiempo (s)")
        self.ax1.set_ylabel("Lectura AD")
        self.ax1.set_ylim(0, 4095)
        self.ax2.set_ylabel("Ángulo (°)")
        self.ax2.set_ylim(-135, 135)
        self.figure.suptitle("Brazo: 3 Potenciómetros + Sensor Capacitivo")

        # Series
        self.time_idx = []
        self.ad_series = [[], [], []]
        self.ang_series = [[], [], []]
        # Series de ángulos originales (del firmware), para exportación
        self.raw_ang_series = [[], [], []]
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c']  # azul, naranja, verde
        self.lines_ad = [self.ax1.plot([], [], color=colors[i], linestyle='-', label=f'AD{i+1}')[0] for i in range(3)]
        self.lines_ang = [self.ax2.plot([], [], color=colors[i], linestyle='--', label=f'Ang{i+1}')[0] for i in range(3)]
        l1, lb1 = self.ax1.get_legend_handles_labels()
        l2, lb2 = self.ax2.get_legend_handles_labels()
        self.ax1.legend(l1 + l2, lb1 + lb2, loc='upper left', fontsize=8)

        # Embeder canvas en la UI
        if hasattr(self.ui, 'grapWid') and self.ui.grapWid is not None:
            if not self.ui.grapWid.layout():
                lay = QVBoxLayout()
                lay.setContentsMargins(5, 5, 5, 5)
                lay.setSpacing(0)
                self.ui.grapWid.setLayout(lay)
            lay = self.ui.grapWid.layout()
            while lay.count():
                ch = lay.takeAt(0)
                if ch.widget():
                    ch.widget().deleteLater()
            self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            lay.addWidget(self.canvas)

        # Mostrar calBox si ya hay calibración
        if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
            any_cal = any(self.cal[i]['ok'] for i in range(3))
            self.ui.calBox.setVisible(any_cal)

        # Timer de UI (solo para refrescar gráfico si quisiéramos agregarlo por polling externo)
        self.timer = QTimer(self)
        self.timer.setInterval(0)  # no usado por ahora

    # ---- Control ----
    def toggle_monitoring(self):
        if not self.monitoreo_activo:
            self.start()
        else:
            self.stop()

    def start(self):
        try:
            esp32_ip = self._get_ip()
            if not esp32_ip:
                QMessageBox.warning(self, "ESP32", "IP no definida. Escribe la IP en el campo correspondiente.")
                return
            # detener previo
            if self.thread and self.thread.isRunning():
                self.thread.stop()
                self.thread = None
            self.thread = AngleArmThread(esp32_ip)
            self.thread.data_received.connect(self._on_data)
            self.thread.connection_status.connect(self._on_status)
            self.thread.start()
            self.monitoreo_activo = True
            if hasattr(self.ui, 'iniciarBt'):
                self.ui.iniciarBt.setText("Pausar")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"No se pudo iniciar monitoreo: {e}")

    def stop(self):
        try:
            self.monitoreo_activo = False
            if self.thread:
                if self.thread.isRunning():
                    self.thread.stop()
                self.thread = None
            if hasattr(self.ui, 'iniciarBt'):
                self.ui.iniciarBt.setText("Iniciar Monitoreo")
        except Exception as e:
            print(f"Error al detener monitoreo: {e}")

    def cleanup(self):
        self.stop()

    # ---- Datos ----
    def _on_status(self, msg: str):
        print(f"[AngleArm] {msg}")

    def _apply_cal(self, idx: int, lectura: int, raw_angle: int | None) -> float:
        if not self.cal[idx]['ok']:
            return float(raw_angle if raw_angle is not None else 0)
        ang = self.cal[idx]['m'] * lectura + self.cal[idx]['b']
        # Clamp físico
        return max(-135.0, min(135.0, ang))

    def _on_data(self, l1, a1, l2, a2, l3, a3, sensor):
        # Guardar últimas
        self.last_l = [l1, l2, l3]
        self.last_a = [a1, a2, a3]
        self.sensor_on = bool(sensor)

        # Calcular calibrados
        ac1 = self._apply_cal(0, l1, a1)
        ac2 = self._apply_cal(1, l2, a2)
        ac3 = self._apply_cal(2, l3, a3)

        # Actualizar labels crudos
        if hasattr(self.ui, 'analogoDtPOT1'): self.ui.analogoDtPOT1.setText(str(l1))
        if hasattr(self.ui, 'analogoDtPOT2'): self.ui.analogoDtPOT2.setText(str(l2))
        if hasattr(self.ui, 'analogoDtPOT3'): self.ui.analogoDtPOT3.setText(str(l3))
        if hasattr(self.ui, 'anguloDtPOT1'): self.ui.anguloDtPOT1.setText(f"{a1}°")
        if hasattr(self.ui, 'anguloDtPOT2'): self.ui.anguloDtPOT2.setText(f"{a2}°")
        if hasattr(self.ui, 'anguloDtPOT3'): self.ui.anguloDtPOT3.setText(f"{a3}°")

        # Mostrar calibrados en calBox (aunque los labels se llamen 'analogoDtCalibradoPOTx', usaremos ángulo calibrado)
        if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
            self.ui.calBox.setVisible(any(self.cal[i]['ok'] for i in range(3)))
            if hasattr(self.ui, 'analogoDtCalibradoPOT1'): self.ui.analogoDtCalibradoPOT1.setText(f"{ac1:.1f}°")
            if hasattr(self.ui, 'analogoDtCalibradoPOT2'): self.ui.analogoDtCalibradoPOT2.setText(f"{ac2:.1f}°")
            if hasattr(self.ui, 'analogoDtCalibradoPOT3'): self.ui.analogoDtCalibradoPOT3.setText(f"{ac3:.1f}°")

        # Estado del sensor capacitivo
        if hasattr(self.ui, 'EstadoDeSensorCapatitivo_ON_OFF'):
            self.ui.EstadoDeSensorCapatitivo_ON_OFF.setText("ON" if self.sensor_on else "OFF")

        # Agregar datos a las series para la gráfica (ventana deslizante)
        t = self.time_idx[-1] + 0.1 if self.time_idx else 0.0
        self.time_idx.append(t)
        for i, v in enumerate((l1, l2, l3)):
            self.ad_series[i].append(v)
            if len(self.ad_series[i]) > 100:
                self.ad_series[i].pop(0)
        for i, v in enumerate((ac1, ac2, ac3)):
            self.ang_series[i].append(v)
            if len(self.ang_series[i]) > 100:
                self.ang_series[i].pop(0)
        # Guardar ángulos originales
        for i, v in enumerate((a1, a2, a3)):
            self.raw_ang_series[i].append(v)
            if len(self.raw_ang_series[i]) > 100:
                self.raw_ang_series[i].pop(0)
        if len(self.time_idx) > 100:
            self.time_idx.pop(0)

        # Actualizar líneas
        for i in range(3):
            self.lines_ad[i].set_data(self.time_idx, self.ad_series[i])
            self.lines_ang[i].set_data(self.time_idx, self.ang_series[i])
        # Ajuste X
        if self.time_idx:
            xmin, xmax = self.time_idx[0], self.time_idx[-1]
            if xmax - xmin < 1:
                xmax = xmin + 1
            self.ax1.set_xlim(xmin, xmax)
        self.canvas.draw()

    # ---- Calibración ----
    def calibrar(self):
        if not self.monitoreo_activo or any(v is None for v in self.last_l):
            QMessageBox.information(self, "Calibración", "Inicia el monitoreo y asegúrate que llegan datos de los 3 potenciómetros.")
            return
        # Número de puntos
        num_points, ok = QInputDialog.getInt(self, "Calibración", "¿Cuántos puntos por potenciómetro? (2-10)", 3, 2, 10)
        if not ok:
            return

        # Calibrar cada canal secuencialmente
        for idx in range(3):
            proceed = QMessageBox.question(self, "Calibración", f"¿Calibrar POT{idx+1} ahora?", QMessageBox.Yes | QMessageBox.No)
            if proceed != QMessageBox.Yes:
                continue
            puntos = []
            for i in range(1, num_points + 1):
                # Pedir ángulo de referencia actual
                angle, ok2 = QInputDialog.getDouble(self, f"POT{idx+1} - Punto {i}/{num_points}",
                                                    "Fija el potenciómetro en el ángulo deseado y escribe el valor (grados):",
                                                    0.0, -135.0, 135.0, 1)
                if not ok2:
                    break
                lectura = int(self.last_l[idx] if self.last_l[idx] is not None else 0)
                puntos.append((lectura, float(angle)))
            if len(puntos) < 2:
                QMessageBox.warning(self, "Calibración", f"POT{idx+1}: se requieren al menos 2 puntos.")
                continue
            # Regresión lineal simple
            xs = [p[0] for p in puntos]
            ys = [p[1] for p in puntos]
            n = len(xs)
            mean_x = sum(xs) / n
            mean_y = sum(ys) / n
            num = sum((xs[i] - mean_x) * (ys[i] - mean_y) for i in range(n))
            den = sum((xs[i] - mean_x) ** 2 for i in range(n))
            if den == 0:
                QMessageBox.warning(self, "Calibración", f"POT{idx+1}: varianza cero en lecturas.")
                continue
            m = num / den
            b = mean_y - m * mean_x
            self.cal[idx]['m'] = float(m)
            self.cal[idx]['b'] = float(b)
            self.cal[idx]['ok'] = True
            self.cal[idx]['points'] = puntos
        # Guardar y reflejar UI
        self._save_calibration()
        if hasattr(self.ui, 'calibrarBt'):
            self.ui.calibrarBt.setText("Calibrado")
        if hasattr(self.ui, 'calBox') and self.ui.calBox is not None:
            self.ui.calBox.setVisible(any(self.cal[i]['ok'] for i in range(3)))

    # ---- Utilidad ----
    def limpiar(self):
        self.time_idx.clear()
        for i in range(3):
            self.ad_series[i].clear()
            self.ang_series[i].clear()
        for ln in self.lines_ad + self.lines_ang:
            ln.set_data([], [])
        # Reset labels
        if hasattr(self.ui, 'analogoDtPOT1'): self.ui.analogoDtPOT1.setText("--")
        if hasattr(self.ui, 'analogoDtPOT2'): self.ui.analogoDtPOT2.setText("--")
        if hasattr(self.ui, 'analogoDtPOT3'): self.ui.analogoDtPOT3.setText("--")
        if hasattr(self.ui, 'anguloDtPOT1'): self.ui.anguloDtPOT1.setText("--")
        if hasattr(self.ui, 'anguloDtPOT2'): self.ui.anguloDtPOT2.setText("--")
        if hasattr(self.ui, 'anguloDtPOT3'): self.ui.anguloDtPOT3.setText("--")
        if hasattr(self.ui, 'analogoDtCalibradoPOT1'): self.ui.analogoDtCalibradoPOT1.setText("--")
        if hasattr(self.ui, 'analogoDtCalibradoPOT2'): self.ui.analogoDtCalibradoPOT2.setText("--")
        if hasattr(self.ui, 'analogoDtCalibradoPOT3'): self.ui.analogoDtCalibradoPOT3.setText("--")
        self.canvas.draw()

    def exportar(self):
        """
        Exporta los datos a Excel (.xlsx) siguiendo el mismo estilo del módulo Simple Angle.
        Incluye datos principales, metadatos, calibración por canal, y una hoja con la gráfica.
        Si faltan librerías, hace fallback a CSV.
        """
        try:
            if not self.time_idx:
                QMessageBox.information(self, "Exportar", "No hay datos para exportar.")
                return

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"AngleArm_{ts}.xlsx"
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exportar datos a Excel",
                default_name,
                "Archivos Excel (*.xlsx);;Archivos CSV (*.csv);;Todos los archivos (*.*)"
            )
            if not filename:
                return

            export_format = 'xlsx' if filename.lower().endswith('.xlsx') else 'csv'
            if export_format == 'xlsx':
                self._export_to_excel(filename)
            else:
                self._export_to_csv(filename)

            QMessageBox.information(self, "Exportar", f"Datos exportados exitosamente a:\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Exportar", f"Error al exportar: {e}")

    def _export_to_excel(self, filename: str):
        """
        Exporta a Excel con múltiples hojas: Datos, Metadatos, Puntos Calibración y Gráfica.
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws_data = wb.active
            ws_data.title = "Datos del Sensor"

            # Preparar datos principales
            data = {
                'Tiempo (s)': self.time_idx.copy(),
                'AD1': self.ad_series[0].copy(),
                'AD2': self.ad_series[1].copy(),
                'AD3': self.ad_series[2].copy(),
                'Ángulo Calibrado 1 (°)': self.ang_series[0].copy(),
                'Ángulo Calibrado 2 (°)': self.ang_series[1].copy(),
                'Ángulo Calibrado 3 (°)': self.ang_series[2].copy(),
            }
            # Incluir ángulos originales si hay
            if any(len(s) for s in self.raw_ang_series):
                data['Ángulo Original 1 (°)'] = self.raw_ang_series[0].copy()
                data['Ángulo Original 2 (°)'] = self.raw_ang_series[1].copy()
                data['Ángulo Original 3 (°)'] = self.raw_ang_series[2].copy()

            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws_data[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Autoajustar columnas
            for column in ws_data.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws_data.column_dimensions[col_letter].width = min(max_length + 2, 18)

            # Hoja de Metadatos
            ws_meta = wb.create_sheet("Metadatos")
            intervalo = 0.1
            metadata = [
                ["INFORMACIÓN DEL EXPERIMENTO", ""],
                ["Fecha de exportación", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["Sensor", "Angle Arm (3 potenciómetros + capacitivo)"],
                ["Muestras totales", len(self.time_idx)],
                ["Intervalo de muestreo", f"{intervalo} segundos"],
                ["Tiempo total", f"{len(self.time_idx) * intervalo:.1f} segundos"],
                ["", ""],
                ["RANGO DE DATOS", ""],
                ["AD1 min", min(self.ad_series[0]) if self.ad_series[0] else 0],
                ["AD1 max", max(self.ad_series[0]) if self.ad_series[0] else 0],
                ["AD2 min", min(self.ad_series[1]) if self.ad_series[1] else 0],
                ["AD2 max", max(self.ad_series[1]) if self.ad_series[1] else 0],
                ["AD3 min", min(self.ad_series[2]) if self.ad_series[2] else 0],
                ["AD3 max", max(self.ad_series[2]) if self.ad_series[2] else 0],
            ]
            # Rango de ángulos calibrados
            for i in range(3):
                if self.ang_series[i]:
                    metadata.append([f"Ángulo Calibrado {i+1} min (°)", f"{min(self.ang_series[i]):.2f}"])
                    metadata.append([f"Ángulo Calibrado {i+1} max (°)", f"{max(self.ang_series[i]):.2f}"])
            # Calibración por canal
            for i in range(3):
                c = self.cal[i]
                metadata.extend([
                    ["", ""],
                    [f"CALIBRACIÓN CANAL {i+1}", ""],
                    ["Estado", "Calibrado" if c['ok'] else "No calibrado"],
                    ["Ecuación", f"y = {c['m']:.6f}x + {c['b']:.3f}"] if c['ok'] else ["Ecuación", "--"],
                    ["Puntos", len(c['points']) if c['points'] else 0],
                ])
            for row in metadata:
                ws_meta.append(row)

            # Hoja Puntos Calibración (si existen)
            any_points = any(self.cal[i]['points'] for i in range(3))
            if any_points:
                ws_cal = wb.create_sheet("Puntos Calibración")
                ws_cal.append(["Canal", "Punto", "Lectura AD", "Ángulo Ref (°)", "Ángulo Calc (°)", "Error (°)"])
                for ch in range(3):
                    pts = self.cal[ch]['points']
                    for j, (lect, ang_ref) in enumerate(pts, 1):
                        if self.cal[ch]['ok']:
                            ang_calc = self.cal[ch]['m'] * lect + self.cal[ch]['b']
                            err = ang_calc - ang_ref
                        else:
                            ang_calc = 0.0
                            err = 0.0
                        ws_cal.append([ch + 1, j, lect, ang_ref, round(ang_calc, 3), round(err, 3)])

            # Hoja de Gráfica
            ws_graph = wb.create_sheet("Gráfica")
            graph_image = self._create_export_graph()
            if graph_image:
                from openpyxl.drawing.image import Image
                img = Image(graph_image)
                img.width = 1000
                img.height = 700
                ws_graph['A1'] = "BRAZO ROBÓTICO - DATOS EXPORTADOS"
                ws_graph['A1'].font = Font(bold=True, size=16, color="2F5597")
                ws_graph['A1'].alignment = Alignment(horizontal="center")
                ws_graph['A3'] = f"Generada: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws_graph['A4'] = f"Muestras mostradas: {len(self.time_idx)}"
                # Mostrar ecuaciones por canal
                for ch in range(3):
                    c = self.cal[ch]
                    if c['ok']:
                        ws_graph[f'A{6+ch}'] = f"Canal {ch+1}: y = {c['m']:.6f}x + {c['b']:.3f}"
                ws_graph.add_image(img, 'A8')
                ws_graph.column_dimensions['A'].width = 60

            wb.save(filename)
            if graph_image:
                try:
                    os.unlink(graph_image)
                except Exception:
                    pass
            print(f"Datos exportados a Excel: {filename}")
        except ImportError:
            # Fallback a CSV si faltan librerías
            QMessageBox.warning(self, "Exportar", "Librerías de Excel no disponibles. Exportando a CSV...")
            self._export_to_csv(filename.replace('.xlsx', '.csv'))
        except Exception as e:
            raise Exception(f"Error al crear archivo Excel: {e}")

    def _export_to_csv(self, filename: str):
        """Exportación básica a CSV como alternativa."""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                f.write(f"# AngleArm exportado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                f.write("# t,AD1,AD2,AD3,AngCal1,AngCal2,AngCal3,AngOrig1,AngOrig2,AngOrig3\n")
                n = len(self.time_idx)
                for i in range(n):
                    t = self.time_idx[i]
                    ad1 = self.ad_series[0][i] if i < len(self.ad_series[0]) else ''
                    ad2 = self.ad_series[1][i] if i < len(self.ad_series[1]) else ''
                    ad3 = self.ad_series[2][i] if i < len(self.ad_series[2]) else ''
                    ang1 = self.ang_series[0][i] if i < len(self.ang_series[0]) else ''
                    ang2 = self.ang_series[1][i] if i < len(self.ang_series[1]) else ''
                    ang3 = self.ang_series[2][i] if i < len(self.ang_series[2]) else ''
                    a1 = self.raw_ang_series[0][i] if i < len(self.raw_ang_series[0]) else ''
                    a2 = self.raw_ang_series[1][i] if i < len(self.raw_ang_series[1]) else ''
                    a3 = self.raw_ang_series[2][i] if i < len(self.raw_ang_series[2]) else ''
                    f.write(f"{t:.2f},{ad1},{ad2},{ad3},{ang1},{ang2},{ang3},{a1},{a2},{a3}\n")
            print(f"Datos exportados a CSV: {filename}")
        except Exception as e:
            raise Exception(f"Error al crear archivo CSV: {e}")

    def _create_export_graph(self):
        """
        Crea imagen PNG de alta calidad con 3 series AD y 3 series de ángulo calibrado.
        Retorna la ruta del archivo temporal o None.
        """
        try:
            from matplotlib.figure import Figure as MplFigure
            fig = MplFigure(figsize=(12, 8), dpi=150, facecolor='white')
            ax1 = fig.add_subplot(111)
            ax2 = ax1.twinx()

            ax1.set_xlabel("Tiempo (segundos)", fontsize=14, fontweight='bold')
            ax1.set_ylabel("Lectura Analógica (AD)", color='blue', fontsize=14, fontweight='bold')
            ax1.tick_params(axis='y', labelcolor='blue', labelsize=12)
            ax1.tick_params(axis='x', labelsize=12)
            ax1.set_ylim(0, 4095)
            ax1.grid(True, alpha=0.3, linestyle='--')

            ax2.set_ylabel("Ángulo Calibrado (°)", color='red', fontsize=14, fontweight='bold')
            ax2.tick_params(axis='y', labelcolor='red', labelsize=12)
            ax2.set_ylim(-135, 135)

            colors = ['#1f77b4', '#ff7f0e', '#2ca02c']
            labels_ad = ['AD1', 'AD2', 'AD3']
            labels_ang = ['AngCal1', 'AngCal2', 'AngCal3']

            if self.time_idx:
                for i in range(3):
                    ax1.plot(self.time_idx, self.ad_series[i], color=colors[i], linestyle='-', linewidth=2.0, alpha=0.8, label=labels_ad[i])
                    ax2.plot(self.time_idx, self.ang_series[i], color=colors[i], linestyle='--', linewidth=2.0, alpha=0.9, label=labels_ang[i])

                if len(self.time_idx) > 1:
                    x_min = min(self.time_idx)
                    x_max = max(self.time_idx)
                    margen = (x_max - x_min) * 0.05 if x_max > x_min else 1
                    ax1.set_xlim(x_min - margen, x_max + margen)
                else:
                    ax1.set_xlim(0, 1)

            # Leyenda combinada
            l1, lb1 = ax1.get_legend_handles_labels()
            l2, lb2 = ax2.get_legend_handles_labels()
            ax1.legend(l1 + l2, lb1 + lb2, loc='upper left', fontsize=12, framealpha=0.9)

            # Título con ecuaciones por canal si hay
            title = "Brazo Robótico - Datos Exportados"
            eqs = []
            for ch in range(3):
                c = self.cal[ch]
                if c['ok']:
                    eqs.append(f"C{ch+1}: y={c['m']:.4f}x+{c['b']:.1f}")
            if eqs:
                title += "\n" + ", ".join(eqs)
            fig.suptitle(title, fontsize=16, fontweight='bold', y=0.95)

            # Información adicional
            info_text = f"Muestras: {len(self.time_idx)} | Intervalo: 0.1s | Tiempo total: {len(self.time_idx) * 0.1:.1f}s"
            fig.text(0.02, 0.02, info_text, fontsize=10, style='italic', alpha=0.7)

            fig.tight_layout()

            import tempfile
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
            img_buffer.seek(0)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            temp_file.write(img_buffer.getvalue())
            temp_file.close()
            img_buffer.close()
            fig.clear()
            return temp_file.name
        except Exception as e:
            print(f"Error creando gráfica de exportación: {e}")
            return None

    # ---- Persistencia ----
    def _cal_path(self):
        base = os.path.dirname(__file__)
        return os.path.join(base, 'angleArm_calibration.json')

    def _save_calibration(self):
        payload = {
            'channels': {
                str(i): {
                    'm': self.cal[i]['m'],
                    'b': self.cal[i]['b'],
                    'ok': self.cal[i]['ok'],
                    'points': self.cal[i]['points'],
                } for i in range(3)
            }
        }
        try:
            with open(self._cal_path(), 'w', encoding='utf-8') as f:
                json.dump(payload, f, indent=2)
            print("Calibración AngleArm guardada")
        except Exception as e:
            print(f"No se pudo guardar calibración AngleArm: {e}")

    def _load_calibration(self):
        path = self._cal_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
            chans = payload.get('channels', {})
            for i in range(3):
                c = chans.get(str(i), {})
                self.cal[i]['m'] = float(c.get('m', 1.0))
                self.cal[i]['b'] = float(c.get('b', 0.0))
                self.cal[i]['ok'] = bool(c.get('ok', False))
                self.cal[i]['points'] = c.get('points', [])
            print("Calibración AngleArm cargada")
        except Exception as e:
            print(f"No se pudo cargar calibración AngleArm: {e}")

    # ---- Aux ----
    def _get_ip(self):
        # Buscar el QLineEdit ipEdit en la ventana principal
        ip_text = None
        try:
            from PySide6.QtWidgets import QLineEdit
            if self.main_window:
                ip_widget = self.main_window.findChild(QLineEdit, "ipEdit")
                if ip_widget:
                    ip_text = ip_widget.text().strip()
        except Exception:
            pass
        return ip_text or "192.168.1.100"
